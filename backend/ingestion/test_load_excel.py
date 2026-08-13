"""Self-check for the ingestion pipeline. Run with: pytest backend/ingestion/test_load_excel.py"""
import os
import shutil
import sys
import tempfile
from datetime import date

import openpyxl
import pytest

from backend.ingestion.column_mapping import parse_assigned_week_order, to_number


def _load_against_fresh_db():
    """backend.db.session reads its DB path once at import time, so each
    test that needs its own throwaway DB must force a re-import after
    pointing PAYMENTS_DB_PATH at a fresh temp file."""
    os.environ["PAYMENTS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH, load

    # Vercel-migration task: load()'s own default is now the DB-backed
    # "current" blob, not this local checked-out file — this helper
    # genuinely wants to bootstrap a fresh DB straight from the real
    # sample file, so it passes the path explicitly (same escape hatch
    # every other test in this codebase already uses).
    return load(excel_path=EXCEL_PATH)


def test_parse_assigned_week_order():
    assert parse_assigned_week_order("W2-P3") == (2, 3)
    assert parse_assigned_week_order(0) == (None, None)
    assert parse_assigned_week_order("0") == (None, None)
    assert parse_assigned_week_order(None) == (None, None)


def test_to_number():
    assert to_number(None) == 0.0
    assert to_number(" ") == 0.0
    assert to_number(5000) == 5000.0
    assert to_number("5000") == 5000.0


def test_load_duplicate_erp_code_same_entity_first_row_wins_and_warns_loudly(tmp_path):
    """Confirmed real bug (2026-07-31 investigation), narrowed by the
    entity-code-collision fix: a SAME-ENTITY ERP code collision (e.g. the
    real sheet's ARS "INC" pair — two different vendors typo'd onto one
    code) is a genuine data-entry error, not two entities legitimately
    sharing a bare code (see the different-entity test below for that
    case). The per-(entity, erp_code) upsert must still not let the later
    row silently overwrite the earlier one. First-row-wins (Sarath's later
    call, superseding the original last-row-wins fix): the later row is
    skipped entirely, never even read — loud either way, via a note naming
    the code, entity, every row, and the winner.
    """
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "dup.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    from backend.ingestion.column_mapping import build_sheet_map

    sheet_map = build_sheet_map(ws)
    # Fixture's two real rows: VT0001/"Alpha Freight Services" (row 2),
    # VT0002/"Bravo Industrial Supplies" (row 3), both entity "ARS" — give
    # row 3 row 2's own erp_code (same entity on both), matching the real
    # sheet's ARS "INC" collision shape exactly.
    ws.cell(row=3, column=sheet_map.erp_code_col, value="VT0001")
    wb.save(path)

    os.environ["PAYMENTS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import load

    report = load(excel_path=path)

    assert len(report["duplicate_erp_code_notes"]) == 1
    note = report["duplicate_erp_code_notes"][0]
    assert "VT0001" in note and "ARS" in note and "Alpha Freight Services" in note  # names code, entity, winner
    assert "distinct ERP code" in note  # explicit, not left looking resolved
    assert note in report["data_quality_notes"]  # also surfaced through the existing notes channel

    session = SessionLocal()
    try:
        assert session.query(Vendor).count() == 1  # collapsed to one row, upload still succeeds
        vendor = session.query(Vendor).filter_by(erp_code="VT0001").one()
        assert vendor.vendor_name == "Alpha Freight Services"  # first row (row 2) won
    finally:
        session.close()


def test_load_duplicate_erp_code_different_entities_both_kept(tmp_path):
    """Entity-code-collision fix, the actual bug: 28 real pairs in the
    master sheet are two DIFFERENT legal entities (ARS/FP) sharing one bare
    numeric code, disambiguated by the sheet's own "Unique" column — this
    is not a duplicate at all, and both vendors must be ingested as
    separate, live rows (Vendor's real uniqueness key is (entity,
    erp_code), models.py), not silently collapsed to one."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "dup_entities.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    from backend.ingestion.column_mapping import build_sheet_map

    sheet_map = build_sheet_map(ws)
    # Same bare erp_code as row 2, but a DIFFERENT entity — this is the
    # ARS/FP "V00149" shape, not a real duplicate.
    ws.cell(row=3, column=sheet_map.erp_code_col, value="VT0001")
    ws.cell(row=3, column=sheet_map.entity_col, value="FP")
    wb.save(path)

    os.environ["PAYMENTS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import load

    report = load(excel_path=path)

    assert report["duplicate_erp_code_notes"] == []  # not flagged — different entities, not a real duplicate

    session = SessionLocal()
    try:
        vendors = session.query(Vendor).filter_by(erp_code="VT0001").order_by(Vendor.entity).all()
        assert len(vendors) == 2  # both entities' vendors kept, not collapsed
        assert {v.entity for v in vendors} == {"ARS", "FP"}
        assert {v.vendor_name for v in vendors} == {"Alpha Freight Services", "Bravo Industrial Supplies"}
    finally:
        session.close()


def test_load_negative_payment_cell_treated_as_overpayment(tmp_path):
    """Confirmed with Finance: a negative Payment cell means an over-payment
    (more than what was billed) recorded with the wrong sign — it must be
    applied as a real payment of that magnitude (balance goes DOWN), not
    added back on as extra debt (the old, backwards behavior)."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "negpay.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    from backend.ingestion.column_mapping import build_sheet_map

    sheet_map = build_sheet_map(ws)
    first_month, payment_col = sheet_map.payment_cols[0]
    ws.cell(row=2, column=payment_col, value=-332000)
    wb.save(path)

    os.environ["PAYMENTS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import load

    report = load(excel_path=path)
    assert any("negative payment cell" in note for note in report["data_quality_notes"])

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code=ws.cell(row=2, column=sheet_map.erp_code_col).value).one()
        row = (
            session.query(MonthlyLedger)
            .filter_by(vendor_id=vendor.id, month=first_month)
            .one()
        )
        assert row.payment == 332000  # stored corrected/positive, never the raw negative
        assert row.closing_balance == row.opening_balance + row.payable - 332000  # balance went DOWN
    finally:
        session.close()


def test_load_real_sheet_reconciles():
    report = _load_against_fresh_db()
    assert report["vendor_count"] == 83
    assert report["ledger_row_count"] == 83 * 14
    # NOT asserting ledger_mismatches == [] here: the real dummy sheet's own
    # Closing Balance/Total Payable/Total Payment cells are formulas whose
    # cached values are currently blank (openpyxl's write-back drops them —
    # see the investigation; accepted as lost for this dummy file). That
    # check is still run and still logged as a sanity signal, but it no
    # longer feeds opening_balance (see the test below) — its count is a
    # property of the sheet's current, external, mutable formula-cache
    # state, not something ingestion's own correctness should be judged
    # against.
    #
    # Confirmed known data-entry slip in the source sheet (see docs/07) —
    # one vendor's W1-W5 amounts don't sum into that row's Total column.
    assert len(report["week_total_mismatches"]) == 1


def test_opening_balance_is_always_the_recomputed_figure_not_the_sheets_cached_formula():
    """Fix 1's confirmed bug: vendor.opening_balance was read from the
    sheet's own Closing Balance cell (data_only=True) — a formula whose
    cached value openpyxl's write-back can silently drop workbook-wide.
    Today it's blank for every one of the 83 real vendors (0.0), yet every
    vendor's opening_balance below is non-trivial and matches the ledger's
    own independently-recomputed closing figure — proving the DB value now
    comes from the tranche walk, not from a sheet figure that happens to
    already agree (it doesn't, right now).
    """
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal

    _load_against_fresh_db()
    session = SessionLocal()
    vendors = session.query(Vendor).all()
    assert len(vendors) == 83
    for vendor in vendors:
        last_ledger_row = (
            session.query(MonthlyLedger)
            .filter_by(vendor_id=vendor.id)
            .order_by(MonthlyLedger.month)
            .all()[-1]
        )
        assert float(vendor.opening_balance) == pytest.approx(float(last_ledger_row.closing_balance), abs=1)
    assert any(float(v.opening_balance) > 1000 for v in vendors)  # not a coincidental all-zero match
    session.close()


def test_assigned_week_survives_reingestion_when_sheet_gives_nothing():
    """Fix 2's confirmed permanent bug: the legacy assigned-week column is a
    VLOOKUP into an external workbook this repo doesn't have, so it reads as
    blank on every load — not just the next one, forever. Without a
    fallback to the existing DB value, every future re-ingestion/rollover
    would silently wipe assigned_week for virtually every vendor.
    Constructed: give a vendor a real assigned_week directly in the DB
    (today's actual, already-correct state), re-ingest the same real sheet
    (whose legacy column is still blank), confirm it survives untouched —
    same shape as test_finance_edited_assigned_week_survives_reingestion,
    but for a vendor with nothing at all from either sheet source.
    """
    _load_against_fresh_db()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import EXCEL_PATH, load

    session = SessionLocal()
    vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
    vendor_id = vendor.id
    vendor.assigned_week = 2  # a real value the sheet itself can no longer confirm or provide
    session.commit()
    session.close()

    load(excel_path=EXCEL_PATH)  # re-ingestion of the same, still-corrupted real sheet, same DB

    session = SessionLocal()
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.assigned_week == 2  # untouched — not wiped to None
    session.close()


def test_finance_edited_assigned_week_survives_reingestion():
    """Fix 1's confirmed bug: load() never upserted by erp_code, so a second
    run crashed outright (UNIQUE constraint on vendors.erp_code) — and even
    setting that aside, assigned_week was read from the stale legacy packed
    column instead of the dedicated column vendor_edits.py writes Finance's
    edits to, so a re-ingestion would have silently clobbered the edit.
    Constructed: edit V00400 via vendor_edits.py (writes to a temp Excel
    copy + this fresh DB), re-ingest that same temp Excel, confirm the edit
    survived and no duplicate vendor row was created.
    """
    _load_against_fresh_db()

    for mod in ("backend.configuration.vendor_edits",):
        sys.modules.pop(mod, None)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import EXCEL_PATH, load

    session = SessionLocal()
    vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
    vendor_id, original_week = vendor.id, vendor.assigned_week
    session.close()

    excel_copy = os.path.join(tempfile.mkdtemp(), "copy.xlsx")
    shutil.copy(EXCEL_PATH, excel_copy)
    new_week = 4 if original_week != 4 else 3
    update_vendor_field(vendor_id, "assigned_week", new_week, excel_path=excel_copy)

    load(excel_path=excel_copy)  # re-ingestion, same DB

    session = SessionLocal()
    assert session.query(Vendor).filter_by(erp_code="V00400").count() == 1  # upserted, not duplicated
    reloaded = session.query(Vendor).filter_by(erp_code="V00400").one()
    assert reloaded.id == vendor_id
    assert reloaded.assigned_week == new_week  # survived — not reset to the legacy column's original value
    session.close()


def _add_unique_code_column(ws, sheet_map, values, header_text="Unique Code"):
    """Appends a brand-new column past the fixture's existing ones, writes
    header_text into the header row, then values (a {row: code} dict) into
    the data rows — mirrors how Finance's real, not-yet-existing column
    would show up on re-ingestion."""
    col = ws.max_column + 1
    ws.cell(row=sheet_map.header_row, column=col, value=header_text)
    for row, code in values.items():
        ws.cell(row=row, column=col, value=code)
    return col


def _fresh_uc_db_and_load():
    os.environ["PAYMENTS_DB_PATH"] = os.path.join(tempfile.mkdtemp(), "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import load

    return load


def test_unique_code_basic_match(tmp_path):
    """Scenario 1: every row gets a distinct Unique Code — persisted
    verbatim onto the matching Vendor row."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "uc_basic.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    _add_unique_code_column(ws, sheet_map, {2: "UC-A", 3: "UC-B"})
    wb.save(path)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    report = load(excel_path=path)
    assert report["vendor_count"] == 2

    session = SessionLocal()
    try:
        assert session.query(Vendor).filter_by(erp_code="VT0001").one().unique_code == "UC-A"
        assert session.query(Vendor).filter_by(erp_code="VT0002").one().unique_code == "UC-B"
    finally:
        session.close()


def test_no_unique_code_column_behaves_exactly_as_before(tmp_path):
    """Scenario 2 — the important regression check: every sheet in this
    repo today has no Unique Code column at all. Must match/ingest exactly
    as before, unique_code stays None, no spurious notes."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "no_uc.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    report = load(excel_path=path)
    assert report["vendor_count"] == 2
    assert not any("Unique Code" in note for note in report["data_quality_notes"])

    session = SessionLocal()
    try:
        vendors = session.query(Vendor).all()
        assert len(vendors) == 2
        assert all(v.unique_code is None for v in vendors)
        assert {v.erp_code for v in vendors} == {"VT0001", "VT0002"}
    finally:
        session.close()


def test_unique_code_backfills_on_second_upload_without_duplicating_vendors(tmp_path):
    """Scenario 3: first load has no Unique Code column (vendors created
    with unique_code=None, matched by (entity, erp_code) as always); a
    second load of the same vendors, now with the column populated, must
    backfill unique_code onto the SAME rows, not create duplicates."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path_v1 = str(tmp_path / "v1.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path_v1)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.column_mapping import build_sheet_map

    report1 = load(excel_path=path_v1)
    assert report1["vendor_count"] == 2

    path_v2 = str(tmp_path / "v2.xlsx")
    shutil.copy(path_v1, path_v2)
    wb = openpyxl.load_workbook(path_v2)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    _add_unique_code_column(ws, sheet_map, {2: "UC-A", 3: "UC-B"})
    wb.save(path_v2)

    report2 = load(excel_path=path_v2)
    assert report2["vendor_count"] == 2  # same 2 rows updated, not duplicated

    session = SessionLocal()
    try:
        assert session.query(Vendor).count() == 2
        assert session.query(Vendor).filter_by(erp_code="VT0001").one().unique_code == "UC-A"
        assert session.query(Vendor).filter_by(erp_code="VT0002").one().unique_code == "UC-B"
    finally:
        session.close()


def test_unique_code_match_survives_erp_code_change_but_soft_deactivates__REAL_BUG(tmp_path):
    """Scenario 4, as specified, expected the vendor's erp_code field to be
    updated to the sheet's new value once matched via Unique Code. That is
    NOT what the code does — confirmed by reading load_excel.py's update
    branch (~line 501-524): it reassigns entity/vendor_name/unique_code/
    category/etc. but never `vendor.erp_code`. Flagging this rather than
    asserting the spec's assumption, per this task's own instruction.

    Worse, independently confirmed by running this exact scenario: because
    vendor.erp_code is left stale while the end-of-load soft-deactivate
    check (~line 704-712) compares (entity, vendor.erp_code) against the
    SHEET's own (entity, erp_code) key set — built from the row's erp_code
    cell, not the matched vendor's stored one — the stale erp_code no
    longer appears in that set. The SAME load() call that just matched and
    "updated" this vendor via Unique Code immediately soft-deactivates it
    afterward. This is a real, functional bug in the Unique-Code-matching
    feature as implemented, not a test-authoring assumption — this test
    pins down the actual (buggy) behavior so a fix must touch this test
    deliberately, not silently regress past it unnoticed.
    """
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path_v1 = str(tmp_path / "v1.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path_v1)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.column_mapping import build_sheet_map

    load(excel_path=path_v1)  # v1: no Unique Code column yet

    path_v2 = str(tmp_path / "v2.xlsx")
    shutil.copy(path_v1, path_v2)
    wb = openpyxl.load_workbook(path_v2)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    _add_unique_code_column(ws, sheet_map, {2: "UC-A", 3: "UC-B"})
    wb.save(path_v2)
    load(excel_path=path_v2)  # v2: backfills unique_code onto both vendors

    path_v3 = str(tmp_path / "v3.xlsx")
    shutil.copy(path_v2, path_v3)
    wb = openpyxl.load_workbook(path_v3)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    ws.cell(row=2, column=sheet_map.erp_code_col, value="VT0001-NEW")  # erp_code changed, Unique Code kept
    wb.save(path_v3)
    report3 = load(excel_path=path_v3)
    assert report3["vendor_count"] == 2  # still processed as 2 rows

    session = SessionLocal()
    try:
        assert session.query(Vendor).count() == 2  # matched via unique_code — not duplicated, that part works
        vendor = session.query(Vendor).filter_by(unique_code="UC-A").one()
        assert vendor.erp_code == "VT0001"  # NOT "VT0001-NEW" — erp_code is never reassigned on match
        assert vendor.is_active is False  # REAL BUG: wrongly soft-deactivated by this same load() call
    finally:
        session.close()


def test_unique_code_exact_duplicate_flags_and_collapses_but_also_hits_the_same_bug(tmp_path):
    """Scenario 5: two rows, different erp_code/entity/vendor_name, same
    Unique Code. Confirmed by running this exact scenario: load() does NOT
    crash, a data_quality_notes entry names the code and both rows, and the
    two rows DO collapse into one Vendor (find_duplicate_unique_codes()'s
    own docstring says this is expected — Unique Code is now the matching
    key, so the second row's upsert finds the first row's just-created
    vendor).

    Also observed, not assumed: the surviving vendor ends up
    is_active=False — the SAME root cause as the erp_code-change test
    above (vendor.erp_code is never reassigned to either row's value after
    the merge, so it satisfies neither row's own (entity, erp_code) key,
    and the end-of-load presence check deactivates it). Flagging this
    combination to Sarath rather than deciding whether exact-duplicate
    Unique Codes should be handled more strictly — see this task's own
    request for that decision.
    """
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "dup_uc.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    ws.cell(row=3, column=sheet_map.entity_col, value="FP")  # different entity too, not just erp_code/name
    _add_unique_code_column(ws, sheet_map, {2: "DUP-1", 3: "DUP-1"})
    wb.save(path)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    report = load(excel_path=path)
    assert any("DUP-1" in note and "collapse" in note for note in report["data_quality_notes"])

    session = SessionLocal()
    try:
        assert session.query(Vendor).count() == 1  # collapsed, as the code's own docstring predicts
        vendor = session.query(Vendor).filter_by(unique_code="DUP-1").one()
        assert vendor.is_active is False  # observed — same underlying bug as the test above
    finally:
        session.close()


def test_unique_code_near_identical_but_not_exact_codes_produce_no_note(tmp_path):
    """Duplicate Unique Code detection is EXACT MATCH ONLY (Sarath's
    explicit decision, 2026-08-13). A fuzzy/difflib-similarity duplicate
    check used to also run here (find_fuzzy_duplicate_unique_codes(), now
    removed from column_mapping.py entirely) and was producing ~5,471
    near-useless warnings on the real production sheet — most vendor codes
    share a formulaic prefix (e.g. "ARSV00400" vs "ARSV00401") that scores
    as "similar" under any generic string-similarity ratio despite being
    completely distinct, real vendor identities.

    This locks in the fix: two rows with different-but-similar-looking
    Unique Codes ("VC-1042" vs "VC-10042" — a 0.933 difflib ratio, well
    above the old 0.85 threshold) must produce NO note at all now, and
    must NOT be merged (two distinct exact codes -> two separate vendors,
    same as any other non-matching pair)."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "similar_not_exact.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    _add_unique_code_column(ws, sheet_map, {2: "VC-1042", 3: "VC-10042"})
    wb.save(path)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    report = load(excel_path=path)
    assert not any("typo" in n or "similar" in n for n in report["data_quality_notes"])

    session = SessionLocal()
    try:
        assert session.query(Vendor).count() == 2  # different exact codes -> never merged
        assert {v.is_active for v in session.query(Vendor).all()} == {True}
    finally:
        session.close()


@pytest.mark.parametrize("header_text", ["Unique ID", "Vendor Unique Code"])
def test_unique_code_header_variants_recognized(tmp_path, header_text):
    """Scenario 7: "Unique Code" is the primary expected header, but "Unique
    ID"/"Vendor Unique Code" are accepted synonyms."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path = str(tmp_path / "variant.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path)

    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    _add_unique_code_column(ws, sheet_map, {2: "UID-A", 3: "UID-B"}, header_text=header_text)
    wb.save(path)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    load(excel_path=path)
    session = SessionLocal()
    try:
        assert session.query(Vendor).filter_by(erp_code="VT0001").one().unique_code == "UID-A"
        assert session.query(Vendor).filter_by(erp_code="VT0002").one().unique_code == "UID-B"
    finally:
        session.close()


def test_unique_code_blank_cell_falls_back_and_never_wipes_a_backfilled_value(tmp_path):
    """Scenario 8: combines with scenario 3's backfill — a vendor already
    backfilled with a real unique_code, then re-loaded with that row's
    Unique Code cell now blank (erp_code/entity unchanged, so the
    (entity, erp_code) fallback still matches correctly and the scenario
    4/5 bug above doesn't apply here) must keep its unique_code untouched,
    not wiped to None."""
    fixtures_dir = os.path.join(os.path.dirname(__file__), "test_fixtures")
    path_v1 = str(tmp_path / "v1.xlsx")
    shutil.copy(os.path.join(fixtures_dir, "single_header_row_sheet.xlsx"), path_v1)

    load = _fresh_uc_db_and_load()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.column_mapping import build_sheet_map

    load(excel_path=path_v1)

    path_v2 = str(tmp_path / "v2.xlsx")
    shutil.copy(path_v1, path_v2)
    wb = openpyxl.load_workbook(path_v2)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    uc_col = _add_unique_code_column(ws, sheet_map, {2: "UC-A", 3: "UC-B"})
    wb.save(path_v2)
    load(excel_path=path_v2)

    path_v3 = str(tmp_path / "v3.xlsx")
    shutil.copy(path_v2, path_v3)
    wb = openpyxl.load_workbook(path_v3)
    ws = wb.active
    ws.cell(row=2, column=uc_col, value=None)  # blank out row 2's Unique Code cell only
    wb.save(path_v3)
    report3 = load(excel_path=path_v3)
    assert report3["vendor_count"] == 2

    session = SessionLocal()
    try:
        v1 = session.query(Vendor).filter_by(erp_code="VT0001").one()
        v2 = session.query(Vendor).filter_by(erp_code="VT0002").one()
        assert v1.unique_code == "UC-A"  # unchanged, not wiped
        assert v2.unique_code == "UC-B"
        assert v1.is_active is True and v2.is_active is True
    finally:
        session.close()


def test_mislabeled_nov_column_resolves_by_position():
    """Regression test: the sheet's 8th payable column is headered "Nov'26"
    (a confirmed real typo) but is structurally Nov-25 — months are derived
    by position from the Apr-25 anchor, not parsed from the header text.
    """
    _load_against_fresh_db()
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
    nov_row = (
        session.query(MonthlyLedger)
        .filter_by(vendor_id=vendor.id, month=date(2025, 11, 1))
        .one()
    )
    assert float(nov_row.payable) == 6506736.0
    session.close()
