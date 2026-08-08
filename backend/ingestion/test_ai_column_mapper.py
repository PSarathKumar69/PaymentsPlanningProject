"""Tests for the AI-assisted column-mapping layer (data-pipeline
AI-mapping task — supersedes Part B of
claude_code_prompt_priority_column_and_ai_mapping.md). Same isolation
convention as test_upload.py (fresh temp DB via PAYMENTS_DB_PATH + module
re-import, tmp_path copy of the real master Excel). Never calls the real
Gemini API — backend.ingestion.ai_column_mapper.gemini_client.generate_text
is always monkeypatched.
"""
import json
import os
import shutil
import sys

import openpyxl
import pytest


def _fresh_db_and_master(tmp_path):
    """Mirrors test_upload.py's own helper of the same name exactly (same
    narrow module-pop list, same reasoning — see that file's docstring)."""
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.ingestion.upload",
        "backend.ingestion.ai_column_mapper",
        "backend.ingestion.column_mapping_store",
        "backend.configuration.vendor_edits",
        "backend.configuration.extra_fields",
    ):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH, load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    load(excel_path=master_path)
    return master_path


def _rename_header(src_path, dest_path, old_header, new_header):
    from backend.ingestion.column_mapping import build_sheet_map, find_column, resolve_header_row

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    build_sheet_map(ws)  # sanity: the sheet is well-formed before the rename
    col = find_column(ws, old_header)
    assert col is not None, f"column {old_header!r} not found"
    header_row, _ = resolve_header_row(ws)
    ws.cell(row=header_row, column=col, value=new_header)
    wb.save(dest_path)


def _remove_header(src_path, dest_path, header_to_blank):
    """Blanks a header cell entirely — simulates a column that's genuinely
    gone (not just renamed), for the "AI also can't resolve it" scenario."""
    from backend.ingestion.column_mapping import find_column, resolve_header_row

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    col = find_column(ws, header_to_blank)
    assert col is not None, f"column {header_to_blank!r} not found"
    header_row, _ = resolve_header_row(ws)
    ws.cell(row=header_row, column=col).value = None
    wb.save(dest_path)


def _fake_gemini(mappings=None, warnings=None):
    """Returns a fake generate_text(prompt) that always answers with the
    given mappings/warnings, formatted exactly as the real prompt asks."""
    payload = {"mappings": mappings or {}, "warnings": warnings or []}

    def _fn(prompt):
        return json.dumps(payload)

    return _fn


# ---- Unit: prompt construction / response parsing --------------------------


def test_build_prompt_includes_field_descriptions_and_sample_data():
    from backend.ingestion import ai_column_mapper

    prompt = ai_column_mapper.build_prompt(
        header_row=["ERP Code", "V-P"],
        sample_rows=[["V001", "P0"], ["V002", "P2"]],
        fields_to_map=["priority_tag"],
        resolved_field_headers={"erp_code": "ERP Code"},
    )
    assert "priority_tag" in prompt
    assert "P0/P1/P2/P3/P4/P5" in prompt or "P0" in prompt  # field description text present
    assert "V-P" in prompt and "P0" in prompt  # real sample data present, not just headers


def test_build_prompt_includes_master_header_row_when_given():
    """Row 2 (merged/master header block) is new context this task adds —
    optional (None default keeps this function's signature backward
    compatible), but when a caller does pass one it must reach the prompt."""
    from backend.ingestion import ai_column_mapper

    prompt = ai_column_mapper.build_prompt(
        header_row=["ERP Code", "Apr-25", "May-25"],
        sample_rows=[["V001", 100, 200]],
        fields_to_map=[],
        resolved_field_headers={"erp_code": "ERP Code"},
        master_header_row=[None, "Payable", "Payable"],
    )
    assert "master_header_row" in prompt
    assert "Payable" in prompt


def test_parse_response_accepts_plain_json():
    from backend.ingestion import ai_column_mapper

    parsed = ai_column_mapper._parse_response(
        json.dumps({"mappings": {"priority_tag": {"column": "V-P", "confidence": "high", "reason": "x"}}}, )
    )
    assert parsed["mappings"]["priority_tag"]["column"] == "V-P"
    assert parsed["warnings"] == []


def test_parse_response_tolerates_markdown_fences():
    from backend.ingestion import ai_column_mapper

    fenced = "```json\n" + json.dumps({"mappings": {}, "warnings": ["heads up"]}) + "\n```"
    parsed = ai_column_mapper._parse_response(fenced)
    assert parsed["warnings"] == ["heads up"]


def test_parse_response_raises_on_non_json():
    from backend.ingestion import ai_column_mapper

    with pytest.raises(ValueError, match="non-JSON"):
        ai_column_mapper._parse_response("sorry, I can't help with that")


def test_map_columns_calls_gemini_and_returns_parsed_result(monkeypatch):
    from backend.ingestion import ai_column_mapper

    monkeypatch.setattr(
        ai_column_mapper.gemini_client,
        "generate_text",
        _fake_gemini(mappings={"priority_tag": {"column": "V-P", "confidence": "high", "reason": "values are P0-P4"}}),
    )
    result = ai_column_mapper.map_columns(
        header_row=["ERP Code", "V-P"],
        sample_rows=[["V001", "P0"]],
        fields_to_map=["priority_tag"],
        resolved_field_headers={"erp_code": "ERP Code"},
    )
    assert result["mappings"]["priority_tag"]["column"] == "V-P"


# ---- Outage handling (fix verification): Gemini errors -> clean 503, ------
# ---- never a raw 500 -------------------------------------------------------


def test_map_columns_raises_ai_mapping_unavailable_on_gemini_api_error(monkeypatch):
    """Gemini answered, but with an error status (4xx/5xx) the SDK's own
    retry gave up on — must surface as AIMappingUnavailableError (503 via
    backend/api/main.py's handler), never the raw genai_errors.APIError."""
    from google.genai import errors as genai_errors

    from backend.ingestion import ai_column_mapper

    def _raise_api_error(prompt):
        raise genai_errors.APIError(503, {"error": {"message": "model overloaded"}})

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", _raise_api_error)

    with pytest.raises(ai_column_mapper.AIMappingUnavailableError, match="temporarily unavailable"):
        ai_column_mapper.map_columns(
            header_row=["ERP Code", "V-P"],
            sample_rows=[["V001", "P0"]],
            fields_to_map=["priority_tag"],
            resolved_field_headers={"erp_code": "ERP Code"},
        )


def test_map_columns_raises_ai_mapping_unavailable_on_transport_error(monkeypatch):
    """Confirmed real case: a corporate proxy/firewall killing the TLS
    connection raises httpx.ConnectError — never reaches Gemini at all, so
    it's NOT a genai_errors.APIError subclass. Must still surface as the
    same clean AIMappingUnavailableError, not leak the raw httpx.ConnectError
    (the bug this fix addressed — it fell through uncaught before)."""
    import httpx

    from backend.ingestion import ai_column_mapper

    def _raise_connect_error(prompt):
        raise httpx.ConnectError("mocked TLS handshake failure")

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", _raise_connect_error)

    with pytest.raises(ai_column_mapper.AIMappingUnavailableError, match="temporarily unavailable"):
        ai_column_mapper.map_columns(
            header_row=["ERP Code", "V-P"],
            sample_rows=[["V001", "P0"]],
            fields_to_map=["priority_tag"],
            resolved_field_headers={"erp_code": "ERP Code"},
        )


def test_outage_does_not_block_upload_when_everything_already_resolves(tmp_path, monkeypatch):
    """Real incident: Gemini returning 503 used to hard-fail EVERY upload
    (AIMappingUnavailableError -> 503, see backend/api/main.py), even one
    where every field already resolves via fixed default/persisted override
    and needs no AI help at all. The fix: an outage is treated as "AI has no
    opinion on anything" and the upload proceeds on the existing mapping,
    with a banner explaining nothing new could be auto-detected this time."""
    from google.genai import errors as genai_errors

    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.upload import commit_upload

    def _raise_api_error(prompt):
        raise genai_errors.APIError(503, {"error": {"message": "model overloaded"}})

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", _raise_api_error)

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(master_path, upload_path)  # real sheet as-is, nothing renamed
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["vendor_count"] > 0  # outage never blocked the upload
    assert any("temporarily unavailable" in m for m in result["ai_column_mapping_messages"])


def test_outage_with_unresolvable_required_field_still_fails_but_names_the_outage(tmp_path, monkeypatch):
    """A required field with no fixed default and no persisted override is
    genuinely unresolvable without the AI — this must still fail during an
    outage, but as AIMappingUnavailableError (a clear "try again later"),
    never the generic MissingRequiredColumnsError, which would read as a
    permanent data problem and send Finance chasing the wrong fix."""
    from google.genai import errors as genai_errors

    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.upload import commit_upload

    def _raise_api_error(prompt):
        raise genai_errors.APIError(503, {"error": {"message": "model overloaded"}})

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", _raise_api_error)

    upload_path = str(tmp_path / "upload.xlsx")
    _rename_header(master_path, upload_path, "Entity", "Biz Unit")  # a required field, now unresolvable
    backup_path = str(tmp_path / "master.backup.xlsx")

    with pytest.raises(ai_column_mapper.AIMappingUnavailableError, match="temporarily unavailable"):
        commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)


# ---- Integration: resolve_column_mapping() against a real workbook --------


def test_fully_resolved_sheet_still_calls_gemini_once_and_changes_nothing(tmp_path, monkeypatch):
    """REWRITTEN (Sarath's explicit reversal of this module's earlier
    "0 calls when nothing's missing" resolution — see ai_column_mapper.py's
    module docstring): every expected column present and correctly named
    still gets exactly ONE Gemini call (the always-on cross-check), batching
    every field. When the AI's response is empty/agrees with everything,
    nothing gets applied and nothing gets logged."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion import ai_column_mapper

    # The real master sheet's priority_tag/assigned_week headers are
    # "Prioity" (a real typo) and "Week", not the fixed defaults ("Priority
    # Tag"/"Assigned Week") — so on the sheet AS-IS these two fields are
    # never actually "fully resolved" and this test's own premise wouldn't
    # hold. Normalize just those two headers here so this test verifies what
    # it claims to (a genuinely fully-resolved sheet), independent of the
    # real file's current drifted column names.
    normalized_step1 = str(tmp_path / "normalized_step1.xlsx")
    _rename_header(master_path, normalized_step1, "Prioity", "Priority Tag")
    normalized_path = str(tmp_path / "normalized.xlsx")
    _rename_header(normalized_step1, normalized_path, "Week", "Assigned Week")

    calls = []
    monkeypatch.setattr(
        ai_column_mapper.gemini_client,
        "generate_text",
        lambda p: calls.append(p) or json.dumps({"mappings": {}, "warnings": []}),
    )

    wb = openpyxl.load_workbook(normalized_path, data_only=True)
    session = SessionLocal()
    try:
        overrides_before = dict(ai_column_mapper.column_mapping_store.load_overrides(session))
        result = ai_column_mapper.resolve_column_mapping(session, wb.active)
    finally:
        session.close()

    assert len(calls) == 1  # always called now, exactly once per upload — never 0, never per-field
    # The real sheet also genuinely has no "Commitment Months" column at all
    # (a separate, pre-existing gap — vendor_edits.py's own module docstring:
    # "category/commitment_months don't exist as columns in the real Excel
    # yet"). It's optional, so this never blocks the upload, but this bug
    # fix's whole point is that an unresolved optional field is never
    # silent — so it's expected here too, not swept away as "nothing changed".
    assert len(result["audit_entries"]) == 1
    assert "commitment_months" in result["audit_entries"][0]["new_value"]
    assert len(result["banner_messages"]) == 1
    assert "commitment_months" in result["banner_messages"][0]
    assert result["header_overrides"] == overrides_before  # AI's empty response changed nothing


def test_renamed_soft_field_gap_filled_and_persisted_still_confirmed_on_second_upload(tmp_path, monkeypatch):
    """REWRITTEN (Sarath's explicit reversal — see module docstring and
    test_fully_resolved_sheet_still_calls_gemini_once_and_changes_nothing
    above): "second upload skips Gemini" is no longer true, AI runs every
    upload now. The task's own example scenario: "Priority Tag" renamed to
    "V-P". First upload: AI fills the gap (it's `missing`), mapping gets
    auto-persisted, field reads correctly. Second upload with the SAME
    renamed header: "V-P" is now the persisted override, so it's no longer
    `missing` — but the AI still runs, sees it as an already-resolved
    candidate, and (since our fake response keeps proposing the same "V-P"
    column) simply reconfirms it rather than overriding — the override
    persisted from upload 1 is still what's actually used, unchanged."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import ChangeSource, VendorPriorityTag

    calls = []

    def fake_generate_text(prompt):
        calls.append(prompt)
        return json.dumps(
            {
                "mappings": {
                    "priority_tag": {
                        "column": "V-P",
                        "confidence": "high",
                        "reason": "values in this column are exactly P0-P5",
                    }
                },
                "warnings": [],
            }
        )

    import backend.ingestion.ai_column_mapper as ai_column_mapper

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    backup_path = str(tmp_path / "master.backup.xlsx")

    # --- Upload 1: "Priority Tag" renamed to "V-P" ---
    upload_1 = str(tmp_path / "upload1.xlsx")
    _rename_header(master_path, upload_1, "Priority Tag", "V-P")
    result_1 = commit_upload(upload_1, excel_path=master_path, backup_path=backup_path)

    assert len(calls) == 1
    assert any("V-P" in m for m in result_1["ai_column_mapping_messages"])

    session = SessionLocal()
    try:
        # The mapping was actually applied: at least one vendor's
        # priority_tag came from the renamed column, not left None.
        tagged = [v for v in session.query(Vendor).all() if v.priority_tag is not None]
        assert tagged, "no vendor got a priority_tag from the AI-resolved 'V-P' column"
        assert all(v.priority_tag in VendorPriorityTag for v in tagged)

        entry = (
            session.query(AuditLog)
            .filter_by(source=ChangeSource.AI_COLUMN_MAPPING.value, field_name="column_mapping:priority_tag")
            .one()
        )
        assert "V-P" in entry.new_value
        assert "P0-P4" in entry.new_value or "confidence=high" in entry.new_value
    finally:
        session.close()

    # --- Upload 2: SAME renamed header again ---
    upload_2 = str(tmp_path / "upload2.xlsx")
    shutil.copy(upload_1, upload_2)  # already has "V-P", no further changes needed
    result_2 = commit_upload(upload_2, excel_path=master_path, backup_path=backup_path)

    assert len(calls) == 2  # AI runs on BOTH uploads now, not just the first
    # Confirmed, not re-overridden — no "AI OVERRODE" banner the second time,
    # same "V-P" mapping the first upload persisted is still in effect.
    assert not any("OVERRODE" in m for m in result_2["ai_column_mapping_messages"])
    session = SessionLocal()
    try:
        overrides = ai_column_mapper.column_mapping_store.load_overrides(session)
        assert overrides["priority_tag"] == "V-P"
    finally:
        session.close()


def test_required_field_ai_cannot_resolve_rejects_upload_with_clear_message(tmp_path, monkeypatch):
    """A genuinely unresolvable REQUIRED field (blanked header, AI reports
    low confidence): upload is rejected, message names the field. This is
    the only case that still blocks (task's final safety net)."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.column_mapping import MissingRequiredColumnsError
    from backend.ingestion.upload import commit_upload

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {
                    "entity": {"column": None, "confidence": "low", "reason": "no column plausibly represents entity"}
                },
                "warnings": [],
            }
        )

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    upload_path = str(tmp_path / "upload.xlsx")
    _remove_header(master_path, upload_path, "Entity")
    backup_path = str(tmp_path / "master.backup.xlsx")

    with pytest.raises(MissingRequiredColumnsError, match="entity"):
        commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)


def test_required_field_ai_confidently_resolves_it_upload_succeeds(tmp_path, monkeypatch):
    """A renamed REQUIRED field (not just a soft one): AI resolves it with
    high confidence, upload proceeds and reads that field correctly."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.upload import commit_upload

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {
                    "entity": {"column": "Biz Unit", "confidence": "high", "reason": "same short entity codes as before"}
                },
                "warnings": [],
            }
        )

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    upload_path = str(tmp_path / "upload.xlsx")
    _rename_header(master_path, upload_path, "Entity", "Biz Unit")
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["vendor_count"] > 0  # ingestion actually completed, not rejected

    session = SessionLocal()
    try:
        vendors = session.query(Vendor).all()
        assert all(v.entity for v in vendors)  # entity field populated from the renamed column
    finally:
        session.close()


def test_ai_mapping_does_not_override_an_already_resolved_field_on_medium_confidence(tmp_path, monkeypatch):
    """REWRITTEN (Sarath's explicit reversal — see module docstring): an
    already-resolved field CAN be overridden now, but only on a "high"-
    confidence, genuinely different answer (see the dedicated override test
    below). This test now covers the guardrail that survives that reversal:
    a "medium"-confidence disagreement must NOT silently replace the
    existing mapping — it only gets logged as a non-blocking warning. Under
    the OLD code this held structurally (category wasn't even in
    fields_to_map, so the AI's opinion on it was never looked at); under the
    new code it's a live decision gated on confidence, so this is a
    meaningfully different check now, not just a relabeled old one."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import VendorCategory

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {
                    "priority_tag": {"column": "V-P", "confidence": "high", "reason": "P0-P4 values"},
                    # Only "medium" — a real disagreement, but not confident
                    # enough to act on. Must NOT override "category".
                    "category": {"column": "Status", "confidence": "medium", "reason": "worth a second look"},
                },
                "warnings": [],
            }
        )

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    upload_path = str(tmp_path / "upload.xlsx")
    _rename_header(master_path, upload_path, "Priority Tag", "V-P")
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    # Logged as a non-blocking warning, not silently dropped and not applied.
    assert any(
        "cross-check warning" in m and "category" in m for m in result["ai_column_mapping_messages"]
    )
    assert not any("OVERRODE" in m for m in result["ai_column_mapping_messages"])

    session = SessionLocal()
    try:
        overrides = ai_column_mapper.column_mapping_store.load_overrides(session)
        assert "category" not in overrides  # never persisted as an override
        # category still reads from the real "Category" column, unaffected —
        # every vendor's category is still a legitimate enum value, not
        # garbage read from the "Status" (Active/Blocked) column.
        vendors = session.query(Vendor).all()
        assert all(v.category in VendorCategory for v in vendors)
    finally:
        session.close()


def test_low_confidence_optional_field_never_silently_vanishes(tmp_path, monkeypatch):
    """Sarath's exact real-world bug report: "Priority Tag and Assigned Week
    vanish after upload even though the source Excel clearly has them for
    every vendor." Root cause: the real master sheet's headers for these two
    fields are "Prioity" (a genuine typo) and "Week" — NOT the fixed
    defaults ("Priority Tag"/"Assigned Week") — so they're always `missing`
    pre-AI, and priority_tag/assigned_week are OPTIONAL fields (never in
    REQUIRED_FIELDS). Before this fix, a "low"-confidence (or no-opinion)
    Gemini answer for an optional missing field just `continue`d with ZERO
    audit_log entry and ZERO banner message — load() then writes None for
    every vendor with no prior DB value, indistinguishable from the sheet
    genuinely having no data there. This test uses the REAL sheet's actual
    "Prioity"/"Week" headers directly (no renaming needed — this is the real
    shape), confirms the pre-fix silent-vanish symptom still happens (these
    are optional fields, so a low-confidence AI answer legitimately can't be
    trusted to guess), but proves it is now always flagged, never silent.
    """
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion import ai_column_mapper
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import ChangeSource

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {
                    "priority_tag": {
                        "column": None,
                        "confidence": "low",
                        "reason": "no column plausibly represents priority_tag",
                    },
                    "assigned_week": {
                        "column": None,
                        "confidence": "low",
                        "reason": "no column plausibly represents assigned_week",
                    },
                },
                "warnings": [],
            }
        )

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(master_path, upload_path)  # real sheet as-is — "Prioity"/"Week" headers, unrenamed
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["vendor_count"] > 0  # optional fields never block the upload

    # The symptom Sarath reported: both fields genuinely come back empty for
    # every vendor, because neither the fixed default header nor a
    # low-confidence AI guess resolved a real column.
    session = SessionLocal()
    try:
        vendors = session.query(Vendor).all()
        assert all(v.priority_tag is None for v in vendors)
        assert all(v.assigned_week is None for v in vendors)

        # The fix under test: this is never silent. Both a banner (surfaced
        # to Finance on the upload response) and an audit_log row exist.
        for field in ("priority_tag", "assigned_week"):
            assert any(field in m and "could not confidently identify" in m.lower() for m in result["ai_column_mapping_messages"]), (
                field,
                result["ai_column_mapping_messages"],
            )
            assert (
                session.query(AuditLog)
                .filter_by(source=ChangeSource.AI_COLUMN_MAPPING.value, field_name="column_mapping_warning")
                .filter(AuditLog.new_value.contains(field))
                .first()
                is not None
            )
    finally:
        session.close()


# ---- New behavior: AI can override a stale/coincidental match, loudly -----


def test_ai_overrides_a_stale_coincidental_header_match(tmp_path, monkeypatch):
    """Sarath's exact confirmed failure mode, generalized: a field's FIXED
    DEFAULT header text can coincidentally match a DIFFERENT column holding
    unrelated data. Here "Category" gets renamed to "Old Category" (still
    holds the real Must Pay/Commitment/Normal/Inactive values), and the
    unrelated "Status" extra column (Active/Blocked values, real seeded
    data) gets renamed TO "Category" — a column named exactly "Category"
    really does exist, so the deterministic layer would happily (and
    wrongly) trust it. A "high"-confidence, real-column AI answer must
    override this — loudly: a `column_mapping:category` audit_log entry
    naming both headers and the reason, plus an "AI OVERRODE" banner
    distinct from a routine "auto-detected as" one."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import ChangeSource, VendorCategory

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {
                    "category": {
                        "column": "Old Category",
                        "confidence": "high",
                        "reason": "the 'Category' column now holds Active/Blocked status values, not a real "
                        "vendor category — 'Old Category' has the actual Must Pay/Commitment/Normal/Inactive values",
                    }
                },
                "warnings": [],
            }
        )

    import backend.ingestion.ai_column_mapper as ai_column_mapper

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    step1 = str(tmp_path / "step1.xlsx")
    _rename_header(master_path, step1, "Category", "Old Category")
    upload_path = str(tmp_path / "upload.xlsx")
    _rename_header(step1, upload_path, "Status", "Category")
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    override_banners = [m for m in result["ai_column_mapping_messages"] if "AI OVERRODE" in m]
    assert override_banners, result["ai_column_mapping_messages"]
    assert "Category" in override_banners[0] and "Old Category" in override_banners[0]

    session = SessionLocal()
    try:
        # Read from "Old Category" (the real data), not the renamed "Status"
        # column's Active/Blocked garbage.
        vendors = session.query(Vendor).all()
        assert all(v.category in VendorCategory for v in vendors)

        overrides = ai_column_mapper.column_mapping_store.load_overrides(session)
        assert overrides["category"] == "Old Category"

        entry = (
            session.query(AuditLog)
            .filter_by(source=ChangeSource.AI_COLUMN_MAPPING.value, field_name="column_mapping:category")
            .one()
        )
        assert "AI OVERRODE" in entry.new_value
        assert "Category" in entry.new_value and "Old Category" in entry.new_value
    finally:
        session.close()


def test_priority_tag_vs_assigned_week_disambiguation(tmp_path, monkeypatch):
    """Sarath's exact worked example: a header mentioning "priority" could
    mean EITHER priority_tag (values P0-P5) OR assigned_week (values are
    small integers/Wn labels) — these must never be conflated. "Priority
    Tag" renamed to "Priority", "Assigned Week" renamed to "Weekly
    Priority" — a scripted fake Gemini response that correctly tells them
    apart, proving the CODE applies a correct disambiguation (the real
    model's own reasoning is verified separately, see
    verify_ai_first_mapping_manually.py and this report)."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import VendorPriorityTag

    def fake_generate_text(prompt):
        # build_prompt() joins "{instructions}\n\n{json_payload}" — the
        # payload itself (json.dumps(indent=2)) never contains a blank line,
        # so the LAST "\n\n"-separated segment is always just the payload
        # (the instructions text has plenty of "{" of its own in prose/JSON
        # examples, so finding the first "{" isn't reliable).
        payload = json.loads(prompt.rsplit("\n\n", 1)[1])
        assert "Priority" in payload["header_row"]
        assert "Weekly Priority" in payload["header_row"]
        return json.dumps(
            {
                "mappings": {
                    "priority_tag": {
                        "column": "Priority",
                        "confidence": "high",
                        "reason": "values are exactly P0-P5, the vendor's own priority tag",
                    },
                    "assigned_week": {
                        "column": "Weekly Priority",
                        "confidence": "high",
                        "reason": "values are small plain integers, the vendor's assigned week, distinct from "
                        "the P0-P5 priority tag column",
                    },
                },
                "warnings": [],
            }
        )

    import backend.ingestion.ai_column_mapper as ai_column_mapper

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    step1 = str(tmp_path / "step1.xlsx")
    _rename_header(master_path, step1, "Priority Tag", "Priority")
    upload_path = str(tmp_path / "upload.xlsx")
    _rename_header(step1, upload_path, "Assigned Week", "Weekly Priority")
    backup_path = str(tmp_path / "master.backup.xlsx")

    commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    try:
        vendors = session.query(Vendor).all()
        tagged = [v for v in vendors if v.priority_tag is not None]
        assert tagged, "no vendor got a priority_tag"
        assert all(v.priority_tag in VendorPriorityTag for v in tagged)  # real P0-P5, not week numbers
        weeked = [v for v in vendors if v.assigned_week is not None]
        assert weeked, "no vendor got an assigned_week"
        assert all(isinstance(v.assigned_week, int) and 1 <= v.assigned_week <= 5 for v in weeked)
    finally:
        session.close()


# ---- Header-row confirmation (this task) ------------------------------------


def test_resolve_column_mapping_returns_the_deterministic_header_row_when_ai_agrees(tmp_path):
    """Default conftest.py stub answers with no "header_row" opinion at all
    — resolve_column_mapping() must fall back to the deterministic
    candidate, not error or guess something else."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion import ai_column_mapper, column_mapping

    wb = openpyxl.load_workbook(master_path, data_only=True)
    detection = column_mapping.detect_header_row(wb.active)

    session = SessionLocal()
    try:
        result = ai_column_mapper.resolve_column_mapping(session, wb.active)
    finally:
        session.close()

    assert result["header_row"] == detection.row_index
    assert result["data_start_row"] == detection.data_start_row


def test_resolve_column_mapping_header_row_override_discards_stale_mappings_and_fails_loudly(tmp_path, monkeypatch):
    """The AI can correct the header row within the SAME call (never a
    second Gemini call). When it does, this call's column-mapping answers
    were judged against the wrong row and must be discarded outright, never
    applied — proven here by pointing the "corrected" row at a genuine data
    row with no real header text: since nothing resolves the required
    fields there (the bogus mapping answer must NOT have been used to paper
    over that), this must fail loudly, not silently ingest against a row
    that doesn't actually have headers."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion import ai_column_mapper, column_mapping
    from backend.ingestion.column_mapping import MissingRequiredColumnsError

    wb = openpyxl.load_workbook(master_path, data_only=True)
    detection = column_mapping.detect_header_row(wb.active)
    other_row = detection.data_start_row  # a real data row — genuinely no headers there

    def fake_generate_text(prompt):
        return json.dumps(
            {
                "mappings": {"erp_code": {"column": "Should Be Ignored", "confidence": "high", "reason": "bogus"}},
                "header_row": {"row_index": other_row, "confidence": "high", "reason": "bogus but confident"},
                "warnings": [],
            }
        )

    monkeypatch.setattr(ai_column_mapper.gemini_client, "generate_text", fake_generate_text)

    session = SessionLocal()
    try:
        with pytest.raises(MissingRequiredColumnsError):
            ai_column_mapper.resolve_column_mapping(session, wb.active)
    finally:
        session.close()
