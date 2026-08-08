"""Tests for dynamic header-row detection (this task — never a fixed
HEADER_ROW/DATA_START_ROW constant, see column_mapping.py's module
docstring). Covers detect_header_row() against every known sheet shape
plus a genuinely ambiguous one, resolve_header_row()'s AI-override
precedence, and the Part 5 sanity-gate backstop (validate_sheet_layout()).
"""
import os

import openpyxl
import pytest

_FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "test_fixtures")


def _load(name):
    wb = openpyxl.load_workbook(os.path.join(_FIXTURES_DIR, name), data_only=True)
    return wb.active


# ---- detect_header_row() across known shapes --------------------------------


def test_detects_header_row_3_shape_two_pre_header_rows():
    """messy_sheet_1.xlsx: row 1 blank, row 2 sparse master-block labels,
    row 3 the real per-column header, data from row 4."""
    from backend.ingestion.column_mapping import detect_header_row

    ws = _load("messy_sheet_1.xlsx")
    detection = detect_header_row(ws)
    assert detection.row_index == 3
    assert detection.data_start_row == 4
    assert detection.confident


def test_detects_header_row_1_shape_single_header_row():
    """single_header_row_sheet.xlsx (this task's new fixture): header on
    row 1, data from row 2, no master-block row at all — the real Finance
    file's shape."""
    from backend.ingestion.column_mapping import detect_header_row

    ws = _load("single_header_row_sheet.xlsx")
    detection = detect_header_row(ws)
    assert detection.row_index == 1
    assert detection.data_start_row == 2
    assert detection.confident


def test_detects_header_row_after_renaming_every_header():
    """Renamed headers are still text, not numbers — detection must not
    depend on any specific header wording matching REQUIRED_HEADER_DEFAULTS
    (a renamed sheet has ~zero literal hits and must still win on the
    text/numeric shape signal alone)."""
    from backend.ingestion.column_mapping import detect_header_row

    ws = _load("single_header_row_sheet.xlsx")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.value = f"Col{c}"
    detection = detect_header_row(ws)
    assert detection.row_index == 1
    assert detection.confident


def test_ambiguous_sheet_is_not_confident():
    """Two rows that both look equally header-like (all-text, full
    coverage, no data row heavy with numbers to tell them apart) must not
    produce a confident guess."""
    from backend.ingestion.column_mapping import detect_header_row

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in (1, 2):
        for c in range(1, 6):
            ws.cell(row=row, column=c, value=f"Label {row}-{c}")
    detection = detect_header_row(ws)
    assert not detection.confident
    assert detection.runner_up is not None


# ---- resolve_header_row(): AI-override precedence ---------------------------


def test_resolve_header_row_uses_deterministic_candidate_with_no_ai_input():
    from backend.ingestion.column_mapping import resolve_header_row

    ws = _load("single_header_row_sheet.xlsx")
    header_row, data_start_row = resolve_header_row(ws)
    assert (header_row, data_start_row) == (1, 2)


def test_resolve_header_row_ignores_low_confidence_ai_disagreement():
    from backend.ingestion.column_mapping import resolve_header_row

    ws = _load("single_header_row_sheet.xlsx")
    header_row, _ = resolve_header_row(ws, ai_override={"row_index": 2, "confidence": "low"})
    assert header_row == 1  # deterministic candidate wins — AI wasn't confident


def test_resolve_header_row_applies_high_confidence_ai_override():
    from backend.ingestion.column_mapping import resolve_header_row

    ws = _load("messy_sheet_1.xlsx")
    # Deterministic candidate is row 3; AI confidently proposes a different,
    # in-bounds row — must be applied even though the deterministic pass
    # was itself already confident (same "AI can override" rule as
    # resolve_column_mapping()'s per-field precedence).
    header_row, data_start_row = resolve_header_row(ws, ai_override={"row_index": 2, "confidence": "high"})
    assert (header_row, data_start_row) == (2, 3)


def test_resolve_header_row_raises_when_genuinely_ambiguous():
    from backend.ingestion.column_mapping import AmbiguousHeaderRowError, resolve_header_row

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in (1, 2):
        for c in range(1, 6):
            ws.cell(row=row, column=c, value=f"Label {row}-{c}")
    with pytest.raises(AmbiguousHeaderRowError):
        resolve_header_row(ws)  # no AI input either — nothing breaks the tie


def test_resolve_header_row_ambiguous_case_resolved_by_high_confidence_ai():
    from backend.ingestion.column_mapping import resolve_header_row

    wb = openpyxl.Workbook()
    ws = wb.active
    for row in (1, 2):
        for c in range(1, 6):
            ws.cell(row=row, column=c, value=f"Label {row}-{c}")
    # The deterministic pass ties row 1 vs row 2 (both equally text-like) and
    # picks row 1 as its best guess but isn't confident — a high-confidence
    # AI answer for the OTHER tied row is a genuinely different, real answer
    # and must resolve the tie rather than raising.
    header_row, data_start_row = resolve_header_row(ws, ai_override={"row_index": 2, "confidence": "high"})
    assert (header_row, data_start_row) == (2, 3)


# ---- Part 5: validate_sheet_layout() sanity gate ----------------------------


def test_validate_sheet_layout_passes_for_a_well_formed_sheet():
    from backend.ingestion.column_mapping import build_sheet_map, validate_sheet_layout

    ws = _load("single_header_row_sheet.xlsx")
    sheet_map = build_sheet_map(ws)
    validate_sheet_layout(ws, sheet_map)  # must not raise


def test_validate_sheet_layout_rejects_a_data_row_mistaken_for_headers():
    """Reproduces this task's actual production bug directly: a SheetMap
    whose header_row/data_start_row point one row too low (a real vendor
    data row masquerading as the header row) must be caught here, even
    though build_sheet_map()'s own column lookups aren't exercised by this
    unit test (they're covered separately — this isolates the Part 5
    backstop itself, the last line of defense if every layer above it
    somehow still produced a wrong SheetMap)."""
    import dataclasses

    from backend.ingestion.column_mapping import SheetLayoutError, build_sheet_map, validate_sheet_layout

    ws = _load("single_header_row_sheet.xlsx")
    good_map = build_sheet_map(ws)
    # Same column positions, but header_row/data_start_row shifted down by
    # one — header_row now points at what's actually the first data row
    # (numbers), data_start_row at what's actually the second data row.
    bad_map = dataclasses.replace(
        good_map, header_row=good_map.data_start_row, data_start_row=good_map.data_start_row + 1
    )
    with pytest.raises(SheetLayoutError):
        validate_sheet_layout(ws, bad_map)


if __name__ == "__main__":
    test_detects_header_row_3_shape_two_pre_header_rows()
    test_detects_header_row_1_shape_single_header_row()
    test_detects_header_row_after_renaming_every_header()
    test_ambiguous_sheet_is_not_confident()
    test_resolve_header_row_uses_deterministic_candidate_with_no_ai_input()
    test_resolve_header_row_ignores_low_confidence_ai_disagreement()
    test_resolve_header_row_applies_high_confidence_ai_override()
    test_resolve_header_row_raises_when_genuinely_ambiguous()
    test_resolve_header_row_ambiguous_case_resolved_by_high_confidence_ai()
    test_validate_sheet_layout_passes_for_a_well_formed_sheet()
    test_validate_sheet_layout_rejects_a_data_row_mistaken_for_headers()
    print("OK")
