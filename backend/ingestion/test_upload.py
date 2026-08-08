"""Tests for the Excel upload/commit/revert flow (data-pipeline-upload
task, docs/07 + docs/11). Same isolation convention as test_load_excel.py
(fresh temp DB via PAYMENTS_DB_PATH + module re-import) plus a tmp_path copy
of the real master Excel for every file operation — the real
data/Vendor's Details.xlsx and backend/db/app.db are never touched.

REVISED (Sarath's course-correction, this task): there is no preview step
any more — commit_upload() is the whole flow, one call. The diff-detection
tests below now exercise that same diff logic through commit_upload()'s own
response instead of a since-removed preview_upload().
"""
import os
import shutil
import sys

import openpyxl
import pytest
from sqlalchemy import func


def _fresh_db_and_master(tmp_path):
    """Fresh temp DB, seeded with a baseline ingestion of a tmp_path copy of
    the real master sheet — mirrors test_load_excel.py's own
    _load_against_fresh_db() helper exactly, including its narrow pop list
    (just these two modules, not "every backend.* module"): popping wider
    than that (confirmed the hard way) leaves any OTHER already-imported
    module — e.g. backend.models.new_model_2.allocator, whose own test file
    does `from ... import generate_plan` at collection time — holding a
    stale reference once this function re-imports load_excel/session fresh,
    so a later test's `import allocator_module` gets a DIFFERENT module
    object than the generate_plan it already imported, and monkeypatching
    the former silently does nothing to the latter. Every test below passes
    explicit excel_path/backup_path overrides anyway, so upload.py's own
    module-level EXCEL_PATH/MASTER_EXCEL_BACKUP_PATH constants are never
    read from a stale copy either way — no need to pop that module too."""
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    # Every module a test below actually imports from and that itself calls
    # SessionLocal() (directly, or via a bare `session=None` default) needs
    # popping here — not just backend.db.session itself, or a stale cached
    # copy of e.g. backend.ingestion.upload would keep calling the OLD
    # SessionLocal it captured on its first import in an earlier test.
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.ingestion.upload",
        "backend.configuration.vendor_edits",
        "backend.configuration.extra_fields",
    ):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH, load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    load(excel_path=master_path)
    return master_path


def _add_unmapped_column(src_path, dest_path, header, value_fn):
    """value_fn(erp_code, row_offset) -> cell value, for every real data row."""
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = ws.max_column + 1
    ws.cell(row=sheet_map.header_row, column=col, value=header)
    offset = 0
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        erp_code = ws.cell(row=row, column=sheet_map.erp_code_col).value
        if erp_code is None:
            continue
        ws.cell(row=row, column=col, value=value_fn(erp_code, offset))
        offset += 1
    wb.save(dest_path)


def _change_vendor_cell(src_path, dest_path, erp_code, header, new_value):
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, header)
    assert col is not None, f"column {header!r} not found"
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            ws.cell(row=row, column=col, value=new_value)
            wb.save(dest_path)
            return
    raise AssertionError(f"vendor {erp_code!r} not found")


def _read_vendor_cell(path, erp_code, header):
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, header)
    assert col is not None, f"column {header!r} not found"
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            return ws.cell(row=row, column=col).value
    raise AssertionError(f"vendor {erp_code!r} not found")


def _remove_vendor_rows(src_path, dest_path, erp_codes):
    """Bug #1 test helper — builds an upload sheet with the given vendors'
    rows entirely absent (not just blanked), the same shape as Sarath's
    real repro (a replacement sheet that doesn't list some previously-known
    vendors at all)."""
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    to_remove = set(erp_codes)
    for row in range(ws.max_row, sheet_map.data_start_row - 1, -1):  # bottom-up so row numbers stay valid
        if ws.cell(row=row, column=sheet_map.erp_code_col).value in to_remove:
            ws.delete_rows(row, 1)
    wb.save(dest_path)


def _remove_vendor_row_by_entity(src_path, dest_path, erp_code, entity):
    """Like _remove_vendor_rows(), but disambiguated by entity too —
    entity-code-collision fix: erp_code alone can match more than one row
    now (two different entities sharing a bare code), so removing "the"
    V00149 row needs to say which one."""
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    for row in range(ws.max_row, sheet_map.data_start_row - 1, -1):  # bottom-up so row numbers stay valid
        if (
            ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code
            and ws.cell(row=row, column=sheet_map.entity_col).value == entity
        ):
            ws.delete_rows(row, 1)
            wb.save(dest_path)
            return
    raise AssertionError(f"vendor {erp_code!r}/{entity!r} not found")


def _fresh_empty_db(tmp_path):
    """Same fresh-temp-DB setup as _fresh_db_and_master(), but skips the
    baseline load() — a genuinely empty DB with no prior ledger month at
    all, for testing the very first upload ever."""
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.ingestion.upload",
        "backend.configuration.vendor_edits",
        "backend.configuration.extra_fields",
    ):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH

    return EXCEL_PATH


def _append_synthetic_month(excel_path, payable=10_000.0, payment=4_000.0):
    """Appends one new month's Payable+Payment column pair immediately
    before the existing Total Payable / Total Payment columns — mirrors
    what a real monthly Excel refresh does (docs/07), just with synthetic
    figures instead of a real new month's data. Row-2 merged header labels
    (e.g. "Monthly Payable" spanning H2:V2) go stale after the column shift
    since openpyxl's insert_cols doesn't adjust merge boundaries — but
    column_mapping.py only ever reads the detected header row and below, so
    this has no effect on ingestion correctness. Moved here from the
    now-deleted backend/month_end/test_rollover.py (merge-rollover-into-
    upload task) — same helper, new home.
    """
    from backend.ingestion.column_mapping import build_sheet_map, to_number

    # The Closing Balance column holds a live formula (e.g. "=G4+V4-AK4"),
    # not a literal — read cached values from a separate data_only load
    # (same convention load_excel.py itself uses) before the formula-
    # preserving workbook below overwrites that cell with a literal number.
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    ws_data = wb_data.active
    data_sheet_map = build_sheet_map(ws_data)
    prior_closing_by_row = {
        row: to_number(ws_data.cell(row=row, column=data_sheet_map.closing_balance_col).value)
        for row in range(data_sheet_map.data_start_row, ws_data.max_row + 1)
        if ws_data.cell(row=row, column=data_sheet_map.erp_code_col).value is not None
    }

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    total_payable_col = sheet_map.total_payable_col
    total_payment_col = sheet_map.total_payment_col

    # Insert both new columns using pre-computed offsets, not a re-called
    # build_sheet_map between the two inserts — the sheet is transiently
    # asymmetric (one new payable column, no matching payment column yet)
    # and build_sheet_map's own equal-block-length check would reject that.
    ws.insert_cols(total_payable_col)
    ws.cell(row=sheet_map.header_row, column=total_payable_col, value="synthetic-payable")
    ws.insert_cols(total_payment_col + 1)
    ws.cell(row=sheet_map.header_row, column=total_payment_col + 1, value="synthetic-payment")

    sheet_map = build_sheet_map(ws)  # symmetric again — safe to rebuild now
    new_payable_col = sheet_map.payable_cols[-1][1]
    new_payment_col = sheet_map.payment_cols[-1][1]

    for row, prior_closing in prior_closing_by_row.items():
        ws.cell(row=row, column=new_payable_col, value=payable)
        ws.cell(row=row, column=new_payment_col, value=payment)
        ws.cell(row=row, column=sheet_map.closing_balance_col, value=prior_closing + payable - payment)

    wb.save(excel_path)


def _duplicate_row_as_new_vendor(src_path, dest_path, source_erp_code, new_erp_code):
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == source_erp_code:
            new_row = ws.max_row + 1
            for col in range(1, ws.max_column + 1):
                ws.cell(row=new_row, column=col, value=ws.cell(row=row, column=col).value)
            ws.cell(row=new_row, column=sheet_map.erp_code_col, value=new_erp_code)
            ws.cell(row=new_row, column=sheet_map.vendor_name_col, value="Test New Vendor")
            wb.save(dest_path)
            return
    raise AssertionError(f"vendor {source_erp_code!r} not found")


def _resave_only_commit_result(tmp_path, subdir):
    """Baseline control: a pure openpyxl load+save round-trip of the real
    master sheet, with zero cell edits, committed through commit_upload().
    Confirmed (not guessed) that even this is enough to shift a handful of
    unrelated vendors' derived fields — openpyxl's resave drops cached
    formula results for the sheet's live Closing Balance formula cells,
    which feeds into a few vendors' opening_balance/category/assigned_week
    on the next load(). That noise is orthogonal to whatever a specific
    test's own deliberate edit is trying to detect, and its exact vendor set
    isn't something to hardcode (it would drift with the real sheet same as
    any other count) — tests below diff against this control instead."""
    control_dir = tmp_path / subdir
    control_dir.mkdir()
    master_path = _fresh_db_and_master(control_dir)
    from backend.ingestion.upload import commit_upload

    import openpyxl

    upload_path = str(control_dir / "resave.xlsx")
    openpyxl.load_workbook(master_path).save(upload_path)
    backup_path = str(control_dir / "resave.backup.xlsx")
    return commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)


# ---- Unit tests: diff computation, exercised through commit_upload() ------
# (no more standalone preview_upload() — Sarath's course-correction, this
# task — but commit_upload() computes and returns this same diff, it just
# also writes at the same time now.)


def test_commit_detects_new_vendor(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload

    session = SessionLocal()
    existing_count = session.query(Vendor).count()  # derived, not hardcoded — real sheet's row count drifts
    session.close()

    upload_path = str(tmp_path / "upload.xlsx")
    _duplicate_row_as_new_vendor(master_path, upload_path, "V00400", "V99999")
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert "V99999" in result["new_vendors"]
    assert result["new_vendor_count"] == 1
    assert result["existing_vendor_count"] == existing_count  # every other real vendor, unchanged


def test_commit_detects_changed_ledger_figure(tmp_path):
    baseline = _resave_only_commit_result(tmp_path, "control")
    baseline_changed = set(baseline["vendors_with_changed_ledger"])

    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.upload import commit_upload

    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", 999_999_999)
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert "V00400" in result["vendors_with_changed_ledger"]
    # Derived against the resave-only baseline rather than a fresh hardcoded
    # count — isolates the delta this test's own edit actually caused from
    # the resave noise documented in _resave_only_commit_result() above.
    assert set(result["vendors_with_changed_ledger"]) == baseline_changed | {"V00400"}
    assert result["vendors_with_changed_ledger_count"] == len(baseline_changed) + 1
    assert any("Apr-25" in c for c in result["changed_fields_by_vendor"]["V00400"])


def test_commit_detects_unmapped_column(tmp_path):
    baseline = _resave_only_commit_result(tmp_path, "control")
    baseline_unmapped = set(baseline["unmapped_columns"])

    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.upload import commit_upload

    upload_path = str(tmp_path / "upload.xlsx")
    _add_unmapped_column(master_path, upload_path, "Status", lambda erp, i: "Active" if i % 2 == 0 else "Inactive")
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    # Derived against the resave-only baseline (same reasoning as
    # test_commit_detects_changed_ledger_figure) instead of hardcoding the
    # real sheet's current unmapped-column set, which drifts as the sheet
    # gains/loses columns — "VN" existed in an earlier version of the real
    # sheet and no longer does.
    assert set(result["unmapped_columns"]) == baseline_unmapped | {"Status"}


def test_commit_classifies_recovered_entity_collision_vendor_as_new_not_anomaly(tmp_path):
    """Entity-code-collision fix, end to end: the real master sheet has
    erp_code "V00149" on two rows — ARS "Instant Screening Services" (row
    103) and FP "G-Task Services" (row 409) — a genuine two-entity code
    reuse, not a duplicate. Before this fix, first-row-wins silently
    dropped the FP row forever; after it, both are live DB rows from the
    very first ingestion. This test simulates the one-time recovery
    moment: a baseline DB seeded WITHOUT the FP row (mirroring the old,
    already-broken state), then a normal upload of the real, complete
    sheet — the previously-missing vendor must be classified as a genuine
    new vendor, not merged into/confused with its same-code ARS sibling.
    """
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.ingestion.upload",
        "backend.configuration.vendor_edits",
        "backend.configuration.extra_fields",
    ):
        sys.modules.pop(mod, None)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import EXCEL_PATH, load
    from backend.ingestion.upload import commit_upload

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    baseline_path = str(tmp_path / "baseline.xlsx")
    _remove_vendor_row_by_entity(master_path, baseline_path, "V00149", "FP")
    shutil.copy(baseline_path, master_path)
    load(excel_path=master_path)  # baseline: FP/"G-Task Services" not in the DB yet

    session = SessionLocal()
    assert session.query(Vendor).filter_by(erp_code="V00149").count() == 1  # only ARS, pre-fix-era state
    session.close()

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(EXCEL_PATH, upload_path)  # the real, complete sheet — both V00149 rows present
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    assert "V00149 (FP)" in result["new_vendors"]  # ambiguous code -> disambiguated label, classified new
    assert "V00149" not in result["new_vendors"]  # never the bare, ambiguous form
    assert "V00149 (ARS)" not in result["new_vendors"]  # ARS sibling already existed — not "new"

    session = SessionLocal()
    try:
        vendors = session.query(Vendor).filter_by(erp_code="V00149").order_by(Vendor.entity).all()
        assert {v.entity for v in vendors} == {"ARS", "FP"}
        assert {v.vendor_name for v in vendors} == {"Instant Screening Services", "G-Task Services"}
    finally:
        session.close()


# ---- Integration tests: full upload -> commit -> DB+Excel state -----------


def test_commit_updates_db_and_master_excel_to_match_upload(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from datetime import date

    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", 12_345_678)
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["vendors_changed"] >= 1

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
        ledger_row = session.query(MonthlyLedger).filter_by(vendor_id=vendor.id, month=date(2025, 4, 1)).one()
        assert float(ledger_row.payable) == pytest.approx(12_345_678)
    finally:
        session.close()

    # Master Excel now IS the uploaded file (atomically replaced) — reload
    # it fresh and confirm the same cell reads back the new value.
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(master_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, "Apr-25")
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == "V00400":
            assert ws.cell(row=row, column=col).value == 12_345_678
            break
    else:
        raise AssertionError("V00400 not found in post-commit master file")


def test_commit_writes_audit_log_entries_with_excel_upload_source(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import AuditLog
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import ChangeSource

    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", 12_345_678)
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    try:
        # Audit-log revamp (Task B): one summary row per upload now, not
        # one per vendor touched — no longer keyed to a specific vendor_id.
        # Doesn't assert an exact "changed" count: a pure openpyxl
        # load+save round-trip of the real master sheet, even with zero
        # OTHER cell edits, already shifts some unrelated vendors' derived
        # fields (Closing Balance's cached formula value, dropped on
        # resave — see CLAUDE.md's Class-1 dataset-drift note) — this test
        # only needs to confirm a summary row exists and mentions "changed".
        entry = (
            session.query(AuditLog)
            .filter_by(vendor_id=None, source=ChangeSource.EXCEL_UPLOAD.value)
            .one()
        )
        assert "changed" in entry.new_value
        assert entry.old_value is None  # summary, not a cell-by-cell before/after log
    finally:
        session.close()


# ---- Integration test: backup -> bad upload -> revert ---------------------


def test_backup_bad_upload_revert_restores_pre_upload_state(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload, revert_upload
    from datetime import date

    session = SessionLocal()
    original_payable = float(
        session.query(MonthlyLedger)
        .join(Vendor)
        .filter(Vendor.erp_code == "V00400", MonthlyLedger.month == date(2025, 4, 1))
        .one()
        .payable
    )
    session.close()

    bad_upload_path = str(tmp_path / "bad_upload.xlsx")
    _change_vendor_cell(master_path, bad_upload_path, "V00400", "Apr-25", -1)  # obviously "bad" data
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(bad_upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    bad_value = float(
        session.query(MonthlyLedger)
        .join(Vendor)
        .filter(Vendor.erp_code == "V00400", MonthlyLedger.month == date(2025, 4, 1))
        .one()
        .payable
    )
    session.close()
    assert bad_value == pytest.approx(-1)  # confirms the bad upload actually landed before we revert it

    revert_result = revert_upload(excel_path=master_path, backup_path=backup_path)
    assert "not covered by this revert" in revert_result["warning"]

    session = SessionLocal()
    try:
        restored_value = float(
            session.query(MonthlyLedger)
            .join(Vendor)
            .filter(Vendor.erp_code == "V00400", MonthlyLedger.month == date(2025, 4, 1))
            .one()
            .payable
        )
        assert restored_value == pytest.approx(original_payable)
    finally:
        session.close()

    wb = openpyxl.load_workbook(master_path)
    ws = wb.active
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    sheet_map = build_sheet_map(ws)
    col = find_column(ws, "Apr-25")
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == "V00400":
            assert ws.cell(row=row, column=col).value != -1
            break


# ---- Bug #1: replace-upload must actually replace, not merge/append -------


def test_commit_soft_deactivates_vendor_missing_from_upload(tmp_path):
    """The repro: a vendor whose erp_code is absent from the newly uploaded
    sheet must disappear from the live vendor list/grid (soft-deactivated),
    not stay live alongside the new data — and never be hard-deleted, since
    PlanAllocation/Payment/AuditLog rows already reference it and SQLite FK
    enforcement is off here (models.py's Vendor.is_active comment).

    Picks its own victim erp_codes from whatever master sheet is actually
    loaded rather than hardcoding "V00400" — the real data/Vendor's
    Details.xlsx this copies from is a shared, externally-mutable file (see
    memory: demo15-dataset-quirks) that has drifted vendor codes before."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import AuditLog, MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.grid import build_master_grid
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import ChangeSource

    session = SessionLocal()
    total_before = session.query(Vendor).count()
    removed = [v.erp_code for v in session.query(Vendor).order_by(Vendor.id).limit(2).all()]
    session.close()
    assert len(removed) == 2  # sanity: master sheet has at least 2 vendors to remove

    upload_path = str(tmp_path / "upload.xlsx")
    _remove_vendor_rows(master_path, upload_path, removed)
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["deactivated_vendor_count"] == 2
    assert set(result["deactivated_erp_codes"]) == set(removed)

    session = SessionLocal()
    try:
        # Row still physically present (not hard-deleted) but flagged inactive,
        # and its sheet-derived rows are dropped.
        for erp_code in removed:
            vendor = session.query(Vendor).filter_by(erp_code=erp_code).one()
            assert vendor.is_active is False
            assert vendor.removed_at is not None
            assert session.query(MonthlyLedger).filter_by(vendor_id=vendor.id).count() == 0

        # Audit-log revamp (Task B): one summary row per upload now, not
        # one per vendor removed — names the count, not each vendor.
        entry = (
            session.query(AuditLog)
            .filter_by(vendor_id=None, source=ChangeSource.EXCEL_UPLOAD.value)
            .one()
        )
        assert "2 removed" in entry.new_value

        # Total row count is unchanged (soft-deactivate, not delete)...
        assert session.query(Vendor).count() == total_before
        # ...but every live-facing read excludes them: Master Data grid,
        # min-funds' shared vendor loader — replaced, not merged.
        grid = build_master_grid(session, excel_path=master_path)
        assert {v["erp_code"] for v in grid["vendors"]} & set(removed) == set()
        assert len(grid["vendors"]) == total_before - 2

        from backend.shared.min_funds import _load_vendors_and_ledger

        live_vendors, _ = _load_vendors_and_ledger(session)
        assert {v.erp_code for v in live_vendors} & set(removed) == set()
    finally:
        session.close()


def test_revert_after_deactivating_upload_restores_the_vendor(tmp_path):
    """The one-slot backup/revert flow must still work exactly as before:
    reverting past a replace-upload that dropped some vendors brings them
    back live (reactivated, not re-inserted as a new row)."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload, revert_upload

    session = SessionLocal()
    total_before = session.query(Vendor).count()
    target = session.query(Vendor).order_by(Vendor.id).first()
    target_erp_code, original_vendor_id = target.erp_code, target.id
    session.close()

    upload_path = str(tmp_path / "upload.xlsx")
    _remove_vendor_rows(master_path, upload_path, [target_erp_code])
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    assert session.query(Vendor).filter_by(erp_code=target_erp_code).one().is_active is False
    session.close()

    revert_upload(excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code=target_erp_code).one()
        assert vendor.is_active is True
        assert vendor.removed_at is None
        assert vendor.id == original_vendor_id  # reactivated the same row, not re-inserted
        assert session.query(Vendor).count() == total_before  # no duplicate row created
    finally:
        session.close()


def test_revert_without_a_prior_upload_raises_clean_error(tmp_path):
    _fresh_db_and_master(tmp_path)
    from backend.ingestion.upload import revert_upload

    with pytest.raises(ValueError, match="no backup available"):
        revert_upload(excel_path=str(tmp_path / "master.xlsx"), backup_path=str(tmp_path / "nonexistent.backup.xlsx"))


# ---- Round-trip test: VendorExtraField (widget derivation + edit) ---------


def test_extra_field_round_trip_widget_derivation_and_edit(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.extra_fields import list_extra_fields
    from backend.configuration.vendor_edits import update_extra_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload

    # 2-value column -> toggle.
    upload_1 = str(tmp_path / "upload1.xlsx")
    _add_unmapped_column(master_path, upload_1, "Toggle Col", lambda erp, i: "Yes" if i % 2 == 0 else "No")
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(upload_1, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    fields = {f["column_name"]: f for f in list_extra_fields(session)}
    session.close()
    assert fields["Toggle Col"]["widget"] == "toggle"
    assert set(fields["Toggle Col"]["options"]) == {"Yes", "No"}

    # 5-value column -> dropdown with those exact 5 options.
    upload_2 = str(tmp_path / "upload2.xlsx")
    five_values = ["A", "B", "C", "D", "E"]
    _add_unmapped_column(master_path, upload_2, "Dropdown Col", lambda erp, i: five_values[i % 5])
    commit_upload(upload_2, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    fields = {f["column_name"]: f for f in list_extra_fields(session)}
    vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
    vendor_id = vendor.id
    session.close()
    assert fields["Dropdown Col"]["widget"] == "dropdown"
    assert set(fields["Dropdown Col"]["options"]) == set(five_values)

    # Edit one vendor's value through the new endpoint-backing function —
    # confirm DB, audit_log, and the next Excel read all agree.
    excel_copy = str(tmp_path / "edit_copy.xlsx")
    shutil.copy(master_path, excel_copy)
    old_value, new_value = update_extra_field(vendor_id, "Dropdown Col", "Z", excel_path=excel_copy)
    assert new_value == "Z"

    session = SessionLocal()
    try:
        fields = {f["column_name"]: f for f in list_extra_fields(session)}
        assert fields["Dropdown Col"]["values_by_vendor_id"][vendor_id] == "Z"
        entry = session.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="Dropdown Col").one()
        assert entry.new_value == "Z"
    finally:
        session.close()

    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(excel_copy)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, "Dropdown Col")
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == "V00400":
            assert ws.cell(row=row, column=col).value == "Z"
            break
    else:
        raise AssertionError("V00400 not found when reading back the edited column")


# ---- End-to-end (Fix 3): uploaded data actually reaches Planning/Payments -
# generate_plan()/calculate_minimum_funds_required()/log_payment() all
# default to SessionLocal() when no session is passed — a stale cached copy
# of THEIR module (backend.models.new_model_2.allocator,
# backend.shared.min_funds, backend.shared.payment_logging) from an earlier
# test elsewhere in the same pytest process would silently keep talking to
# an old temp DB via that internal default instead of this test's fresh
# one. Popping those modules to dodge that is what test_min_funds.py's own
# object-identity tests exist to catch (confirmed the hard way: doing so
# here broke test_model2_advisor_reexports_the_same_function_objects and
# test_model1_scorer_does_not_import_model2_package when the full suite ran
# — those modules import from min_funds and compare `is` identity, so
# popping min_funds out from under them mid-session breaks that check for
# reasons unrelated to any real bug). Fix: build ONE session from this
# file's own already-fresh backend.db.session.SessionLocal and pass it
# EXPLICITLY into every call below — the callee then never reaches its own
# internal SessionLocal default at all, so which module copy it happens to
# be attached to stops mattering, and no other test file's cache needs
# touching.


def test_upload_data_reaches_new_model_2_generate_plan(tmp_path):
    """Fix 3 — a vendor's ledger figure changed by an upload must show up
    in the very next New Model 2 generate_plan() call for that vendor, not
    a stale pre-upload number. Proves there's no cached session/vendor
    object anywhere between the upload path and the planning path."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.models.new_model_2.allocator import generate_plan
    from backend.shared.min_funds_v2 import calculate_minimum_funds_required_v2

    session = SessionLocal()
    try:
        before = calculate_minimum_funds_required_v2(session=session)
        before_breakdown = before["breakdown"]
        before_required = float(
            before_breakdown[before_breakdown["erp_code"] == "V00400"]["required_amount"].iloc[0]
        )
    finally:
        session.close()

    original_apr_payable = float(_read_vendor_cell(master_path, "V00400", "Apr-25") or 0)
    delta = 50_000_000.0
    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", original_apr_payable + delta)
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    try:
        after = calculate_minimum_funds_required_v2(session=session)
        after_breakdown = after["breakdown"]
        after_required = float(
            after_breakdown[after_breakdown["erp_code"] == "V00400"]["required_amount"].iloc[0]
        )
        # Directional, not an exact-delta prediction: required_amount()'s
        # carry-forward walk (backend/shared/min_funds.py) can fold the
        # bumped tranche in alongside others already in its window, so the
        # exact increase isn't a simple +delta — that formula has its own
        # dedicated tests elsewhere. What matters here is that the figure
        # actually moved off the pre-upload value, proving generate_plan()
        # below reads live data, not something cached from before the upload.
        assert after_required > before_required
        assert after_required != pytest.approx(before_required)

        plan = generate_plan(available_funds=after["total"] * 2, session=session)  # absolute funds -> 100% for everyone
        row = plan["allocations"][plan["allocations"]["erp_code"] == "V00400"].iloc[0]
        assert float(row["required_amount"]) == pytest.approx(after_required, abs=1)
        assert float(row["allocated_amount"]) == pytest.approx(after_required, abs=1)
    finally:
        session.close()


def test_upload_data_reaches_payment_logging(tmp_path):
    """Fix 3 — log_payment() must apply against the vendor's balance as it
    stands AFTER the most recent upload. payment_amount below is bigger
    than the OLD (pre-upload) balance but well within the NEW (post-upload)
    one — if log_payment() read a stale pre-upload balance for its
    ledger-integrity guard, this would wrongly raise "payment exceeds what's
    owed" instead of succeeding."""
    master_path = _fresh_db_and_master(tmp_path)
    from datetime import date

    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.payment_logging import log_payment

    session = SessionLocal()
    vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
    old_opening_balance = float(vendor.opening_balance)
    vendor_id = vendor.id
    session.close()

    original_apr_payable = float(_read_vendor_cell(master_path, "V00400", "Apr-25") or 0)
    delta = 50_000_000.0
    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", original_apr_payable + delta)
    backup_path = str(tmp_path / "master.backup.xlsx")
    commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    new_opening_balance = float(session.query(Vendor).filter_by(id=vendor_id).one().opening_balance)
    session.close()
    assert new_opening_balance == pytest.approx(old_opening_balance + delta, abs=1)

    payment_amount = old_opening_balance + 10_000_000.0  # > old balance, well within the new one
    session = SessionLocal()
    try:
        log_payment(vendor_id, payment_amount, today=date(2026, 5, 8), session=session)  # raises if stale-balance-based
        session.commit()
    finally:
        session.close()

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(id=vendor_id).one()
        assert float(vendor.paid_so_far_this_month) == pytest.approx(payment_amount)
        live_outstanding = float(vendor.opening_balance) - float(vendor.paid_so_far_this_month)
        assert live_outstanding == pytest.approx(new_opening_balance - payment_amount, abs=1)
    finally:
        session.close()


# ---- Month-end cycle reset, folded into commit_upload() (merge-rollover-
# into-upload task) — replaces the now-deleted standalone
# backend/month_end/test_rollover.py, since rollover_month() itself is gone
# and this behavior now lives inside commit_upload() instead. ----------------


def test_commit_upload_resets_cycle_state_on_genuine_new_month(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import PaymentStatus

    session = SessionLocal()
    original_vendor_count = session.query(Vendor).count()
    # Give one real vendor actual payment-cycle state before the upload, so
    # the reset is proven against a genuinely non-default prior state.
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id = vendor.id
    vendor.payment_status = PaymentStatus.PARTIAL
    vendor.paid_so_far_this_month = 250_000.0
    vendor.override_amount = 999_999.0
    vendor.week_distribution_plan = {"1": 999_999.0}
    vendor.finalized_budget_amount = 999_999.0
    session.commit()
    session.close()

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(master_path, upload_path)
    _append_synthetic_month(upload_path)
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["cycle_reset"] is True
    assert result["vendors_reset"] == original_vendor_count

    session = SessionLocal()
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == PaymentStatus.NOT_PAID  # before: PARTIAL
    assert float(reloaded.paid_so_far_this_month) == 0.0  # before: 250,000
    assert reloaded.override_amount is None  # before: 999,999
    assert reloaded.week_distribution_plan is None  # before: {"1": 999,999}
    assert reloaded.finalized_budget_amount is None  # before: 999,999
    # The new month's own ledger rows genuinely landed too — this isn't
    # just a reset with the ingestion skipped.
    latest_month = session.query(func.max(MonthlyLedger.month)).scalar()
    assert len(session.query(MonthlyLedger).filter_by(month=latest_month).all()) == original_vendor_count
    session.close()


def test_commit_upload_same_month_correction_does_not_reset_or_error(tmp_path):
    """The routine case now that every re-upload takes this same path —
    must NOT raise (the old standalone rollover_month() refused this with a
    ValueError; that framing no longer applies once this is the only
    upload path) and must NOT reset anything."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload
    from backend.shared.enums import PaymentStatus

    session = SessionLocal()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id = vendor.id
    vendor.payment_status = PaymentStatus.PARTIAL
    vendor.paid_so_far_this_month = 99_999.0
    session.commit()
    session.close()

    # No new month appended — an ordinary same-month correction re-upload.
    upload_path = str(tmp_path / "upload.xlsx")
    _change_vendor_cell(master_path, upload_path, "V00400", "Apr-25", 555_555.0)
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)  # must not raise
    assert result["cycle_reset"] is False
    assert result["vendors_reset"] == 0

    session = SessionLocal()
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == PaymentStatus.PARTIAL  # unchanged
    assert float(reloaded.paid_so_far_this_month) == 99_999.0  # unchanged
    session.close()


def test_commit_upload_first_upload_ever_does_not_reset_or_error(tmp_path):
    """No prior ledger month to compare against at all — must not attempt a
    reset and must not error either."""
    excel_path = _fresh_empty_db(tmp_path)
    from backend.ingestion.upload import commit_upload

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(excel_path, upload_path)
    master_path = str(tmp_path / "master.xlsx")  # doesn't exist yet — a genuinely fresh environment
    backup_path = str(tmp_path / "master.backup.xlsx")

    result = commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)
    assert result["cycle_reset"] is False
    assert result["vendors_reset"] == 0


def test_commit_upload_reset_failure_rolls_back_ingestion_too(tmp_path, monkeypatch):
    """A genuine failure during the reset step — after the new-month check
    has already confirmed a real new month — must roll back BOTH the reset
    attempt and this call's re-ingestion together. Without that, the new
    month's ledger data would already be committed, and a retry would
    immediately look like a same-month correction (no reset attempted
    again) with no way to force the reset through short of manual DB
    surgery."""
    master_path = _fresh_db_and_master(tmp_path)
    import backend.month_end.rollover as rollover_module
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.upload import commit_upload

    session = SessionLocal()
    original_latest_month = session.query(func.max(MonthlyLedger.month)).scalar()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id, original_status = vendor.id, vendor.payment_status
    session.close()

    upload_path = str(tmp_path / "upload.xlsx")
    shutil.copy(master_path, upload_path)
    _append_synthetic_month(upload_path)
    backup_path = str(tmp_path / "master.backup.xlsx")

    class _ExplodingPaymentStatus:
        @property
        def NOT_PAID(self):
            raise RuntimeError("simulated failure during the reset loop")

    monkeypatch.setattr(rollover_module, "PaymentStatus", _ExplodingPaymentStatus())

    with pytest.raises(RuntimeError):
        commit_upload(upload_path, excel_path=master_path, backup_path=backup_path)

    session = SessionLocal()
    # Both the reset AND the re-ingestion it was sharing a transaction with
    # must be undone — the ledger must still show the ORIGINAL latest
    # month, not the new one the failed attempt tried to land.
    assert session.query(func.max(MonthlyLedger.month)).scalar() == original_latest_month
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == original_status
    session.close()


if __name__ == "__main__":
    # ponytail: no framework/fixtures — a minimal manual run for a quick
    # sanity check (pytest's tmp_path fixture isn't available standalone).
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as d:
        tmp_path = Path(d)
        test_commit_detects_new_vendor(tmp_path)
    with tempfile.TemporaryDirectory() as d:
        tmp_path = Path(d)
        test_extra_field_round_trip_widget_derivation_and_edit(tmp_path)
    with tempfile.TemporaryDirectory() as d:
        tmp_path = Path(d)
        test_upload_data_reaches_new_model_2_generate_plan(tmp_path)
    with tempfile.TemporaryDirectory() as d:
        tmp_path = Path(d)
        test_upload_data_reaches_payment_logging(tmp_path)
    print("upload.py self-checks passed")
