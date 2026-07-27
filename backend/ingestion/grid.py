"""Master Data grid read model (Part C's backend support, docs/07 + 11).

Builds one row per vendor, one column per real Excel column, in the file's
own column order — read-only reconstruction from the live workbook's header
row (for column identity/order) plus the DB (for values), never a second
copy of ingestion's own parsing logic. category/commitment_months/
assigned_week/priority_tag are shown here for reference only
(editable=False) — Configuration-module scope decision (docs/11, Task 2):
editing these four vendor tags moved to the Planning screen's inline
per-row dropdowns (still calling the same PATCH /vendors/{id} ->
update_vendor_field() write-back+audit-log path, not a second
implementation). Raw ledger figures (Payable/Payment/Totals/Closing/W1-W5)
are display-only in this grid too — CLAUDE.md's ledger-integrity rules mean
those aren't hand-editable cell by cell anywhere in this codebase, this
grid doesn't change that.

AMBIGUOUS SPEC DETAIL, flagged rather than guessed (house style, see
allocator.py's FALLBACK_BUCKET comment): the sheet's W1-W5/Total/legacy
"Assigned Week-Order" columns are a snapshot of ONE plan_run at ingestion
time, not a live per-vendor DB figure the way category/opening_balance are
— there's no single "current" value for them outside a specific plan_run.
Returned as None here; Finance already has the live weekly view in the
Planning tab's own tables, so this grid intentionally doesn't duplicate it.
"""
import openpyxl

from backend.db.models import MonthlyLedger, Vendor, VendorExtraField
from backend.ingestion.load_excel import EXCEL_PATH
from backend.ingestion.column_mapping import FIELD_TO_EXCEL_HEADER, build_sheet_map, find_column, unmapped_header_columns
from backend.configuration.extra_fields import list_extra_fields


def _describe_columns(ws, sheet_map, category_col, commitment_months_col, assigned_week_col, priority_tag_col, unmapped):
    payable_month_by_col = {col: month for month, col in sheet_map.payable_cols}
    payment_month_by_col = {col: month for month, col in sheet_map.payment_cols}
    week_number_by_col = {col: wk for wk, col in sheet_map.week_cols}
    unmapped_by_col = dict(unmapped)

    columns = []
    for cell in (ws.cell(row=3, column=c) for c in range(1, ws.max_column + 1)):
        header = str(cell.value).strip() if cell.value is not None else None
        if header is None:
            continue
        c = cell.column
        if c == sheet_map.erp_code_col:
            columns.append({"header": header, "kind": "erp_code", "editable": False})
        elif c == sheet_map.entity_col:
            columns.append({"header": header, "kind": "entity", "editable": False})
        elif c == sheet_map.vendor_name_col:
            columns.append({"header": header, "kind": "vendor_name", "editable": False})
        elif c == sheet_map.opening_balance_col:
            columns.append({"header": header, "kind": "opening_balance", "editable": False})
        elif c in payable_month_by_col:
            columns.append({"header": header, "kind": "payable", "month": payable_month_by_col[c].isoformat(), "editable": False})
        elif c == sheet_map.total_payable_col:
            columns.append({"header": header, "kind": "total_payable", "editable": False})
        elif c in payment_month_by_col:
            columns.append({"header": header, "kind": "payment", "month": payment_month_by_col[c].isoformat(), "editable": False})
        elif c == sheet_map.total_payment_col:
            columns.append({"header": header, "kind": "total_payment", "editable": False})
        elif c == sheet_map.closing_balance_col:
            columns.append({"header": header, "kind": "closing_balance", "editable": False})
        elif c in week_number_by_col:
            columns.append({"header": header, "kind": "week", "week_number": week_number_by_col[c], "editable": False})
        elif c == sheet_map.week_total_col:
            columns.append({"header": header, "kind": "week_total", "editable": False})
        elif c == sheet_map.assigned_week_order_col:
            columns.append({"header": header, "kind": "assigned_week_order_legacy", "editable": False})
        elif c == category_col:
            columns.append({"header": header, "kind": "category", "field": "category", "editable": False})
        elif c == commitment_months_col:
            columns.append({"header": header, "kind": "commitment_months", "field": "commitment_months", "editable": False})
        elif c == assigned_week_col:
            columns.append({"header": header, "kind": "assigned_week", "field": "assigned_week", "editable": False})
        elif c == priority_tag_col:
            columns.append({"header": header, "kind": "priority_tag", "field": "priority_tag", "editable": False})
        elif c in unmapped_by_col:
            columns.append({"header": header, "kind": "extra", "column_name": unmapped_by_col[c], "editable": True})
        # else: a header cell that's none of the above shouldn't happen —
        # unmapped_header_columns() already covers "everything not known".
    return columns


def build_master_grid(session, excel_path=None):
    """Read-only. `session` supplies every value; `excel_path` (defaults to
    the real master) is opened only to learn column identity/order —
    values always come from the DB, never straight from this workbook read,
    so a grid built mid-preview never shows uncommitted data."""
    excel_path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    category_col = find_column(ws, FIELD_TO_EXCEL_HEADER["category"])
    commitment_months_col = find_column(ws, FIELD_TO_EXCEL_HEADER["commitment_months"])
    assigned_week_col = find_column(ws, FIELD_TO_EXCEL_HEADER["assigned_week"])
    priority_tag_col = find_column(ws, FIELD_TO_EXCEL_HEADER["priority_tag"])
    unmapped = unmapped_header_columns(
        ws, sheet_map, extra_cols=(category_col, commitment_months_col, assigned_week_col, priority_tag_col)
    )
    columns = _describe_columns(
        ws, sheet_map, category_col, commitment_months_col, assigned_week_col, priority_tag_col, unmapped
    )
    extra_widgets = {f["column_name"]: {"widget": f["widget"], "options": f["options"]} for f in list_extra_fields(session)}

    vendors_out = []
    # is_active.isnot(False), not ==True: a vendor row that predates this
    # column (Bug #1 fix, models.py) reads NULL, which must count as active.
    for vendor in session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).all():
        ledger_by_month = {
            row.month.isoformat(): row for row in session.query(MonthlyLedger).filter_by(vendor_id=vendor.id)
        }
        extra_values = {
            row.column_name: row.value for row in session.query(VendorExtraField).filter_by(vendor_id=vendor.id)
        }
        earliest_month = min(ledger_by_month) if ledger_by_month else None

        values = {}
        for col in columns:
            h, kind = col["header"], col["kind"]
            if kind == "erp_code":
                values[h] = vendor.erp_code
            elif kind == "entity":
                values[h] = vendor.entity
            elif kind == "vendor_name":
                values[h] = vendor.vendor_name
            elif kind == "opening_balance":
                # The sheet's own starting balance, reconstructed from the
                # earliest ledger month's own opening_balance — vendor.opening_balance
                # itself is redefined as the LIVE current figure (see load_excel.py's
                # own comment on that field), not this sheet's original one.
                values[h] = float(ledger_by_month[earliest_month].opening_balance) if earliest_month else None
            elif kind == "payable":
                row = ledger_by_month.get(col["month"])
                values[h] = float(row.payable) if row else None
            elif kind == "total_payable":
                values[h] = sum(float(r.payable) for r in ledger_by_month.values())
            elif kind == "payment":
                row = ledger_by_month.get(col["month"])
                values[h] = float(row.payment) if row else None
            elif kind == "total_payment":
                values[h] = sum(float(r.payment) for r in ledger_by_month.values())
            elif kind == "closing_balance":
                values[h] = float(vendor.opening_balance)  # live current == latest month's closing
            elif kind in ("week", "week_total", "assigned_week_order_legacy"):
                values[h] = None  # see module docstring's flagged ambiguity
            elif kind == "category":
                values[h] = vendor.category.value
            elif kind == "commitment_months":
                values[h] = vendor.commitment_months
            elif kind == "assigned_week":
                values[h] = vendor.assigned_week
            elif kind == "priority_tag":
                values[h] = vendor.priority_tag  # already a plain string (or None)
            elif kind == "extra":
                values[h] = extra_values.get(col["column_name"])
        vendors_out.append({"vendor_id": vendor.id, "erp_code": vendor.erp_code, "values": values})

    return {"columns": columns, "vendors": vendors_out, "extra_field_widgets": extra_widgets}
