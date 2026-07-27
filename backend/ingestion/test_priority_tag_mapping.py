"""Part A prerequisite tests (AI-mapping task): priority_tag wired as a
real, dedicated Excel-mapped column — mirrors category/commitment_months/
assigned_week exactly. Same isolation convention as test_upload.py.
Dataset-agnostic (picks real erp_codes/values from whatever's loaded,
never hardcodes "V00400" — the real master Excel here is a shared,
externally-mutable file that has already drifted more than once this
project, see memory: demo15-dataset-quirks).
"""
import os
import shutil
import sys

import openpyxl
import pytest


def _fresh_db_and_master(tmp_path):
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.ingestion.upload",
        "backend.configuration.vendor_edits",
        "backend.configuration.extra_fields",
    ):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH, load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    load(excel_path=master_path)
    return master_path


def _change_vendor_cell(src_path, dest_path, erp_code, header, new_value):
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, header)
    assert col is not None, f"column {header!r} not found"
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            ws.cell(row=row, column=col, value=new_value)
            wb.save(dest_path)
            return
    raise AssertionError(f"vendor {erp_code!r} not found")


def test_priority_tag_reads_from_dedicated_column_not_left_as_passthrough(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import Vendor, VendorExtraField
    from backend.db.session import SessionLocal
    from backend.shared.enums import VendorPriorityTag

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).order_by(Vendor.id).first()
        assert vendor is not None
        assert vendor.priority_tag is None or vendor.priority_tag in VendorPriorityTag

        # Not sitting in VendorExtraField as an unmapped passthrough column —
        # that would mean Part A's mapping isn't actually wired.
        extra = session.query(VendorExtraField).filter_by(column_name="Priority Tag").first()
        assert extra is None, "Priority Tag is still landing in VendorExtraField — Part A mapping not applied"
    finally:
        session.close()


def test_priority_tag_edit_writes_db_excel_and_audit_log(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.column_mapping import build_sheet_map, find_column
    from backend.shared.enums import ChangeSource, VendorPriorityTag

    session = SessionLocal()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id, erp_code = vendor.id, vendor.erp_code
    session.close()

    excel_copy = str(tmp_path / "edit_copy.xlsx")
    shutil.copy(master_path, excel_copy)

    old_value, new_value = update_vendor_field(vendor_id, "priority_tag", "P3", excel_path=excel_copy)
    assert new_value == VendorPriorityTag.P3

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(id=vendor_id).one()
        assert vendor.priority_tag == VendorPriorityTag.P3

        entry = session.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="priority_tag").one()
        assert entry.new_value == "P3"
        assert entry.source == ChangeSource.UI_EDIT.value
    finally:
        session.close()

    wb = openpyxl.load_workbook(excel_copy)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, "Priority Tag")
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            assert ws.cell(row=row, column=col).value == "P3"
            break
    else:
        raise AssertionError(f"{erp_code} not found when reading back the edited column")


def test_priority_tag_can_be_cleared_to_none(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    vendor_id = session.query(Vendor).order_by(Vendor.id).first().id
    session.close()

    excel_copy = str(tmp_path / "edit_copy.xlsx")
    shutil.copy(master_path, excel_copy)
    update_vendor_field(vendor_id, "priority_tag", "P1", excel_path=excel_copy)
    _, new_value = update_vendor_field(vendor_id, "priority_tag", None, excel_path=excel_copy)
    assert new_value is None

    session = SessionLocal()
    try:
        assert session.query(Vendor).filter_by(id=vendor_id).one().priority_tag is None
    finally:
        session.close()


def test_invalid_priority_tag_rejected_before_touching_db_or_excel(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id, before_tag = vendor.id, vendor.priority_tag
    session.close()

    with pytest.raises(ValueError, match="invalid priority_tag"):
        update_vendor_field(vendor_id, "priority_tag", "not_a_real_tag", excel_path=master_path)

    session = SessionLocal()
    try:
        assert session.query(Vendor).filter_by(id=vendor_id).one().priority_tag == before_tag
    finally:
        session.close()


def test_unrecognized_priority_tag_value_in_sheet_kept_previous_and_logged(tmp_path):
    """A garbage cell value must never crash the whole upload — same
    forgiving-defaults convention as the rest of load_excel.py."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.ingestion.load_excel import load

    session = SessionLocal()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    erp_code, before_tag = vendor.erp_code, vendor.priority_tag
    session.close()

    bad_path = str(tmp_path / "bad.xlsx")
    _change_vendor_cell(master_path, bad_path, erp_code, "Priority Tag", "not-a-tag")

    report = load(excel_path=bad_path)
    assert any(erp_code in note and "Priority Tag" in note for note in report["data_quality_notes"])

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code=erp_code).one()
        assert vendor.priority_tag == before_tag  # unchanged, not clobbered by the garbage value
    finally:
        session.close()
