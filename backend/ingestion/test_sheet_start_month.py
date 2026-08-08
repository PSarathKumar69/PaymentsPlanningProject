"""P1 demo-readiness task, Part A: SHEET_START_MONTH is now config-driven
(backend/ingestion/column_mapping_store.py's get_/set_sheet_start_month()),
not a bare module constant — plus the non-blocking sanity-check warning
(column_mapping.sheet_start_month_warning()) that fires when the sheet's own
month headers disagree with the configured start month. Same isolation
convention as test_priority_tag_mapping.py (dataset-agnostic, real master
sheet, fresh temp DB).
"""
import os
import shutil
import sys
from datetime import date

import openpyxl


def _fresh_db_and_master(tmp_path):
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel"):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import EXCEL_PATH, load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    load(excel_path=master_path)
    return master_path


def test_get_sheet_start_month_seeds_default_on_fresh_db(tmp_path):
    _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion import column_mapping_store
    from backend.ingestion.column_mapping import DEFAULT_SHEET_START_MONTH

    session = SessionLocal()
    try:
        assert column_mapping_store.get_sheet_start_month(session) == DEFAULT_SHEET_START_MONTH
    finally:
        session.close()


def test_set_sheet_start_month_roundtrip_normalizes_to_first_of_month(tmp_path):
    _fresh_db_and_master(tmp_path)
    from backend.db.session import SessionLocal
    from backend.ingestion import column_mapping_store

    session = SessionLocal()
    try:
        column_mapping_store.set_sheet_start_month(session, date(2024, 6, 15))
        session.commit()
        assert column_mapping_store.get_sheet_start_month(session) == date(2024, 6, 1)
    finally:
        session.close()


def test_build_sheet_map_defaults_to_the_seed_constant_when_not_given_one(tmp_path):
    """Confirms Part A didn't silently change ingestion's own numbers for
    every existing caller/test that builds a sheet map with no explicit
    sheet_start_month — same default as before this task."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.column_mapping import DEFAULT_SHEET_START_MONTH, build_sheet_map

    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    assert sheet_map.payable_cols[0][0] == DEFAULT_SHEET_START_MONTH


def test_build_sheet_map_honors_an_explicit_start_month_positions_unchanged(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    default_map = build_sheet_map(ws)
    shifted_map = build_sheet_map(ws, sheet_start_month=date(2024, 1, 1))

    assert shifted_map.payable_cols[0][0] == date(2024, 1, 1)
    # Only the calendar label shifts — column positions never do.
    assert [col for _, col in shifted_map.payable_cols] == [col for _, col in default_map.payable_cols]


def test_sheet_start_month_warning_silent_for_the_correctly_configured_sheet(tmp_path):
    """The real sheet's one confirmed typo (module docstring) must not, by
    itself, trip the warning — only a systematic disagreement should."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.column_mapping import DEFAULT_SHEET_START_MONTH, build_sheet_map, sheet_start_month_warning

    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    assert sheet_start_month_warning(ws, sheet_map, DEFAULT_SHEET_START_MONTH) is None


def test_sheet_start_month_warning_fires_when_configured_start_month_is_wrong(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.ingestion.column_mapping import build_sheet_map, sheet_start_month_warning

    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    wrong_start = date(2020, 1, 1)
    sheet_map = build_sheet_map(ws, sheet_start_month=wrong_start)
    warning = sheet_start_month_warning(ws, sheet_map, wrong_start)
    assert warning is not None
    assert "2020" in warning


if __name__ == "__main__":
    import tempfile
    from pathlib import Path

    tmp = Path(tempfile.mkdtemp())
    test_get_sheet_start_month_seeds_default_on_fresh_db(tmp)
    test_sheet_start_month_warning_fires_when_configured_start_month_is_wrong(Path(tempfile.mkdtemp()))
    print("OK")
