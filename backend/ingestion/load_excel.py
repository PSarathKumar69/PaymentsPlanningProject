"""Ingestion: load data/Vendor's Details.xlsx into the SQLite DB.

Populates vendors + monthly_ledger (one row per vendor per month) and one
plan_runs/plan_allocations snapshot from the sheet's existing W1-W5 +
combined assigned-week/order column. Then verifies ledger integrity and
the W1-W5-vs-Total reconciliation, logging (not fixing) any mismatches.

Safe to re-run: docs/07 confirms the master Excel is re-loaded/refreshed at
the start of each month, so this is an upsert by (entity, erp_code) —
Vendor's real uniqueness key, entity-code-collision fix (update the
existing vendor + replace their monthly_ledger rows from the fresh sheet),
never a blind re-insert.
Finance's edits to category/commitment_months/assigned_week (vendor_edits.py)
are read back from the SAME dedicated columns they were written to, not
recomputed/reset to a default — CLAUDE.md rule 6: the Excel is the single
source of truth precisely because edits are written back into it.

Run with: python -m backend.ingestion.load_excel
"""
import logging
from datetime import datetime

import openpyxl

from backend.db.models import AuditLog, MonthlyLedger, PlanAllocation, PlanRun, Vendor, VendorExtraField
from backend.db.session import SessionLocal, init_db
from backend.ingestion import column_mapping, column_mapping_store
from backend.ingestion.column_mapping import (
    CATEGORY_EXCEL_LABEL,
    FIELD_TO_EXCEL_HEADER,
    build_sheet_map,
    category_from_label,
    derive_category_and_priority_tag,
    valid_priority_tag_values,
    find_column,
    find_duplicate_erp_codes,
    normalize_entity_text,
    parse_assigned_week_order,
    parse_week_number,
    resolve_header,
    to_number,
    unmapped_header_columns,
)
from backend.shared.constants import MONEY_EPSILON
from backend.shared.enums import ChangeSource, VendorCategory

EXCEL_PATH = "data/Vendor's Details.xlsx"

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("ingestion")


def load(excel_path=EXCEL_PATH, session=None, header_overrides=None, sheet_start_month=None, header_row=None, data_start_row=None):
    """session: optional externally-supplied session — e.g.
    backend/month_end/rollover.py passes its own session so re-ingestion and
    its subsequent payment-cycle reset commit together as one transaction.
    Defaults to owning its own session/commit, unchanged for every other
    existing caller (same owns_session convention already used by
    vendor_edits.py/payment_logging.py/planner.py/regeneration.py).

    header_overrides: optional {logical_field: header_text} (AI-mapping
    task) — a persisted AI-resolved header name, consulted before each
    field's fixed default. None (the default) reproduces this function's
    exact pre-AI-mapping behavior; every existing caller that doesn't pass
    this gets identical results to before.

    sheet_start_month: optional explicit override (P1 demo-readiness task,
    CLAUDE.md rule 7) — None (every existing caller) resolves the live
    `config` value itself (column_mapping_store.get_sheet_start_month()),
    seeding today's default on first read. A caller that already looked it
    up for its own purposes (e.g. upload.py's commit_upload(), which also
    needs it for the month-header sanity check) passes it through explicitly
    so both uses agree on the exact same value within one upload.

    header_row/data_start_row: the resolved header/data-start row (this
    task, never a fixed row number — see column_mapping.py's module
    docstring). None (every caller with no AI adjudication available —
    revert_upload(), tests, direct script runs) lets build_sheet_map()
    resolve them itself via column_mapping.resolve_header_row()
    (deterministic-only). upload.py's commit_upload() passes both
    explicitly, already reconciled with the AI's confirmation/override.
    """
    init_db()
    owns_session = session is None
    session = session or SessionLocal()
    if sheet_start_month is None:
        sheet_start_month = column_mapping_store.get_sheet_start_month(session)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    sheet_map = build_sheet_map(
        ws,
        header_overrides=header_overrides,
        sheet_start_month=sheet_start_month,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    column_mapping.validate_sheet_layout(ws, sheet_map)  # Part 5 backstop — before any DB write below

    # Found once, not created — a re-ingestion before any Finance edit has
    # ever happened must behave exactly as it did before these columns
    # existed (find_column never mutates the workbook).
    category_col = find_column(ws, resolve_header("category", header_overrides), sheet_map.header_row)
    commitment_months_col = find_column(ws, resolve_header("commitment_months", header_overrides), sheet_map.header_row)
    assigned_week_override_col = find_column(ws, resolve_header("assigned_week", header_overrides), sheet_map.header_row)
    priority_tag_col = find_column(ws, resolve_header("priority_tag", header_overrides), sheet_map.header_row)

    # Generic passthrough columns (data-pipeline-upload task, docs/11):
    # anything the mapping layer above doesn't recognize gets stored per
    # vendor as-is, never fed into any model (CLAUDE.md rule 3). Computed
    # once, walked per vendor in the same pass below — not a second parser.
    unmapped_columns = unmapped_header_columns(
        ws,
        sheet_map,
        extra_cols=(category_col, commitment_months_col, assigned_week_override_col, priority_tag_col),
    )

    # Prefetched once (not per-vendor-per-column) so an existing extra-field
    # value is updated in place rather than duplicated — a brand-new vendor
    # simply has no entry here yet, which is fine, it just gets a fresh row.
    existing_extra_fields = {(r.vendor_id, r.column_name): r for r in session.query(VendorExtraField).all()}

    ledger_mismatches = []
    week_total_mismatches = []
    data_quality_notes = []
    category_priority_conflicts = []
    erp_codes_in_file = []
    # Entity-code-collision fix: erp_codes_in_file (above) stays bare
    # erp_code strings — it's part of MasterDataCommitResponse's public API
    # contract (list[str]) via upload.py's diff. vendor_keys_in_file is the
    # internal (entity, erp_code) counterpart, used for the soft-deactivate
    # membership check below and by upload.py's own diff — never exposed
    # through the API schema (undeclared keys are silently dropped by
    # FastAPI's response_model, same as every other internal-only report key).
    vendor_keys_in_file = []
    category_present_count = 0
    category_unrecognized_erp_codes = []

    # Duplicate ERP code detection (confirmed real bug, 2026-07-31
    # investigation; first-row-wins per Sarath's later call, superseding
    # the original last-row-wins decision): done once, up front, over the
    # whole erp_code column — not folded into the per-row loop below, which
    # upserts by (entity, erp_code) and would otherwise let a later row
    # silently overwrite an earlier one with zero trace.
    #
    # Entity-code-collision fix: find_duplicate_erp_codes() now groups by
    # (entity, erp_code), not bare erp_code — 28 of the 29 groups this used
    # to flag were NEVER real duplicates at all, just two different legal
    # entities (ARS/FP) reusing the same bare numeric code (disambiguated
    # by the sheet's own "Unique" column); those now ingest as two separate
    # vendor rows and never appear here. What's left is a genuine
    # same-entity data-entry collision (confirmed real case: ARS "INC") —
    # still first-row-wins, still loud, but Finance needs to assign the
    # skipped vendor(s) a real, distinct ERP code; this fix doesn't resolve
    # that automatically.
    duplicate_erp_code_notes = []
    duplicate_groups = find_duplicate_erp_codes(ws, sheet_map)
    duplicate_skip_rows = {r["row"] for g in duplicate_groups for r in g["rows"][1:]}
    if duplicate_groups:
        if len(duplicate_groups) > 5:
            # Whole-sheet aggregation (same principle as the
            # whole-column-unrecognized-category note below) — one summary
            # rather than one line per duplicate when there are many.
            duplicate_erp_code_notes.append(
                f"{len(duplicate_groups)} ERP codes each appear more than once for the SAME entity in the "
                "sheet — a genuine data-entry collision, not cross-entity code reuse. For each, only the FIRST "
                "row is kept and the later row(s) are skipped entirely; Finance needs to assign the skipped "
                "vendor(s) a distinct ERP code: " + ", ".join(f"{g['erp_code']} ({g['entity']})" for g in duplicate_groups)
            )
        else:
            for group in duplicate_groups:
                code, entity, rows = group["erp_code"], group["entity"], group["rows"]
                winner = rows[0]
                skipped = rows[1:]
                duplicate_erp_code_notes.append(
                    f"{code} ({entity}): same entity appears in {len(rows)} rows (rows {[r['row'] for r in rows]}) "
                    f"— a genuine ERP-code collision, not cross-entity code reuse. Only row {winner['row']} "
                    f"({winner['vendor_name']!r}) was kept; row(s) {[r['row'] for r in skipped]} "
                    f"({[r['vendor_name'] for r in skipped]}) were skipped entirely — Finance needs to assign "
                    "a distinct ERP code before this vendor can be tracked."
                )
        data_quality_notes.extend(duplicate_erp_code_notes)

    plan_run = PlanRun(
        created_at=datetime.now(),
        month=sheet_map.payable_cols[-1][0],
        model_used=ChangeSource.INGESTION.value,
        funds_figure=None,
    )
    session.add(plan_run)
    session.flush()  # get plan_run.id

    vendor_count = 0
    ledger_row_count = 0
    allocation_count = 0

    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        erp_code = ws.cell(row=row, column=sheet_map.erp_code_col).value
        if erp_code is None:
            continue  # trailing blank row
        if row in duplicate_skip_rows:
            continue  # later occurrence of an already-seen (entity, erp_code) pair — first row wins, this one is ignored entirely
        entity = ws.cell(row=row, column=sheet_map.entity_col).value
        erp_codes_in_file.append(erp_code)
        vendor_keys_in_file.append((entity, erp_code))

        vendor_name = ws.cell(row=row, column=sheet_map.vendor_name_col).value

        raw_opening = ws.cell(row=row, column=sheet_map.opening_balance_col).value
        if raw_opening is None:
            data_quality_notes.append(f"{erp_code}: blank Opening Balance, treated as 0")
        opening_balance = to_number(raw_opening)

        sheet_total_payable = to_number(ws.cell(row=row, column=sheet_map.total_payable_col).value)
        sheet_total_payment = to_number(ws.cell(row=row, column=sheet_map.total_payment_col).value)
        sheet_closing = to_number(ws.cell(row=row, column=sheet_map.closing_balance_col).value)

        raw_assigned_order = ws.cell(row=row, column=sheet_map.assigned_week_order_col).value
        try:
            legacy_assigned_week, within_week_order = parse_assigned_week_order(raw_assigned_order)
        except ValueError:
            # Same defensive-fallback pattern as priority_tag below: a sheet
            # with no legacy packed "W2-P3" column at all can still resolve
            # assigned_week_order_col to some unrelated real column by
            # position (e.g. a plain "W2" assigned-week tag) — log it and
            # fall back rather than crash the whole upload over it.
            data_quality_notes.append(
                f"{erp_code}: unrecognized Assigned Week Order value {raw_assigned_order!r}, kept previous value"
            )
            legacy_assigned_week, within_week_order = None, None

        # Entity-code-collision fix: erp_code alone is no longer unique
        # (models.py) — 28 real cases are two different entities (ARS/FP)
        # sharing one bare code, so the upsert lookup must match both. A
        # plain compound filter_by(erp_code=, entity=) would miss a
        # whitespace/case-only difference between the sheet and the stored
        # value, silently inserting a duplicate row instead of updating the
        # existing one — matched in Python via normalize_entity_text()
        # instead, same as find_duplicate_erp_codes()'s own grouping.
        existing = None
        for candidate in session.query(Vendor).filter_by(erp_code=erp_code).all():
            if normalize_entity_text(candidate.entity) == normalize_entity_text(entity):
                existing = candidate
                break

        # category/priority_tag (this task): two views of the same Finance
        # signal, reconciled by derive_category_and_priority_tag() below once
        # BOTH raw cells are read (priority_tag_raw further down) — a blank
        # cell still unambiguously means "never edited" (see
        # commitment_months's own comment below for why), same fallback
        # convention as every other Finance-edit field here.
        category_raw = ws.cell(row=row, column=category_col).value if category_col else None
        category_label = category_from_label(category_raw, session=session)
        if category_raw is not None:
            category_present_count += 1
            if category_label is None:
                # Doesn't match Must Pay/Commitment/Normal/Inactive under any
                # accepted variant (e.g. the live sheet's actual Category
                # column: "Vendor"/"Non Vendor", a different vocabulary
                # entirely) — aggregated into one note after the vendor loop
                # rather than one line per vendor if the WHOLE column is like
                # this (see after the loop, below).
                category_unrecognized_erp_codes.append(erp_code)

        commitment_months_raw = (
            ws.cell(row=row, column=commitment_months_col).value if commitment_months_col else None
        )
        if commitment_months_raw is not None:
            commitment_months = int(to_number(commitment_months_raw))
        elif existing is not None:
            commitment_months = existing.commitment_months
        else:
            commitment_months = None  # let the model's own column default (1) apply on insert

        # assigned_week: the dedicated column (Finance's UI edits) wins over
        # the legacy packed column whenever either has a value for this
        # vendor; if BOTH are blank, fall back to the existing DB value —
        # same pattern as category/commitment_months above. This isn't just
        # a one-off safety net for a single bad load: the legacy column is a
        # VLOOKUP into an external workbook this repo doesn't have, so it
        # reads as blank on EVERY future load, permanently. This fallback is
        # therefore the permanent mechanism that keeps assigned_week intact
        # across every future re-ingestion/rollover, not a stopgap.
        assigned_week_override_raw = (
            ws.cell(row=row, column=assigned_week_override_col).value if assigned_week_override_col else None
        )
        if assigned_week_override_raw is not None:
            sheet_assigned_week = parse_week_number(assigned_week_override_raw)
            if sheet_assigned_week is None:
                # Column-mapping resolved assigned_week to a real column, but
                # this particular value is neither a bare int nor "W<n>" —
                # same defensive-fallback pattern as priority_tag/
                # assigned_week_order_col above: log it, fall back rather
                # than crash the whole upload over one cell.
                data_quality_notes.append(
                    f"{erp_code}: unrecognized Assigned Week value {assigned_week_override_raw!r}, "
                    "ignored (falling back to the legacy column/existing value)"
                )
                sheet_assigned_week = legacy_assigned_week
        else:
            sheet_assigned_week = legacy_assigned_week
        if sheet_assigned_week is not None:
            assigned_week = sheet_assigned_week
        elif existing is not None:
            assigned_week = existing.assigned_week
        else:
            assigned_week = None  # brand-new vendor, sheet gave nothing to seed from

        # priority_tag: same blank-means-never-edited fallback as category
        # above — read here, then reconciled WITH category_label in one shot
        # via derive_category_and_priority_tag() (this task). Vendor.priority_tag
        # is a plain String column — tag_value is already the exact validated
        # value, no need to round-trip through the enum.
        priority_tag_raw = ws.cell(row=row, column=priority_tag_col).value if priority_tag_col else None
        tag_value = None
        if priority_tag_raw is not None:
            candidate = str(priority_tag_raw).strip().upper()
            # valid_priority_tag_values(session), not the old fixed P0-P5
            # set — a custom tag Finance added via Configuration (and wrote
            # back to this exact cell through the UI edit's Excel
            # write-back) must still round-trip correctly on re-ingestion.
            if candidate in valid_priority_tag_values(session):
                tag_value = candidate

        derived_category, derived_tag, conflict = derive_category_and_priority_tag(category_label, tag_value, session=session)

        if conflict is not None:
            conflict_tag, conflict_category = conflict
            note = (
                f"{erp_code}: Category and Priority Tag disagree — tag column said {conflict_tag!r}, category "
                f"column said {category_raw!r} (-> {CATEGORY_EXCEL_LABEL.get(conflict_category, conflict_category)!r}); category label "
                f"wins, using category={CATEGORY_EXCEL_LABEL.get(derived_category, derived_category)!r}/priority_tag={derived_tag!r}"
            )
            data_quality_notes.append(note)
            category_priority_conflicts.append(note)

        if priority_tag_raw is not None and tag_value is None:
            # Unrecognized tag cell — derive from category if we can, else
            # (same as before this task) log and keep whatever's already
            # there rather than guessing.
            if category_label is not None:
                data_quality_notes.append(
                    f"{erp_code}: unrecognized Priority Tag {priority_tag_raw!r}, derived {derived_tag!r} from "
                    f"Category ({category_raw!r}) instead"
                )
            else:
                data_quality_notes.append(
                    f"{erp_code}: unrecognized Priority Tag {priority_tag_raw!r}, kept previous value"
                )

        if derived_category is not None:
            category = derived_category
        elif existing is not None:
            category = existing.category
        else:
            category = VendorCategory.NORMAL

        if derived_tag is not None:
            priority_tag = derived_tag
        elif existing is not None:
            priority_tag = existing.priority_tag
        else:
            priority_tag = None

        if existing is not None:
            vendor = existing
            vendor.entity = entity
            vendor.vendor_name = vendor_name
            # opening_balance is set below, after the ledger walk recomputes
            # it — never from sheet_closing directly (see that assignment
            # for why).
            vendor.assigned_week = assigned_week
            vendor.category = category
            vendor.priority_tag = priority_tag
            # Present in this sheet -> live, whether or not a prior upload
            # had deactivated it (Bug #1 fix, models.py's is_active comment).
            vendor.is_active = True
            vendor.removed_at = None
            if commitment_months is not None:
                vendor.commitment_months = commitment_months
            # Ledger figures are sheet-derived, never Finance-edited — safe
            # to drop and rebuild rather than diff month-by-month.
            session.query(MonthlyLedger).filter_by(vendor_id=vendor.id).delete()
        else:
            vendor = Vendor(
                erp_code=erp_code,
                entity=entity,
                vendor_name=vendor_name,
                opening_balance=0.0,  # placeholder — overwritten below, after the ledger walk recomputes it
                assigned_week=assigned_week,
                category=category,
                priority_tag=priority_tag,
                reconsider=None,
                score=None,
                current_aging_bucket=None,
            )
            if commitment_months is not None:
                vendor.commitment_months = commitment_months
            session.add(vendor)
        session.flush()  # get vendor.id
        vendor_count += 1

        if conflict is not None:
            conflict_tag, conflict_category = conflict
            session.add(
                AuditLog(
                    timestamp=datetime.now(),
                    vendor_id=vendor.id,
                    field_name="category_priority_conflict",
                    old_value=f"tag={conflict_tag}",
                    new_value=f"category={CATEGORY_EXCEL_LABEL.get(conflict_category, conflict_category)} (won); priority_tag={derived_tag}",
                    source=ChangeSource.INGESTION.value,
                )
            )

        touched_extra_columns = set()
        for col_idx, header_text in unmapped_columns:
            raw = ws.cell(row=row, column=col_idx).value
            value_text = None if raw is None else str(raw).strip()
            key = (vendor.id, header_text)
            extra_row = existing_extra_fields.get(key)
            if extra_row is not None:
                extra_row.value = value_text
            else:
                extra_row = VendorExtraField(vendor_id=vendor.id, column_name=header_text, value=value_text)
                session.add(extra_row)
                existing_extra_fields[key] = extra_row
            touched_extra_columns.add(header_text)

        # A column that existed on a PRIOR load but is absent from this
        # sheet entirely (e.g. revert to a backup taken before that column
        # existed) is dropped for this vendor — same "sheet-derived, rebuilt
        # every load" convention monthly_ledger already follows above, not
        # a diff/merge against whatever used to be there.
        for (vid, column_name), extra_row in list(existing_extra_fields.items()):
            if vid == vendor.id and column_name not in touched_extra_columns:
                session.delete(extra_row)
                del existing_extra_fields[(vid, column_name)]

        # Walk the monthly columns to build the per-month ledger, then
        # verify against the sheet's own reported totals/closing balance.
        running_opening = opening_balance
        sum_payable = 0.0
        sum_payment = 0.0
        for (month, payable_col), (_, payment_col) in zip(sheet_map.payable_cols, sheet_map.payment_cols):
            payable = to_number(ws.cell(row=row, column=payable_col).value)
            payment = to_number(ws.cell(row=row, column=payment_col).value)
            if isinstance(ws.cell(row=row, column=payment_col).value, str):
                data_quality_notes.append(
                    f"{erp_code}: non-numeric payment cell for {month:%b-%y}, treated as 0"
                )
            if payment < 0:
                # Confirmed with Finance: a negative Payment cell means an
                # over-payment (beyond what was billed) recorded with the
                # wrong sign, not extra debt — treat it as a real payment
                # of that magnitude.
                data_quality_notes.append(
                    f"{erp_code}: negative payment cell for {month:%b-%y} (₹{abs(payment):,.2f}), "
                    "treated as a payment of that amount"
                )
                payment = abs(payment)
            closing = running_opening + payable - payment
            session.add(
                MonthlyLedger(
                    vendor_id=vendor.id,
                    month=month,
                    payable=payable,
                    payment=payment,
                    opening_balance=running_opening,
                    closing_balance=closing,
                )
            )
            ledger_row_count += 1
            sum_payable += payable
            sum_payment += payment
            running_opening = closing

        # Current balance = latest recomputed closing, carried forward.
        # Deliberately NOT sheet_closing: that's a formula cell
        # (Closing Balance = Op. Balance + Total Payable - Total Payment),
        # and openpyxl's load-modify-save cycle (vendor_edits.py's Excel
        # write-back) silently drops every formula cell's cached value
        # workbook-wide on save — not just the cell that was edited. This
        # figure is computed purely from the raw (literal, never-a-formula)
        # Payable/Payment columns walked above, so it can never be affected
        # by that bug. The comparison below is kept as a sanity-check signal
        # only, never as the value that actually lands in the DB.
        vendor.opening_balance = running_opening

        if abs(running_opening - sheet_closing) > MONEY_EPSILON:
            ledger_mismatches.append((erp_code, running_opening, sheet_closing))
        if abs(sum_payable - sheet_total_payable) > MONEY_EPSILON:
            data_quality_notes.append(
                f"{erp_code}: recomputed Total Payable {sum_payable:.2f} != sheet's {sheet_total_payable:.2f}"
            )
        if abs(sum_payment - sheet_total_payment) > MONEY_EPSILON:
            data_quality_notes.append(
                f"{erp_code}: recomputed Total Payment {sum_payment:.2f} != sheet's {sheet_total_payment:.2f}"
            )

        # W1-W5 breakdown -> plan_allocations, one row per non-zero week.
        # A vendor can have money in more than one week in the same cycle
        # (e.g. a partial payment topped up later) — see docs/06. Optional
        # (docs/07): a fresh sheet with no week block yet has nothing to
        # seed here — no history to compare, nothing to fall out of sync.
        if sheet_map.week_cols:
            week_values = [
                (week_no, to_number(ws.cell(row=row, column=col).value))
                for week_no, col in sheet_map.week_cols
            ]
            sheet_week_total = to_number(ws.cell(row=row, column=sheet_map.week_total_col).value)
            actual_week_sum = sum(v for _, v in week_values)
            if abs(actual_week_sum - sheet_week_total) > MONEY_EPSILON:
                week_total_mismatches.append((erp_code, actual_week_sum, sheet_week_total))

            for week_no, amount in week_values:
                if amount == 0:
                    continue
                session.add(
                    PlanAllocation(
                        plan_run_id=plan_run.id,
                        vendor_id=vendor.id,
                        assigned_week=week_no,
                        within_week_order=within_week_order,
                        allocated_amount=amount,
                    )
                )
                allocation_count += 1

    # Replace, not merge (Bug #1 fix): any currently-active vendor whose
    # erp_code didn't turn up anywhere in this sheet is soft-deactivated —
    # see models.py's Vendor.is_active comment for why not a hard delete.
    # Their ledger/extra-field rows are sheet-derived and safe to drop, same
    # as the existing-vendor branch above already does for vendors still
    # present.
    # Category column, whole-column-unrecognized case (this task): one
    # aggregated note rather than one line per vendor when literally every
    # row with a Category value failed to match any known label (e.g. the
    # live sheet's actual Category column: "Vendor"/"Non Vendor", an
    # entity-type flag, not Finance's vocabulary at all). A MIXED case
    # (some rows recognized, some not) still gets one note per vendor —
    # that's a genuine per-row data issue, not "this whole column means
    # something else."
    if category_unrecognized_erp_codes:
        if len(category_unrecognized_erp_codes) == category_present_count:
            data_quality_notes.append(
                f"Category column values don't match any known label for {category_present_count} vendors, "
                "ignoring it"
            )
        else:
            for code in category_unrecognized_erp_codes:
                data_quality_notes.append(f"{code}: unrecognized Category value, ignoring it for this vendor")

    # Entity-code-collision fix: membership keyed by (normalized entity,
    # erp_code), not bare erp_code — else a currently-active vendor whose
    # code is shared with a DIFFERENT entity elsewhere in the sheet would
    # look "present" (bare code matches) even if its own (entity, erp_code)
    # row is actually the one missing, and vice versa.
    erp_codes_in_file_set = {(normalize_entity_text(e), c) for e, c in vendor_keys_in_file}
    deactivated_erp_codes = []
    for vendor in session.query(Vendor).filter(Vendor.is_active.isnot(False)).all():
        if (normalize_entity_text(vendor.entity), vendor.erp_code) not in erp_codes_in_file_set:
            vendor.is_active = False
            vendor.removed_at = datetime.now()
            session.query(MonthlyLedger).filter_by(vendor_id=vendor.id).delete()
            session.query(VendorExtraField).filter_by(vendor_id=vendor.id).delete()
            deactivated_erp_codes.append(vendor.erp_code)

    if owns_session:
        session.commit()
        session.close()

    return {
        "vendor_count": vendor_count,
        "ledger_row_count": ledger_row_count,
        "allocation_count": allocation_count,
        "ledger_mismatches": ledger_mismatches,
        "week_total_mismatches": week_total_mismatches,
        "data_quality_notes": data_quality_notes,
        "category_priority_conflicts": category_priority_conflicts,
        "duplicate_erp_code_notes": duplicate_erp_code_notes,
        "erp_codes_in_file": erp_codes_in_file,
        "vendor_keys_in_file": vendor_keys_in_file,
        "deactivated_vendor_count": len(deactivated_erp_codes),
        "deactivated_erp_codes": deactivated_erp_codes,
        "unmapped_columns": [header for _, header in unmapped_columns],
    }


def print_report(report):
    log.info("Vendors loaded: %d", report["vendor_count"])
    log.info("Monthly ledger rows loaded: %d", report["ledger_row_count"])
    log.info("Plan allocation rows loaded: %d", report["allocation_count"])

    if report["ledger_mismatches"]:
        log.warning("Ledger-integrity mismatches (Closing != Opening + Payable - Payment):")
        for erp_code, computed, expected in report["ledger_mismatches"]:
            log.warning("  %s: computed=%.2f, sheet=%.2f", erp_code, computed, expected)
    else:
        log.info("Ledger integrity: OK for all vendors.")

    if report["week_total_mismatches"]:
        log.warning("W1-W5 sum vs sheet's Total column mismatches (known entry-error pattern):")
        for erp_code, computed, expected in report["week_total_mismatches"]:
            log.warning("  %s: W1-W5 sum=%.2f, sheet Total=%.2f", erp_code, computed, expected)
    else:
        log.info("W1-W5 vs Total: OK for all vendors.")

    if report["data_quality_notes"]:
        log.warning("Other data-quality notes:")
        for note in report["data_quality_notes"]:
            log.warning("  %s", note)


if __name__ == "__main__":
    print_report(load())
