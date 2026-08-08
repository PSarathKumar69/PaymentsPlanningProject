"""GET /models/5/min-funds-verification-export ("Verify Min Funds" download
task) — router-level test. Same fresh-temp-DB isolation convention as
test_new_model_2_export.py (duplicated here per this project's existing
per-file-fixture convention).
"""
import sys

import os
import openpyxl
import pytest
import pandas as pd
from datetime import date
from io import BytesIO
from fastapi.testclient import TestClient


def _fresh_app(tmp_path):
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for name in list(sys.modules):
        if name == "backend" or name.startswith("backend."):
            del sys.modules[name]
    from backend.api.main import app

    return app


@pytest.fixture
def client(tmp_path):
    return TestClient(_fresh_app(tmp_path))


def test_export_400_before_any_planning_month_resolvable(client):
    resp = client.get("/models/5/min-funds-verification-export")
    assert resp.status_code == 400


def test_export_aging_columns_are_dynamic_and_zero_vs_blank_is_correct(client):
    """One vendor whose oldest debt goes 5 months back ("151-180", with a
    real billing gap in between that must surface as 0.0 not be skipped),
    one vendor with a short 1-month history ("31-60" at most) — confirms
    the column set is the union across both (not a hardcoded 5), and that
    "beyond this vendor's own oldest month" (blank) is never confused with
    "this vendor's own month, genuinely zero" (0.0)."""
    from backend.db.session import SessionLocal
    from backend.db.models import MonthlyLedger, Vendor
    from backend.shared.min_funds_v2 import required_amount_v2
    from backend.shared.planning_month_store import set_current_planning_month

    session = SessionLocal()
    vendor_a = Vendor(erp_code="A1", entity="E", vendor_name="Old Debt Co", opening_balance=580)
    vendor_b = Vendor(erp_code="B1", entity="E", vendor_name="Recent Co", opening_balance=250)
    session.add_all([vendor_a, vendor_b])
    session.flush()

    as_of = date(2026, 6, 1)  # planning month 2026-07 minus one calendar month

    def month_for(months_back):
        total = as_of.year * 12 + (as_of.month - 1) - months_back
        year, month0 = divmod(total, 12)
        return date(year, month0 + 1, 1)

    # Vendor A: real ledger rows only at 5 months back and current (0) —
    # months 1-4 back have no MonthlyLedger row at all (a genuine billing
    # gap), which monthly_breakdown must still report as 0.0.
    session.add_all([
        MonthlyLedger(vendor_id=vendor_a.id, month=month_for(5), payable=500, payment=0, opening_balance=0, closing_balance=500),
        MonthlyLedger(vendor_id=vendor_a.id, month=month_for(0), payable=80, payment=0, opening_balance=0, closing_balance=80),
    ])
    # Vendor B: short history — oldest only 1 month back.
    session.add_all([
        MonthlyLedger(vendor_id=vendor_b.id, month=month_for(1), payable=200, payment=0, opening_balance=0, closing_balance=200),
        MonthlyLedger(vendor_id=vendor_b.id, month=month_for(0), payable=50, payment=0, opening_balance=0, closing_balance=50),
    ])
    set_current_planning_month(session, "2026-07")
    session.commit()

    ledger_a = session.query(MonthlyLedger).filter_by(vendor_id=vendor_a.id).order_by(MonthlyLedger.month).all()
    ledger_b = session.query(MonthlyLedger).filter_by(vendor_id=vendor_b.id).order_by(MonthlyLedger.month).all()
    min_funds_a, _ = required_amount_v2(vendor_a, ledger_a, as_of=as_of)
    min_funds_b, _ = required_amount_v2(vendor_b, ledger_b, as_of=as_of)
    session.close()

    resp = client.get("/models/5/min-funds-verification-export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    df = pd.read_excel(BytesIO(resp.content))
    # The full point of this fix: dynamic, sized to the single oldest
    # vendor in the file (here vendor A's "151-180"), not a hardcoded 5.
    assert list(df.columns) == [
        "Vendor Name", "Outstanding",
        "0-30", "31-60", "61-90", "91-120", "121-150", "151-180",
        "Min Funds Required",
    ]

    row_a = df[df["Vendor Name"] == "Old Debt Co"].iloc[0]
    assert row_a["Outstanding"] == pytest.approx(580)
    assert row_a["0-30"] == pytest.approx(80)
    assert row_a["31-60"] == pytest.approx(0)      # gap month, applicable, genuinely zero
    assert row_a["61-90"] == pytest.approx(0)
    assert row_a["91-120"] == pytest.approx(0)
    assert row_a["121-150"] == pytest.approx(0)
    assert row_a["151-180"] == pytest.approx(500)
    assert row_a["Min Funds Required"] == pytest.approx(min_funds_a)

    row_b = df[df["Vendor Name"] == "Recent Co"].iloc[0]
    assert row_b["Outstanding"] == pytest.approx(250)
    assert row_b["0-30"] == pytest.approx(50)
    assert row_b["31-60"] == pytest.approx(200)
    assert pd.isna(row_b["61-90"])                 # beyond vendor B's own oldest month: blank, not 0
    assert pd.isna(row_b["91-120"])
    assert pd.isna(row_b["121-150"])
    assert pd.isna(row_b["151-180"])
    assert row_b["Min Funds Required"] == pytest.approx(min_funds_b)

    # Exceptions tab (this task): a second WORKSHEET, never extra rows on
    # the sheet pd.read_excel() just parsed above — this is the actual
    # regression proof that appending it can't corrupt the main sheet's
    # own column dtypes (df's numeric columns are still all numeric,
    # confirmed by every pytest.approx() call above succeeding at all).
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["Sheet1", "Exceptions"]
    assert wb["Sheet1"].max_row == 3  # header + vendor_a + vendor_b, untouched by the new tab
    assert wb.active.title == "Sheet1"  # unchanged — Exceptions is never made the active sheet
