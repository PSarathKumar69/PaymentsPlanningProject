"""New Model 2 router — priority-bucket ceiling allocation (docs/14-new-model-2.md).

Additive alongside Models 1/2/3 and New Model 1 (this build's own
guardrail — see CLAUDE.md: none of those get touched). Mirrors
backend/api/routers/new_model_1.py almost exactly — same override/locking
mechanism (an overridden vendor is excluded from this round's computation
via eligible_vendor_ids, and their override amount is subtracted from
available_funds up front), same frozen-Suggested-figure-on-lock pattern.
See new_model_1.py's own docstrings for the full rationale; not
re-explained here.

Planning month (docs/14, new): New Model 2 needs Finance's own explicitly-
picked planning month (month + year, no day) before anything else — never
_current_cycle_month() (backend/shared/payment_logging.py), the ledger-
ingestion-anchored value every other model still uses unchanged. See
_resolve_planning_month()/_month_minus_one() below.
"""
import os
import tempfile
from datetime import date, datetime
from typing import Any

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy import func
from starlette.background import BackgroundTask

from backend.configuration.vendor_edits import _find_or_create_column, _find_vendor_row
from backend.db.models import AuditLog, MonthlyLedger, PlanAllocation, PlanRun, Vendor
from backend.db.session import SessionLocal
from backend.ingestion import column_mapping_store
from backend.ingestion.column_mapping import build_sheet_map
from backend.ingestion.load_excel import EXCEL_PATH, sync_excel_path_from_db
from backend.models.new_model_2.allocator import generate_plan
from backend.month_end.rollover import _reset_vendor_cycle_state
from backend.shared.aging import compute_vendor_aging
from backend.shared.constants import MONEY_EPSILON, NEW_MODEL_2_PLAN_RUN_LABEL
from backend.shared.enums import AllocationStatus, ChangeSource
from backend.shared.export_exceptions import build_exceptions, write_exceptions_sheet
from backend.shared.min_funds import live_outstanding_balance
from backend.shared.min_funds_v2 import (
    calculate_minimum_funds_required_v2,
    min_funds_breakdown_v2,
    required_amount_v2,
)
from backend.shared.payment_logging import finalize_plan
from backend.shared.plan_history import build_plan_run_history_response, latest_week_distribution_for_vendor
from backend.shared.planning_month_store import get_current_planning_month
from backend.weekly_planning.planner import build_weekly_view

from ..schemas.new_model_2 import (
    AllVendorMinFundsRequiredV2Response,
    CurrentPlanningMonthResponse,
    FinalizeCheckResponse,
    NewModel2GeneratePlanRequest,
    NewModel2MinimumFundsRequiredResponse,
    NewModel2PlanAndWeeklyViewResponse,
    ResetCycleResponse,
    SuggestedPlanningMonthResponse,
    VendorMinFundsRequiredV2Response,
)
from ..schemas.plan_runs import PlanRunHistoryResponse

router = APIRouter(tags=["new_model_2"])


def _plan_response(result: dict) -> dict:
    return {**result, "allocations": result["allocations"].to_dict(orient="records")}


def _parse_planning_month(value: str) -> date:
    """"YYYY-MM" (docs/14: month + year, no day) -> the first-of-month date
    PlanRun.month/aging as_of both use. Raises ValueError on a bad format —
    callers turn that into an HTTP 400."""
    try:
        year_str, month_str = value.split("-")
        return date(int(year_str), int(month_str), 1)
    except (ValueError, AttributeError):
        raise ValueError(f"planning_month must be 'YYYY-MM', got {value!r}") from None


def _month_minus_one(d: date) -> date:
    """The calendar month immediately before `d` (both as first-of-month
    dates) — docs/14: "the aging reference point is planning month minus
    one calendar month"."""
    if d.month == 1:
        return date(d.year - 1, 12, 1)
    return date(d.year, d.month - 1, 1)


def _month_plus_one(d: date) -> date:
    """The calendar month immediately after `d` (both as first-of-month
    dates) — the inverse of _month_minus_one(), used to derive the
    suggested planning month from the sheet's real last recorded ledger
    month (this task: "if the sheet's data goes up to April, Finance plans
    for May")."""
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _last_recorded_ledger_month(session) -> date | None:
    """The sheet's real last recorded ledger month — the global max across
    every vendor's monthly_ledger, not per-vendor (planning month is one
    shared value for the whole cycle, not per-vendor). None only when the
    DB has no ledger data loaded at all yet."""
    return session.query(func.max(MonthlyLedger.month)).scalar()


def _resolve_planning_month(session, requested: date | None) -> date:
    """docs/14: required on the first Generate Plan of a cycle, optional
    (reused) on every regeneration after that. Fallback order (Main-tab
    upload-confirm task, section 3): (1) explicitly requested, (2) latest
    New Model 2 plan_run this cycle, (3) the persisted
    planning.current_planning_month config value (set the moment Finance
    confirms it in the upload modal, before any plan_run exists), (4) raise
    — callers turn that into an HTTP 400 (generate-plan) or an empty/no-op
    response (plan-runs history, finalize — nothing to show/finalize yet
    either).

    Reset-cycle fix: (2) and (3) are compared by calendar month, not treated
    as a strict priority order — whichever is the LATER month wins, rather
    than always preferring "latest plan_run" outright. Before Reset existed,
    the latest plan_run's month was always >= the stored config's month (new
    cycles only ever move forward), so a strict priority order and a
    "later of the two" comparison gave the same answer and this distinction
    never mattered. Reset (backend/api/routers/new_model_2.py's
    reset-cycle endpoint) is the first code path that can leave an OLDER
    plan_run as the only one remaining while the stored config still points
    at the newer, now-plan-less current cycle — a strict "plan_run always
    wins" rule would then resolve every endpoint back to the stale old
    cycle instead of the fresh post-reset state. Comparing by month keeps
    the existing, tested "generate for a later month supersedes the stored
    config" behavior (test_api.py::
    test_current_planning_month_endpoint_reflects_persisted_config_then_latest_plan_run)
    while also fixing the reset case."""
    if requested is not None:
        return requested
    latest = (
        session.query(PlanRun)
        .filter(PlanRun.model_used == NEW_MODEL_2_PLAN_RUN_LABEL)
        .order_by(PlanRun.created_at.desc(), PlanRun.id.desc())
        .first()
    )
    stored = get_current_planning_month(session)
    stored_date = _parse_planning_month(stored) if stored is not None else None
    if latest is not None and stored_date is not None:
        return max(latest.month, stored_date)
    if latest is not None:
        return latest.month
    if stored_date is not None:
        return stored_date
    raise ValueError("planning_month is required for the first Generate Plan of a cycle")


def _latest_new_model_2_plan_run(session, planning_month: date | None):
    """New Model 2's own "this cycle's plan_run" (docs/14 fix) — by the
    resolved planning month, never _current_cycle_month() (the ledger-
    ingestion-anchored value every other model still uses unchanged)."""
    if planning_month is None:
        return None
    return (
        session.query(PlanRun)
        .filter(PlanRun.month == planning_month, PlanRun.model_used == NEW_MODEL_2_PLAN_RUN_LABEL)
        .order_by(PlanRun.created_at.desc(), PlanRun.id.desc())
        .first()
    )


def _last_new_model_2_allocated_amount(session, vendor_id, planning_month):
    """Mirrors new_model_1.py's _last_new_model_1_allocated_amount exactly,
    scoped to this model's own plan_run family — the frozen Suggested
    figure an overridden vendor had right before they got excluded, not
    recomputed."""
    latest_run = _latest_new_model_2_plan_run(session, planning_month)
    if latest_run is None:
        return 0.0
    allocation = (
        session.query(PlanAllocation)
        .filter(PlanAllocation.plan_run_id == latest_run.id, PlanAllocation.vendor_id == vendor_id)
        .first()
    )
    return float(allocation.allocated_amount) if allocation is not None else 0.0


@router.get("/models/5/suggested-planning-month", response_model=SuggestedPlanningMonthResponse)
def get_new_model_2_suggested_planning_month():
    """docs/14 + this task: the sensible planning month is always (the
    sheet's real last recorded ledger month) + 1 — never an arbitrary pick
    disconnected from what data actually exists. Frontend pre-fills Card 1's
    planning-month picker with this rather than leaving it blank/today's
    real calendar date."""
    session = SessionLocal()
    try:
        last_recorded = _last_recorded_ledger_month(session)
        suggested = _month_plus_one(last_recorded) if last_recorded else None
        return {
            "last_recorded_month": last_recorded.strftime("%Y-%m") if last_recorded else None,
            "suggested_planning_month": suggested.strftime("%Y-%m") if suggested else None,
        }
    finally:
        session.close()


@router.get("/models/5/current-planning-month", response_model=CurrentPlanningMonthResponse)
def get_new_model_2_current_planning_month():
    """Main-tab upload-confirm task (section 3/4): read-only lookup for
    whatever _resolve_planning_month() would fall back to right now, with no
    plan_run needed and nothing recomputed — the Main-tab confirm modal uses
    this to pre-fill Planning Month from the current cycle's already-set
    value, and the Planning tab uses it to show that value read-only instead
    of asking Finance to re-enter it. None only when nothing's resolvable
    yet at all (no plan_run this cycle, nothing ever confirmed via upload)."""
    session = SessionLocal()
    try:
        try:
            resolved = _resolve_planning_month(session, None)
        except ValueError:
            return {"planning_month": None}
        return {"planning_month": resolved.strftime("%Y-%m")}
    finally:
        session.close()


@router.get("/models/5/plan-runs", response_model=PlanRunHistoryResponse)
def get_new_model_2_plan_runs():
    session = SessionLocal()
    try:
        try:
            planning_month = _resolve_planning_month(session, None)
        except ValueError:
            # No New Model 2 plan_run exists yet this cycle at all — nothing
            # to show, same "empty history" shape every other model's own
            # history endpoint returns before its own first Generate.
            return {"plan_runs": [], "vendor_week_distribution_plans": {}}
        return build_plan_run_history_response(session, 5, cycle_month=planning_month)
    finally:
        session.close()


@router.get("/models/5/minimum-funds-required", response_model=NewModel2MinimumFundsRequiredResponse)
def get_new_model_2_minimum_funds_required(planning_month: str | None = None):
    """docs/14 Card 2 — normally called with Finance's just-picked planning
    month directly (Card 1). Made optional (Main-tab upload-confirm task,
    section 3): resolves through the same _resolve_planning_month() fallback
    (latest plan_run, then the persisted planning.current_planning_month
    config value) every other New Model 2 endpoint already uses, so a caller
    with nothing explicit to pass still gets a real figure rather than a
    forced 400. Existing callers that always pass planning_month explicitly
    (PlanningView.tsx's handleCalcFundsRequired) are unaffected — `requested`
    resolves immediately and this fallback path never runs for them."""
    session = SessionLocal()
    try:
        try:
            requested = _parse_planning_month(planning_month) if planning_month else None
            planning_month_date = _resolve_planning_month(session, requested)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        planning_month = planning_month_date.strftime("%Y-%m")

        as_of = _month_minus_one(planning_month_date)
        result = calculate_minimum_funds_required_v2(session=session, as_of=as_of)

        # Non-blocking sanity check (this task, same "loud, never blocking"
        # convention as column_mapping.sheet_start_month_warning()): flags
        # when Finance's manually-picked planning month isn't exactly last
        # recorded ledger month + 1 — there may be legitimate reasons to
        # plan differently, so this never stops Card 1/2 from proceeding.
        planning_month_warning = None
        last_recorded = _last_recorded_ledger_month(session)
        if last_recorded is not None:
            expected = _month_plus_one(last_recorded)
            if planning_month_date != expected:
                planning_month_warning = (
                    f"Planning month {planning_month} isn't the sheet's suggested next month "
                    f"({expected:%b-%Y}, one month after the last recorded ledger month "
                    f"{last_recorded:%b-%Y}) — double check this is intentional."
                )

        # Audit-log revamp (Task B) — Card 2's explicit "Calculate Minimum
        # Funds Required" action (CLAUDE.md rule 4a), not the Planning
        # table's own per-vendor column (a different, far more frequent
        # caller of calculate_minimum_funds_required_v2() that never logs).
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="min_funds_calculated",
                old_value=None,
                new_value=f"Rs {result['total']:.2f} (planning_month={planning_month})",
                source=ChangeSource.MIN_FUNDS_CALC.value,
            )
        )
        session.commit()

        return {
            "total": result["total"],
            "breakdown": result["breakdown"].to_dict(orient="records"),
            "planning_month": planning_month,
            "as_of": as_of,
            "planning_month_warning": planning_month_warning,
        }
    finally:
        session.close()


@router.get("/models/5/vendors/{vendor_id}/min-funds-required", response_model=VendorMinFundsRequiredV2Response)
def get_new_model_2_vendor_min_funds_required(vendor_id: int, planning_month: str | None = None):
    """docs/14's vendor detail card "Min Funds Required" section — live,
    real-computed, never hardcoded. planning_month is optional: before
    Finance's first Generate Plan of a cycle, the frontend passes whatever
    it currently has picked in Card 1; afterwards it can omit it and this
    reuses the cycle's stored planning month, same reuse rule as generate-
    plan-and-weekly-view."""
    session = SessionLocal()
    try:
        try:
            requested = _parse_planning_month(planning_month) if planning_month else None
            planning_month_date = _resolve_planning_month(session, requested)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e

        as_of = _month_minus_one(planning_month_date)
        vendor = session.query(Vendor).filter_by(id=vendor_id).one()
        ledger_rows = (
            session.query(MonthlyLedger).filter_by(vendor_id=vendor_id).order_by(MonthlyLedger.month).all()
        )
        detail = min_funds_breakdown_v2(vendor, ledger_rows, as_of=as_of)
        return {**detail, "planning_month": planning_month_date.strftime("%Y-%m"), "as_of": as_of}
    finally:
        session.close()


@router.get("/models/5/vendor-min-funds-required", response_model=AllVendorMinFundsRequiredV2Response)
def get_new_model_2_all_vendor_min_funds_required(planning_month: str | None = None):
    """Fix: New Model 2's own planning table "Total Min Funds" column
    previously always read GET /vendors/{id}/aging's field — the OLD
    carry-forward rule (backend/shared/min_funds.py), same one every other
    tab correctly uses, but New Model 2's own table needs required_amount_v2
    instead. Bulk (one call, every vendor) rather than the vendor-detail
    endpoint's one-at-a-time pattern, since the table renders all of them
    at once.

    Unlike the vendor-detail and Card-2 endpoints, this one does NOT 400
    when no planning month is resolvable yet (no plan_run this cycle, none
    passed) — the table itself must still render before Finance's first
    Generate Plan, just with this column blank ("—") rather than erroring
    on every render.
    """
    session = SessionLocal()
    try:
        try:
            requested = _parse_planning_month(planning_month) if planning_month else None
            planning_month_date = _resolve_planning_month(session, requested)
        except ValueError:
            return {"planning_month": None, "as_of": None, "breakdown": []}

        as_of = _month_minus_one(planning_month_date)
        result = calculate_minimum_funds_required_v2(session=session, as_of=as_of)
        return {
            "planning_month": planning_month_date.strftime("%Y-%m"),
            "as_of": as_of,
            "breakdown": result["breakdown"].to_dict(orient="records"),
        }
    finally:
        session.close()


@router.get("/models/5/min-funds-verification-export")
def get_new_model_2_min_funds_verification_export():
    """Downloadable manual-check sheet (Vendor Name, Outstanding, one column
    per aging window, Min Funds Required) for every active vendor — every
    value reuses the same functions/as_of resolution GET /models/5/vendor-
    min-funds-required ("Cal Min Funds") already uses, so this export
    always agrees with what Finance sees on screen. Brand-new workbook, not
    derived from EXCEL_PATH's own rows/columns/formatting — unlike
    finalized-plan-export above. It does now open EXCEL_PATH once, read-only
    (never saved back to), purely to run the same duplicate-ERP-code check
    the Exceptions tab needs (see below) — the "unrelated to EXCEL_PATH"
    framing above is about not writing into/deriving from it, not about
    never reading it at all.

    Aging columns are dynamic, not a fixed 5 — sized to the single oldest
    vendor in this export (aging.monthly_breakdown already walks every
    calendar month, oldest -> newest, incl. zero-balance ones). A vendor's
    cell is 0.0 for a month that genuinely applies to them with nothing
    outstanding, but blank (None) for a label beyond their own oldest
    month — those two are not the same thing.

    Duplicate handling: the one duplicate check in this codebase is
    ERP-code-based, at ingestion (backend/ingestion/load_excel.py) — a
    duplicate ERP code is skipped entirely at that stage and never becomes
    a Vendor row, so there's nothing left to flag on the main sheet itself.
    It IS surfaced, though: the "Exceptions" tab this export now always
    includes (backend/shared/export_exceptions.py) names the dropped
    vendor(s) by re-running that same duplicate check against the live
    sheet, alongside every other genuine anomaly (capped required amounts,
    no-ledger-history-before-planning-month vendors, inactive vendors still
    carrying a real balance) — a second WORKSHEET, not extra rows appended
    to the main table, so it can never corrupt this sheet's own column
    dtypes when read back (e.g. via pandas' read_excel()).
    """
    session = SessionLocal()
    try:
        try:
            planning_month = _resolve_planning_month(session, None)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        as_of = _month_minus_one(planning_month)

        vendors = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).all()
        ledger_by_vendor = {}
        vendor_rows = []
        mb_to_label = {}
        for vendor in vendors:
            ledger_rows = (
                session.query(MonthlyLedger).filter_by(vendor_id=vendor.id).order_by(MonthlyLedger.month).all()
            )
            ledger_by_vendor[vendor.id] = ledger_rows
            paid_so_far = float(vendor.paid_so_far_this_month)
            aging = compute_vendor_aging(ledger_rows, as_of=as_of, already_paid_this_cycle=paid_so_far)
            min_funds_required, _ = required_amount_v2(vendor, ledger_rows, as_of=as_of)
            breakdown_by_mb = {entry["months_back"]: entry["amount"] for entry in aging.monthly_breakdown}
            mb_to_label.update({entry["months_back"]: entry["label"] for entry in aging.monthly_breakdown})
            vendor_rows.append((vendor, live_outstanding_balance(vendor), breakdown_by_mb, min_funds_required))

        # Ascending months_back -> newest ("0-30") to oldest, left to right.
        ordered_mbs = sorted(mb_to_label)
        bucket_labels = [mb_to_label[mb] for mb in ordered_mbs]

        rows = []
        for vendor, outstanding, breakdown_by_mb, min_funds_required in vendor_rows:
            row = {"Vendor Name": vendor.vendor_name, "Outstanding": outstanding}
            for mb, label in zip(ordered_mbs, bucket_labels):
                row[label] = breakdown_by_mb.get(mb)  # None = not applicable; 0.0 = applicable, zero balance
            row["Min Funds Required"] = min_funds_required
            rows.append(row)

        df = pd.DataFrame(
            rows, columns=["Vendor Name", "Outstanding", *bucket_labels, "Min Funds Required"]
        )
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        df.to_excel(tmp_path, index=False)

        # Exceptions tab (investigation task, this task's own implementation)
        # — a second worksheet, added by reopening the just-written file
        # rather than reworking this function's pandas-based main sheet.
        # Read-only open of the live master sheet, purely for the
        # duplicate-ERP-code check inside build_exceptions() — never saved
        # back to EXCEL_PATH.
        sync_excel_path_from_db()  # Vercel fix — this container may be warm from before the latest upload
        dup_wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        exceptions = build_exceptions(session, dup_wb.active, as_of)
        wb = openpyxl.load_workbook(tmp_path)
        write_exceptions_sheet(wb, exceptions)
        wb.save(tmp_path)

        filename = f"Min Funds Verification - {planning_month:%b-%Y} - {datetime.now():%d-%b-%Y %H%M%S}.xlsx"

        # Audit-log revamp (Task B) — every export, which type.
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="export",
                old_value=None,
                new_value=f"Min Funds Verification export ({filename})",
                source=ChangeSource.EXPORT.value,
            )
        )
        session.commit()

        return FileResponse(
            tmp_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(os.remove, tmp_path),
        )
    finally:
        session.close()


@router.post("/models/5/generate-plan-and-weekly-view", response_model=NewModel2PlanAndWeeklyViewResponse)
def post_new_model_2_generate_plan_and_weekly_view(body: NewModel2GeneratePlanRequest):
    """New Model 2 has no separate regeneration endpoint (docs/14, same
    reason as New Model 1: the allocation function itself never branches on
    first-vs-regeneration), so every "Generate"/"Regenerate" click for this
    tab calls this same endpoint fresh.
    """
    session = SessionLocal()
    try:
        try:
            requested = _parse_planning_month(body.planning_month) if body.planning_month else None
            planning_month = _resolve_planning_month(session, requested)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        as_of = _month_minus_one(planning_month)

        all_vendors = session.query(Vendor).order_by(Vendor.id).all()
        overridden = [v for v in all_vendors if v.override_amount is not None]
        overridden_ids = {v.id for v in overridden}
        eligible_ids = {v.id for v in all_vendors} - overridden_ids

        total_overridden = sum(float(v.override_amount) for v in overridden)
        effective_funds = max(body.available_funds - total_overridden, 0.0)

        plan = generate_plan(effective_funds, session=session, as_of=as_of, eligible_vendor_ids=eligible_ids)

        if overridden:
            ledger_by_vendor_id = {}
            for row in session.query(MonthlyLedger).filter(MonthlyLedger.vendor_id.in_(overridden_ids)).order_by(
                MonthlyLedger.vendor_id, MonthlyLedger.month
            ):
                ledger_by_vendor_id.setdefault(row.vendor_id, []).append(row)

            frozen_rows = []
            for vendor in overridden:
                ledger_rows = ledger_by_vendor_id.get(vendor.id, [])
                amount, rule = required_amount_v2(vendor, ledger_rows, as_of=as_of)
                frozen_rows.append(
                    {
                        "vendor_id": vendor.id,
                        "erp_code": vendor.erp_code,
                        "vendor_name": vendor.vendor_name,
                        "category": vendor.category,
                        "outstanding_balance": live_outstanding_balance(vendor),
                        "required_amount": amount,
                        "rule": rule,
                        "priority_tag": vendor.priority_tag,  # already a plain string (or None)
                        # Frozen — the model never touched this vendor this
                        # round, so their Suggested figure stays exactly
                        # what it was before the override, not recomputed.
                        "allocated_amount": _last_new_model_2_allocated_amount(session, vendor.id, planning_month),
                        "status": AllocationStatus.OVERRIDE_LOCKED.value,
                    }
                )
            plan["allocations"] = pd.concat([plan["allocations"], pd.DataFrame(frozen_rows)], ignore_index=True)

        weekly_view = build_weekly_view(
            plan["allocations"],
            session=session,
            as_of=as_of,
            model_used=NEW_MODEL_2_PLAN_RUN_LABEL,
            funds_figure=body.available_funds,
            # Bug fix (this task): without this, build_weekly_view() always
            # pulled every vendor forward to TODAY's real week, even when
            # Finance is planning a future month — meaningless across two
            # different months' W1-W5. See planner.py's own docstring.
            planning_month=planning_month,
        )
        if weekly_view["plan_run_id"] is not None:
            # build_weekly_view() (shared infra, untouched) stamps `as_of`
            # onto the plan_run it persists — New Model 2's own PlanRun.month
            # must be the actual planning month Finance picked (docs/14),
            # not the aging reference point one calendar month behind it.
            # Fixed up here, post-hoc, rather than by changing planner.py's
            # own shared behavior for every other model.
            #
            # Analytics funds-trend task: min_funds_required recorded
            # alongside funds_figure on this same row, from this Generate
            # Plan onward — reuses calculate_minimum_funds_required_v2()
            # (same required_amount_v2() summation Card 2's own endpoint
            # uses, same as_of) rather than a second copy of that math. Old
            # PlanRun rows are NOT backfilled — real trend data starts here.
            min_funds_total = calculate_minimum_funds_required_v2(session=session, as_of=as_of)["total"]
            # Funds Left card fix: leftover_remaining persisted alongside
            # min_funds_required on this same row/update call, so a
            # Planning-tab remount can read the real post-allocation figure
            # back instead of approximating it (PlanningView.tsx's loadAll()).
            session.query(PlanRun).filter_by(id=weekly_view["plan_run_id"]).update(
                {
                    "month": planning_month,
                    "min_funds_required": min_funds_total,
                    "leftover_remaining": plan["leftover_remaining"],
                }
            )
        # Audit-log revamp (Task B) — two distinct entries per Generate/
        # Regenerate click: the funds figure Finance typed in (its own
        # source/filter category, even though there's no separate "submit
        # funds" endpoint in this codebase — it always arrives bundled with
        # this same call), and the plan outcome itself.
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="available_funds",
                old_value=None,
                new_value=f"{body.available_funds:.2f} (planning_month={planning_month.strftime('%Y-%m')})",
                source=ChangeSource.FUNDS_INPUT.value,
            )
        )
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="generate_plan",
                old_value=None,
                new_value=(
                    f"{len(plan['allocations'])} vendors, "
                    f"Rs {plan['allocations']['allocated_amount'].sum():.2f} allocated "
                    f"(planning_month={planning_month.strftime('%Y-%m')})"
                ),
                source=ChangeSource.GENERATE_PLAN.value,
            )
        )
        session.commit()
        return {
            "plan": _plan_response(plan),
            "weekly_view": {
                "weekly_summary": weekly_view["weekly_summary"].reset_index().to_dict(orient="records"),
                "detail": weekly_view["detail"].to_dict(orient="records"),
                "plan_run_id": weekly_view["plan_run_id"],
            },
            "total_overridden": total_overridden,
            "funds_left_for_regeneration": effective_funds,
        }
    finally:
        session.close()


@router.post("/models/5/finalize", response_model=FinalizeCheckResponse)
def post_new_model_2_finalize():
    """docs/14's shortfall check, mirrors New Model 1's own
    POST /models/4/finalize (see that router's docstring — not
    re-explained here): reconciles the CURRENT state (any overrides
    applied) of New Model 2's latest plan_run against the funds figure it
    was generated with, and publishes the Budget snapshot
    (finalize_plan(model_number=5)) unconditionally — CLAUDE.md rule 2 /
    docs/14: Finalize is a suggestion-surfacing action, never a block, so a
    shortfall is reported back (`ok`/`over_by`) but never stops the
    snapshot from being taken (Sarath's explicit call — an earlier version
    of this endpoint withheld finalize_plan() on a shortfall, which was
    the actual restriction being removed here).

    `over_by` is the same cuttable-shortfall figure the Funds Left card
    already shows (-leftover_remaining, floored at 0) — recomputed live
    here via the exact same generate_plan() call the Generate/Regenerate
    endpoint uses (overridden vendors excluded, their amount pre-
    subtracted from available_funds), NOT summed from this plan_run's
    PlanAllocation rows. Two real bugs that formula-from-PlanAllocation-
    rows approach had, both found chasing real mismatches against the
    Funds Left card for the same plan:
    (1) a flat total_committed-vs-available_funds sum double-counts the
    guaranteed (Must Pay/Commitment) tier's own inherent shortfall as if
    Finance could fix it by cutting Normal/Inactive vendors (~₹1.19Cr
    shown vs Funds Left's correct ~₹44.9L);
    (2) even after excluding the guaranteed tier correctly, PlanAllocation
    rows only exist for vendors planner.py's build_weekly_view() actually
    placed into a week (Vendor.assigned_week set) — any vendor with no
    assigned_week yet is silently absent from this plan_run's allocations
    entirely, so a PlanAllocation-only sum under-counts real production
    data with such vendors, while generate_plan()'s own leftover_remaining
    (what Funds Left reads) is computed over every vendor with an
    outstanding balance regardless of assigned_week.
    """
    session = SessionLocal()
    try:
        try:
            planning_month = _resolve_planning_month(session, None)
        except ValueError:
            planning_month = None
        latest_run = _latest_new_model_2_plan_run(session, planning_month)
        if latest_run is None or latest_run.funds_figure is None:
            vendor_count = finalize_plan(session, model_number=5)
            session.commit()
            return {
                "ok": True,
                "total_committed": 0.0,
                "available_funds": 0.0,
                "over_by": 0.0,
                "responsible_vendors": [],
                "vendor_count": vendor_count,
            }

        allocations = session.query(PlanAllocation).filter_by(plan_run_id=latest_run.id).all()
        available_funds = float(latest_run.funds_figure)

        total_committed = 0.0
        responsible: list[dict[str, Any]] = []
        for allocation in allocations:
            suggested = float(allocation.allocated_amount)
            override = float(allocation.override_amount) if allocation.override_amount is not None else None
            effective = override if override is not None else suggested
            total_committed += effective
            if override is not None and override > suggested + MONEY_EPSILON:
                vendor = session.query(Vendor).filter_by(id=allocation.vendor_id).one()
                responsible.append(
                    {
                        "vendor_id": vendor.id,
                        "erp_code": vendor.erp_code,
                        "vendor_name": vendor.vendor_name,
                        "suggested_amount": suggested,
                        "override_amount": override,
                    }
                )

        all_vendors = session.query(Vendor).order_by(Vendor.id).all()
        overridden = [v for v in all_vendors if v.override_amount is not None]
        overridden_ids = {v.id for v in overridden}
        eligible_ids = {v.id for v in all_vendors} - overridden_ids
        total_overridden = sum(float(v.override_amount) for v in overridden)
        effective_funds = max(available_funds - total_overridden, 0.0)
        as_of = _month_minus_one(latest_run.month)
        live_plan = generate_plan(effective_funds, session=session, as_of=as_of, eligible_vendor_ids=eligible_ids)
        over_by = max(-live_plan["leftover_remaining"], 0.0)
        ok = over_by <= MONEY_EPSILON
        vendor_count = finalize_plan(session, model_number=5)
        session.commit()
        return {
            "ok": ok,
            "total_committed": total_committed,
            "available_funds": available_funds,
            "over_by": over_by,
            "responsible_vendors": responsible,
            "vendor_count": vendor_count,
        }
    finally:
        session.close()


@router.post("/models/5/reset-cycle", response_model=ResetCycleResponse)
def post_new_model_2_reset_cycle():
    """Reset button (this task) — confirmed-with-Sarath scope (CLAUDE.md):
    clears the CURRENT planning cycle's calculations only, taking Finance
    back to the state right after upload, before Min Funds was ever
    calculated (rule 4's starting point). Does NOT touch vendor master
    data/Excel (never writes to it), Configuration, already-finalized past
    months' plan history (different PlanRun.month, left alone), the
    confirmed planning month itself, or the audit log's own history — this
    only ever ADDS one entry recording the reset.

    Deletes every PlanRun row (+ cascaded PlanAllocation rows, no ORM
    cascade configured so done explicitly — same two-step delete
    backend/api/routers/plan_runs.py's own hard-delete already uses) whose
    month matches the current planning month, then reuses
    _reset_vendor_cycle_state() (backend/month_end/rollover.py) — the exact
    same per-vendor reset already used at month-end rollover — rather than
    reimplementing it. All in one transaction: a failure partway through
    rolls back everything.
    """
    session = SessionLocal()
    try:
        try:
            planning_month = _resolve_planning_month(session, None)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e

        plan_run_ids = [
            row.id
            for row in session.query(PlanRun.id)
            .filter(PlanRun.model_used == NEW_MODEL_2_PLAN_RUN_LABEL, PlanRun.month == planning_month)
            .all()
        ]
        deleted_allocations = 0
        if plan_run_ids:
            deleted_allocations = (
                session.query(PlanAllocation)
                .filter(PlanAllocation.plan_run_id.in_(plan_run_ids))
                .delete(synchronize_session=False)
            )
            session.query(PlanRun).filter(PlanRun.id.in_(plan_run_ids)).delete(synchronize_session=False)

        vendors_reset = _reset_vendor_cycle_state(session)

        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="planning_cycle_reset",
                old_value=f"planning_month={planning_month.strftime('%Y-%m')}, plan_runs={len(plan_run_ids)}",
                new_value="reset",
                source=ChangeSource.RESET_CYCLE.value,
            )
        )
        session.commit()
        return {
            "planning_month": planning_month.strftime("%Y-%m"),
            "deleted_plan_runs": len(plan_run_ids),
            "deleted_allocations": deleted_allocations,
            "vendors_reset": vendors_reset,
        }
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


@router.get("/models/5/finalized-plan-export")
def get_new_model_2_finalized_plan_export():
    """Downloadable copy of the REAL master workbook (every existing
    column/row/sheet/formatting untouched) with seven new columns appended,
    in this exact left-to-right order: W1, W2, W3, W4, W5, then "Suggested
    Plan Percentage", then "Suggested Plan Amount" — for every vendor
    Finance has finalized a budget for this cycle
    (Vendor.finalized_budget_amount is not None). Vendors never finalized
    this cycle are left blank in every new column, not 0.

    Each week column holds this vendor's latest plan_run allocation's own
    week_distribution_plan value for that week (bug fix: reading
    Vendor.week_distribution_plan directly here left W1-W5 blank for any
    vendor Finance hadn't individually edited — see
    latest_week_distribution_for_vendor()'s docstring) — the vendor's latest
    override amount if Finance has set one, else the model's own suggested
    amount, same "effective amount" this field already tracks everywhere
    else (see planner.py/plan_allocations.py's override endpoint) — blank
    for any week not present in that vendor's distribution, not 0.

    Percentage = finalized_budget_amount / required_amount_v2() * 100 — the
    exact same (budget, min_funds_required) pairing GET /vendors/payment-
    tracking already computes, just against the finalized snapshot instead
    of actual payments made, so this export always agrees with what Finance
    already sees on screen.

    Read-only: this only ever loads EXCEL_PATH and saves the modified copy
    to a fresh temp file — it must NEVER write back to EXCEL_PATH itself
    (see test_new_model_2_export.py's byte-identical regression test).

    Also always includes an "Exceptions" tab (investigation task, this
    task's own implementation) — a second WORKSHEET added to this same
    workbook, never extra rows appended to the vendor sheet above (that
    would grow `ws.max_row` and risk corrupting anything reading this
    export back), listing genuine anomalies: dropped duplicate-ERP-code
    vendors, capped required amounts, no-ledger-history-before-planning-
    month vendors, and inactive vendors still carrying a real balance —
    see backend/shared/export_exceptions.py. Runs its own all-vendor pass,
    separate from the `finalized_vendors`-only loop above.
    """
    session = SessionLocal()
    try:
        finalized_vendors = (
            session.query(Vendor).filter(Vendor.finalized_budget_amount.isnot(None)).order_by(Vendor.id).all()
        )
        if not finalized_vendors:
            raise HTTPException(status_code=409, detail="Finalize the plan before downloading")

        try:
            planning_month = _resolve_planning_month(session, None)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e

        rows = []
        for vendor in finalized_vendors:
            ledger_rows = (
                session.query(MonthlyLedger).filter_by(vendor_id=vendor.id).order_by(MonthlyLedger.month).all()
            )
            min_funds_required, _ = required_amount_v2(vendor, ledger_rows, as_of=None, paid_so_far_override=0.0)
            amount = float(vendor.finalized_budget_amount)
            percentage = round(amount / min_funds_required * 100, 2) if min_funds_required > MONEY_EPSILON else None
            # Bug fix: Vendor.week_distribution_plan (the "sticky" field) is
            # only ever written by a manual per-vendor Finance edit
            # (plan_allocations.py) — reading it alone left W1-W5 blank for
            # every vendor nobody had individually edited, even though
            # planner.py computes a correct default onto every
            # PlanAllocation row at generate time. Vendor's own sticky value
            # still wins when Finance HAS edited it directly (same
            # precedence planner.py itself uses when stamping a fresh
            # allocation); latest_week_distribution_for_vendor() is only the
            # fallback for the common case nobody has.
            distribution = vendor.week_distribution_plan or latest_week_distribution_for_vendor(session, vendor.id) or {}
            week_amounts = [distribution.get(str(week)) for week in range(1, 6)]
            rows.append((vendor.erp_code, vendor.entity, amount, percentage, week_amounts))

        header_overrides = column_mapping_store.load_overrides(session)
        sheet_start_month = column_mapping_store.get_sheet_start_month(session)

        sync_excel_path_from_db()  # Vercel fix — this container may be warm from before the latest upload
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        sheet_map = build_sheet_map(ws, header_overrides=header_overrides, sheet_start_month=sheet_start_month)

        # Order matters: W1-W5 created first, then the two Suggested
        # columns, so a fresh sheet gets them left-to-right in exactly this
        # sequence (_find_or_create_column always appends at the sheet's
        # current right edge).
        week_cols = [_find_or_create_column(ws, f"W{week}", sheet_map.header_row) for week in range(1, 6)]
        pct_col = _find_or_create_column(ws, "Suggested Plan Percentage", sheet_map.header_row)
        amt_col = _find_or_create_column(ws, "Suggested Plan Amount", sheet_map.header_row)

        for erp_code, entity, amount, percentage, week_amounts in rows:
            row = _find_vendor_row(ws, sheet_map, erp_code, entity)
            for col, week_amount in zip(week_cols, week_amounts):
                ws.cell(row=row, column=col).value = week_amount
            ws.cell(row=row, column=amt_col).value = amount
            ws.cell(row=row, column=pct_col).value = percentage

        # Exceptions tab (investigation task, this task's own implementation)
        # — a second worksheet on this same wb, reusing the ws/sheet_map
        # already open above for the duplicate-ERP-code check. Unlike the
        # rows/columns loop above, this needs EVERY vendor (duplicates,
        # inactive-with-balance, and capped vendors aren't necessarily
        # finalized this cycle), not just `finalized_vendors` — a genuinely
        # separate pass, done inside build_exceptions() via its own query.
        as_of = _month_minus_one(planning_month)
        exceptions = build_exceptions(session, ws, as_of)
        write_exceptions_sheet(wb, exceptions)

        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(tmp_path)

        filename = f"Vendor Payment Plan - {planning_month:%b-%Y} - {datetime.now():%d-%b-%Y %H%M%S}.xlsx"

        # Audit-log revamp (Task B) — every export, which type.
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=None,
                field_name="export",
                old_value=None,
                new_value=f"Finalized Plan export ({filename})",
                source=ChangeSource.EXPORT.value,
            )
        )
        session.commit()

        return FileResponse(
            tmp_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(os.remove, tmp_path),
        )
    finally:
        session.close()
