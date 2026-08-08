"""GET /models/5/finalized-plan-export (Excel download task) — router-level
tests. Same fresh-temp-DB/excel-copy isolation convention as
test_new_model_2_router.py (duplicated here rather than shared via
conftest.py, per this project's existing per-file-fixture convention).

Most important test here: get_new_model_2_finalized_plan_export() must
NEVER write back to the real EXCEL_PATH — it only ever loads it read-only
and saves the modified copy to a fresh temp file (see
test_export_never_touches_the_real_excel_file below).
"""
import shutil
import sys

import os
import pytest
import openpyxl
from io import BytesIO
from fastapi.testclient import TestClient

from backend.ingestion.load_excel import EXCEL_PATH


def _fresh_app(tmp_path):
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for name in list(sys.modules):
        if name == "backend" or name.startswith("backend."):
            del sys.modules[name]
    from backend.api.main import app

    return app


@pytest.fixture
def client(tmp_path):
    return TestClient(_fresh_app(tmp_path))


@pytest.fixture
def excel_copy(tmp_path):
    dest = tmp_path / "copy.xlsx"
    shutil.copy(EXCEL_PATH, dest)
    return str(dest)


@pytest.fixture
def seeded_client(client, excel_copy):
    resp = client.post("/ingestion/load", json={"excel_path": excel_copy})
    assert resp.status_code == 200
    return client


def _finalize_a_plan(seeded_client):
    generate = seeded_client.post(
        "/models/5/generate-plan-and-weekly-view",
        json={"available_funds": 1_000_000_000, "planning_month": "2030-01"},
    )
    assert generate.status_code == 200
    finalize = seeded_client.post("/models/5/finalize", json={})
    assert finalize.status_code == 200
    assert finalize.json()["ok"] is True


def test_export_is_409_before_anything_is_finalized(seeded_client):
    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 409


def test_export_never_touches_the_real_excel_file(seeded_client):
    """Provably safe: reads EXCEL_PATH's own bytes/mtime before and after
    calling the endpoint — this must be a no-op against the real file no
    matter what the endpoint does internally."""
    before_bytes = open(EXCEL_PATH, "rb").read()
    before_mtime = os.path.getmtime(EXCEL_PATH)

    _finalize_a_plan(seeded_client)
    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    after_bytes = open(EXCEL_PATH, "rb").read()
    after_mtime = os.path.getmtime(EXCEL_PATH)
    assert before_bytes == after_bytes
    assert before_mtime == after_mtime


def test_export_has_the_two_new_columns_with_correct_values(seeded_client):
    """Sets finalized_budget_amount directly via the DB rather than relying
    on a real generate-plan-and-weekly-view allocation landing on a
    non-zero amount — the real dataset's week-assignment step doesn't
    reliably produce one in this sandbox (memory/demo15_dataset_quirks.md,
    a pre-existing, unrelated dataset-drift issue: see
    test_new_model_2_router.py::test_override_is_editable_on_the_latest_
    new_model_2_plan_run, which fails the same way on main). This test only
    needs SOME vendor with a real finalized_budget_amount and SOME vendor
    with None — how that amount got set is irrelevant to what this endpoint
    itself does with it.
    """
    from backend.db.session import SessionLocal
    from backend.db.models import MonthlyLedger, Vendor
    from backend.ingestion.column_mapping import build_sheet_map, find_column
    from backend.shared.min_funds_v2 import required_amount_v2

    _finalize_a_plan(seeded_client)  # establishes a plan_run so planning_month resolves

    session = SessionLocal()
    vendors = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).limit(2).all()
    finalized_vendor, blanked_vendor = vendors[0], vendors[1]
    finalized_vendor.finalized_budget_amount = 12345.67
    blanked_vendor.finalized_budget_amount = None
    ledger_rows = (
        session.query(MonthlyLedger).filter_by(vendor_id=finalized_vendor.id).order_by(MonthlyLedger.month).all()
    )
    min_funds_required, _ = required_amount_v2(finalized_vendor, ledger_rows, as_of=None, paid_so_far_override=0.0)
    finalized_erp_code, blanked_erp_code = finalized_vendor.erp_code, blanked_vendor.erp_code
    session.commit()
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    pct_col = find_column(ws, "Suggested Plan Percentage", sheet_map.header_row)
    amt_col = find_column(ws, "Suggested Plan Amount", sheet_map.header_row)
    assert pct_col is not None
    assert amt_col is not None

    def _row_for(erp_code):
        for row in range(sheet_map.data_start_row, ws.max_row + 1):
            if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
                return row
        raise AssertionError(f"{erp_code} not found")

    blanked_row = _row_for(blanked_erp_code)
    assert ws.cell(row=blanked_row, column=amt_col).value is None
    assert ws.cell(row=blanked_row, column=pct_col).value is None

    finalized_row = _row_for(finalized_erp_code)
    assert ws.cell(row=finalized_row, column=amt_col).value == pytest.approx(12345.67)
    expected_pct = round(12345.67 / min_funds_required * 100, 2)
    assert ws.cell(row=finalized_row, column=pct_col).value == pytest.approx(expected_pct)


def test_export_includes_exceptions_worksheet_without_touching_the_main_sheet(seeded_client):
    """Exceptions tab (this task): a second WORKSHEET, added alongside the
    vendor sheet this file's other tests already exercise — must never grow
    that sheet's own used range or change which sheet is active, so every
    existing ws.max_row/sheet_map/_find_vendor_row-based lookup in this
    file keeps working unmodified."""
    source_ws = openpyxl.load_workbook(EXCEL_PATH).active
    source_max_row = source_ws.max_row

    _finalize_a_plan(seeded_client)

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert "Exceptions" in wb.sheetnames
    assert wb.active.title != "Exceptions"  # the vendor sheet stays active

    main_ws = wb.active
    # This export only ever appends new COLUMNS to existing rows on the main
    # sheet, never new rows — its own row count must match the untouched
    # source file exactly, proving the Exceptions tab (a separate worksheet)
    # added nothing to it.
    assert main_ws.max_row == source_max_row


def test_export_resolves_correct_row_for_each_vendor_in_an_erp_code_collision_pair(seeded_client):
    """Entity-code-collision fix: the real sheet's erp_code "V00149" is on
    two rows — ARS "Instant Screening Services" and FP "G-Task Services" —
    both live vendors post-fix. The export's own _find_vendor_row() call
    must land each vendor's Suggested Plan Amount in ITS OWN row, never a
    same-code sibling's."""
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    _finalize_a_plan(seeded_client)

    session = SessionLocal()
    ars_vendor = session.query(Vendor).filter_by(erp_code="V00149", entity="ARS").one()
    fp_vendor = session.query(Vendor).filter_by(erp_code="V00149", entity="FP").one()
    ars_vendor.finalized_budget_amount = 11111.11
    fp_vendor.finalized_budget_amount = 22222.22
    session.commit()
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    amt_col = find_column(ws, "Suggested Plan Amount", sheet_map.header_row)
    assert amt_col is not None

    ars_row = fp_row = None
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value != "V00149":
            continue
        entity = ws.cell(row=row, column=sheet_map.entity_col).value
        if entity == "ARS":
            ars_row = row
        elif entity == "FP":
            fp_row = row
    assert ars_row is not None and fp_row is not None

    assert ws.cell(row=ars_row, column=amt_col).value == pytest.approx(11111.11)
    assert ws.cell(row=fp_row, column=amt_col).value == pytest.approx(22222.22)


def test_export_has_five_week_columns_before_the_two_suggested_columns(seeded_client):
    """W1-W5 (Vendor.week_distribution_plan's own per-week value — the
    latest override if Finance set one, else the model's suggested figure,
    blank for any week absent from that vendor's distribution) must appear
    immediately to the left of Suggested Plan Percentage/Amount, in that
    exact order — 7 new columns total, left to right: W1, W2, W3, W4, W5,
    Suggested Plan Percentage, Suggested Plan Amount."""
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    _finalize_a_plan(seeded_client)

    session = SessionLocal()
    vendors = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).limit(2).all()
    finalized_vendor, blanked_vendor = vendors[0], vendors[1]
    finalized_vendor.finalized_budget_amount = 5000.0
    finalized_vendor.week_distribution_plan = {"1": 1111.0, "3": 3333.0}
    blanked_vendor.finalized_budget_amount = None
    finalized_erp_code, blanked_erp_code = finalized_vendor.erp_code, blanked_vendor.erp_code
    session.commit()
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    week_cols = [find_column(ws, f"W{week}", sheet_map.header_row) for week in range(1, 6)]
    pct_col = find_column(ws, "Suggested Plan Percentage", sheet_map.header_row)
    amt_col = find_column(ws, "Suggested Plan Amount", sheet_map.header_row)
    assert all(c is not None for c in week_cols)
    # Exact left-to-right order: W1..W5 immediately followed by the two
    # Suggested columns, no gaps.
    assert week_cols == sorted(week_cols)
    assert pct_col == week_cols[-1] + 1
    assert amt_col == pct_col + 1

    def _row_for(erp_code):
        for row in range(sheet_map.data_start_row, ws.max_row + 1):
            if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
                return row
        raise AssertionError(f"{erp_code} not found")

    finalized_row = _row_for(finalized_erp_code)
    values = [ws.cell(row=finalized_row, column=c).value for c in week_cols]
    assert values == [1111.0, None, 3333.0, None, None]

    blanked_row = _row_for(blanked_erp_code)
    assert [ws.cell(row=blanked_row, column=c).value for c in week_cols] == [None] * 5


def _seed_allocation_only_distribution(session, vendor, plan_run_id, distribution, assigned_week=1):
    """Simulates the common real case this bug fix targets: planner.py has
    already computed a per-cycle week_distribution_plan onto a
    PlanAllocation row, but Vendor.week_distribution_plan itself was never
    touched (no vendor has ever gone through plan_allocations.py's manual
    edit endpoints) — confirmed real state via the actual production DB
    (419/419 PlanAllocation rows had one, 0/419 matching Vendor rows did)."""
    from backend.db.models import PlanAllocation

    assert vendor.week_distribution_plan is None
    allocation = PlanAllocation(
        plan_run_id=plan_run_id,
        vendor_id=vendor.id,
        assigned_week=assigned_week,
        allocated_amount=vendor.finalized_budget_amount,
        week_distribution_plan=distribution,
    )
    session.add(allocation)
    session.commit()


def _row_and_week_cols(ws, sheet_map, erp_code):
    from backend.ingestion.column_mapping import find_column

    week_cols = [find_column(ws, f"W{week}", sheet_map.header_row) for week in range(1, 6)]
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            return row, week_cols
    raise AssertionError(f"{erp_code} not found")


def test_export_falls_back_to_latest_allocation_distribution_when_vendor_field_never_set__dataset_1(seeded_client):
    """Bug fix, test data set 1: a single-week split, week 2 only. Vendor's
    own week_distribution_plan is never touched (stays None, the real,
    almost-universal production case) — the export must still show the
    figure planner.py already computed onto the vendor's latest
    PlanAllocation row, not blank."""
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.column_mapping import build_sheet_map
    from io import BytesIO
    import openpyxl

    _finalize_a_plan(seeded_client)  # establishes a plan_run so planning_month resolves

    session = SessionLocal()
    from backend.db.models import PlanRun

    plan_run_id = session.query(PlanRun).order_by(PlanRun.id.desc()).first().id
    vendor = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).first()
    vendor.finalized_budget_amount = 7000.0
    session.commit()
    _seed_allocation_only_distribution(session, vendor, plan_run_id, {"2": 7000.0}, assigned_week=2)
    erp_code = vendor.erp_code
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    row, week_cols = _row_and_week_cols(ws, sheet_map, erp_code)
    values = [ws.cell(row=row, column=c).value for c in week_cols]
    assert values == [None, 7000.0, None, None, None]


def test_export_falls_back_to_latest_allocation_distribution_when_vendor_field_never_set__dataset_2(seeded_client):
    """Bug fix, test data set 2: a DIFFERENT vendor, a DIFFERENT (multi-week)
    split across weeks 1 and 4, different amounts — same fallback must hold
    regardless of which weeks/values are involved."""
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.column_mapping import build_sheet_map
    from io import BytesIO
    import openpyxl

    _finalize_a_plan(seeded_client)

    session = SessionLocal()
    from backend.db.models import PlanRun

    plan_run_id = session.query(PlanRun).order_by(PlanRun.id.desc()).first().id
    vendor = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).offset(2).first()
    vendor.finalized_budget_amount = 55000.0
    session.commit()
    _seed_allocation_only_distribution(session, vendor, plan_run_id, {"1": 21000.5, "4": 34000.25}, assigned_week=1)
    erp_code = vendor.erp_code
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    row, week_cols = _row_and_week_cols(ws, sheet_map, erp_code)
    values = [ws.cell(row=row, column=c).value for c in week_cols]
    assert values == [21000.5, None, None, 34000.25, None]


def test_export_prefers_vendors_own_sticky_field_over_the_allocation_snapshot_when_finance_has_edited_it(seeded_client):
    """Regression guard for the fallback's precedence: once Finance HAS
    manually edited a vendor's distribution (Vendor.week_distribution_plan
    set directly, same as plan_allocations.py's edit endpoints would leave
    it), that sticky value must still win over whatever stale figure sits
    on the PlanAllocation snapshot — matches planner.py's own precedence
    when it stamps a fresh allocation."""
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor
    from backend.ingestion.column_mapping import build_sheet_map
    from io import BytesIO
    import openpyxl

    _finalize_a_plan(seeded_client)

    session = SessionLocal()
    from backend.db.models import PlanRun

    plan_run_id = session.query(PlanRun).order_by(PlanRun.id.desc()).first().id
    vendor = session.query(Vendor).filter(Vendor.is_active.isnot(False)).order_by(Vendor.id).offset(4).first()
    vendor.finalized_budget_amount = 9000.0
    session.commit()
    from backend.db.models import PlanAllocation

    stale_allocation = PlanAllocation(
        plan_run_id=plan_run_id, vendor_id=vendor.id, assigned_week=1,
        allocated_amount=9000.0, week_distribution_plan={"1": 999.0},
    )
    session.add(stale_allocation)
    vendor.week_distribution_plan = {"3": 9000.0}  # Finance's own later manual edit
    session.commit()
    erp_code = vendor.erp_code
    session.close()

    resp = seeded_client.get("/models/5/finalized-plan-export")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    row, week_cols = _row_and_week_cols(ws, sheet_map, erp_code)
    values = [ws.cell(row=row, column=c).value for c in week_cols]
    assert values == [None, None, 9000.0, None, None]
