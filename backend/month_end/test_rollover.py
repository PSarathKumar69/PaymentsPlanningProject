"""Self-check for month-end rollover (backend/month_end/rollover.py).

Never touches the real backend/db/app.db or data/Vendor's Details.xlsx:
each test points PAYMENTS_DB_PATH at a fresh copy of the real, already-
ingested app.db (so rollover has real pre-existing vendor/ledger state to
reset against, not a blank DB) and mutates a temp copy of the real Excel.
Run with: pytest backend/month_end/test_rollover.py
"""
import os
import shutil
import sys
import tempfile

import openpyxl
import pytest
from sqlalchemy import func

from backend.ingestion.column_mapping import HEADER_ROW, build_sheet_map, to_number
from backend.ingestion.load_excel import EXCEL_PATH


def _fresh_copies():
    """Re-import backend.db.session (and everything that caches its
    SessionLocal) against a fresh temp copy of the real app.db — same
    re-import trick backend/ingestion/test_load_excel.py already uses for
    its own throwaway-DB tests, except this DB starts pre-populated instead
    of empty."""
    real_db_path = os.path.join(os.path.dirname(__file__), "..", "db", "app.db")
    tmp_dir = tempfile.mkdtemp()
    tmp_db_path = os.path.join(tmp_dir, "test.db")
    shutil.copy(real_db_path, tmp_db_path)
    os.environ["PAYMENTS_DB_PATH"] = tmp_db_path

    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.month_end.rollover",
        "backend.shared.payment_logging",
    ):
        sys.modules.pop(mod, None)

    tmp_excel_path = os.path.join(tmp_dir, "copy.xlsx")
    shutil.copy(EXCEL_PATH, tmp_excel_path)

    from backend.month_end.rollover import rollover_month

    return rollover_month, tmp_excel_path


def _append_synthetic_month(excel_path, payable=10_000.0, payment=4_000.0):
    """Appends one new month's Payable+Payment column pair immediately
    before the existing Total Payable / Total Payment columns — mirrors
    what a real monthly Excel refresh does (docs/07), just with synthetic
    figures instead of a real new month's data. Row-2 merged header labels
    (e.g. "Monthly Payable" spanning H2:V2) go stale after the column shift
    since openpyxl's insert_cols doesn't adjust merge boundaries — but
    column_mapping.py only ever reads row 3 (HEADER_ROW) and below, so this
    has no effect on ingestion correctness.
    """
    # The Closing Balance column holds a live formula (e.g. "=G4+V4-AK4"),
    # not a literal — read cached values from a separate data_only load
    # (same convention load_excel.py itself uses) before the formula-
    # preserving workbook below overwrites that cell with a literal number.
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    ws_data = wb_data.active
    data_sheet_map = build_sheet_map(ws_data)
    prior_closing_by_row = {
        row: to_number(ws_data.cell(row=row, column=data_sheet_map.closing_balance_col).value)
        for row in range(data_sheet_map.data_start_row, ws_data.max_row + 1)
        if ws_data.cell(row=row, column=data_sheet_map.erp_code_col).value is not None
    }

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    total_payable_col = sheet_map.total_payable_col
    total_payment_col = sheet_map.total_payment_col

    # Insert both new columns using pre-computed offsets, not a re-called
    # build_sheet_map between the two inserts — the sheet is transiently
    # asymmetric (one new payable column, no matching payment column yet)
    # and build_sheet_map's own equal-block-length check would reject that.
    ws.insert_cols(total_payable_col)
    ws.cell(row=HEADER_ROW, column=total_payable_col, value="synthetic-payable")
    ws.insert_cols(total_payment_col + 1)
    ws.cell(row=HEADER_ROW, column=total_payment_col + 1, value="synthetic-payment")

    sheet_map = build_sheet_map(ws)  # symmetric again — safe to rebuild now
    new_payable_col = sheet_map.payable_cols[-1][1]
    new_payment_col = sheet_map.payment_cols[-1][1]

    for row, prior_closing in prior_closing_by_row.items():
        ws.cell(row=row, column=new_payable_col, value=payable)
        ws.cell(row=row, column=new_payment_col, value=payment)
        ws.cell(row=row, column=sheet_map.closing_balance_col, value=prior_closing + payable - payment)

    wb.save(excel_path)


def test_rollover_resets_payment_state_only_on_genuine_new_month():
    rollover_month, tmp_excel_path = _fresh_copies()
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.shared.enums import PaymentStatus

    session = SessionLocal()
    original_vendor_count = session.query(Vendor).count()
    # Give one real vendor actual payment history before rollover, so the
    # reset is proven against a genuinely non-default prior state.
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id = vendor.id
    vendor.payment_status = PaymentStatus.PARTIAL
    vendor.paid_so_far_this_month = 250_000.0
    # Sticky, per-vendor, whole-month decisions (docs/06's plan-history
    # feature) — must reset at month-end same as payment_status above, a
    # new cycle starts clean with no override/distribution carried over.
    vendor.override_amount = 999_999.0
    vendor.week_distribution_plan = {"1": 999_999.0}
    # Same whole-month reset — a stale "Finalize Plan" snapshot must not
    # survive into a new cycle's Vendor Payment Table (docs/06 pattern).
    vendor.finalized_budget_amount = 999_999.0
    session.commit()
    session.close()

    _append_synthetic_month(tmp_excel_path)
    result = rollover_month(excel_path=tmp_excel_path)

    assert result["post_month"] > result["pre_month"]
    assert result["vendors_reset"] == original_vendor_count

    session = SessionLocal()
    new_month = result["post_month"]
    ledger_rows = session.query(MonthlyLedger).filter_by(month=new_month).all()
    assert len(ledger_rows) == original_vendor_count  # every vendor got a row for the new month

    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == PaymentStatus.NOT_PAID  # before rollover: PARTIAL
    assert float(reloaded.paid_so_far_this_month) == 0.0  # before rollover: 250,000
    assert reloaded.override_amount is None  # before rollover: 999,999
    assert reloaded.week_distribution_plan is None  # before rollover: {"1": 999,999}
    assert reloaded.finalized_budget_amount is None  # before rollover: 999,999

    # opening_balance carried forward: new month's ledger opening == prior
    # month's closing, and the vendor's live opening_balance == the new closing.
    ledger_row = next(r for r in ledger_rows if r.vendor_id == vendor_id)
    prior_month_row = session.query(MonthlyLedger).filter_by(vendor_id=vendor_id, month=result["pre_month"]).one()
    assert float(ledger_row.opening_balance) == pytest.approx(float(prior_month_row.closing_balance), abs=1)
    assert float(reloaded.opening_balance) == pytest.approx(float(ledger_row.closing_balance), abs=1)
    session.close()


def test_rollover_refuses_same_month_reupload_and_changes_nothing():
    rollover_month, tmp_excel_path = _fresh_copies()
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.shared.enums import PaymentStatus

    session = SessionLocal()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id = vendor.id
    vendor.payment_status = PaymentStatus.PARTIAL
    vendor.paid_so_far_this_month = 99_999.0
    session.commit()
    session.close()

    # No new month appended — an ordinary same-month correction re-upload.
    with pytest.raises(ValueError):
        rollover_month(excel_path=tmp_excel_path)

    session = SessionLocal()
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == PaymentStatus.PARTIAL  # unchanged
    assert float(reloaded.paid_so_far_this_month) == 99_999.0  # unchanged
    session.close()


def test_rollover_same_month_correction_updates_ledger_but_skips_reset():
    """Check 2: rollover_month() doubles as a same-month correction path —
    not just a no-op guard on an untouched re-upload. A genuinely CHANGED
    figure for the CURRENT (non-new) month must still land via load(),
    while payment_status/paid_so_far_this_month stay untouched (the reset
    only fires on a genuine new month).
    """
    rollover_month, tmp_excel_path = _fresh_copies()
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal
    from backend.shared.enums import PaymentStatus

    session = SessionLocal()
    latest_month = session.query(func.max(MonthlyLedger.month)).scalar()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id, erp_code = vendor.id, vendor.erp_code
    original_row = session.query(MonthlyLedger).filter_by(vendor_id=vendor_id, month=latest_month).one()
    original_payable = float(original_row.payable)
    vendor.payment_status = PaymentStatus.PARTIAL
    vendor.paid_so_far_this_month = 42_000.0
    session.commit()
    session.close()

    # Change the CURRENT (latest) month's Payable figure for this vendor —
    # a real correction, no new month column added.
    new_payable = original_payable + 12_345.0
    wb = openpyxl.load_workbook(tmp_excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    target_col = sheet_map.payable_cols[-1][1]
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            ws.cell(row=row, column=target_col, value=new_payable)
            break
    wb.save(tmp_excel_path)

    with pytest.raises(ValueError):
        rollover_month(excel_path=tmp_excel_path)  # same month -> guard refuses the reset

    session = SessionLocal()
    updated_row = session.query(MonthlyLedger).filter_by(vendor_id=vendor_id, month=latest_month).one()
    assert float(updated_row.payable) == pytest.approx(new_payable)  # correction landed via load()

    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == PaymentStatus.PARTIAL  # reset skipped
    assert float(reloaded.paid_so_far_this_month) == 42_000.0  # reset skipped
    session.close()


def test_rollover_reset_failure_rolls_back_ingestion_too(monkeypatch):
    """Check 1: a genuine failure during the reset step — after the
    new-month guard has already confirmed a real new month — must roll
    back BOTH the reset attempt and this call's re-ingestion together.
    Without that, the new month's ledger data would already be committed,
    and a retry would immediately hit the "not strictly newer" guard with
    no way to force the reset through again.
    """
    rollover_month, tmp_excel_path = _fresh_copies()
    import backend.month_end.rollover as rollover_module
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    original_latest_month = session.query(func.max(MonthlyLedger.month)).scalar()
    vendor = session.query(Vendor).order_by(Vendor.id).first()
    vendor_id, original_status = vendor.id, vendor.payment_status
    session.close()

    _append_synthetic_month(tmp_excel_path)

    class _ExplodingPaymentStatus:
        @property
        def NOT_PAID(self):
            raise RuntimeError("simulated failure during the reset loop")

    monkeypatch.setattr(rollover_module, "PaymentStatus", _ExplodingPaymentStatus())

    with pytest.raises(RuntimeError):
        rollover_month(excel_path=tmp_excel_path)

    session = SessionLocal()
    # Both the reset AND the re-ingestion it was sharing a transaction with
    # must be undone — the ledger must still show the ORIGINAL latest
    # month, not the new one the failed attempt tried to land.
    assert session.query(func.max(MonthlyLedger.month)).scalar() == original_latest_month
    reloaded = session.query(Vendor).filter_by(id=vendor_id).one()
    assert reloaded.payment_status == original_status
    session.close()
