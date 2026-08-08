"""Fresh temp DB per test, seeded by ingesting a temp COPY of the real
Excel — never touches the real backend/db/app.db or the real
data/Vendor's Details.xlsx. Same isolation convention as
test_category_priority_auto_update.py/test_upload.py/test_load_excel.py.

REVISED (audit task, 2026-07-31): this file used to run every "proves
DB+Excel actually landed" test with session=None directly against the
real app.db, then hand-revert the touched field via a hardcoded old value
in a finally block. That revert was only correct as long as the
hardcoded value actually matched the vendor's true prior state — once the
real uploaded sheet's data drifted from what these tests assumed at
authoring time (V00400 is genuinely Must Pay in the real data now, not
Normal), a test's own assertion failed BEFORE the hardcoded revert ran,
and the revert then forced the real vendor to the WRONG (stale-assumed)
value anyway, permanently corrupting real Finance data with zero audit
trail (the revert deletes the audit_log row it's "undoing" too). That's a
structural risk in the "run for real, then revert" pattern itself, not
just a bug in one test's revert value — so this file now uses the same
isolated-fresh-DB pattern every other test file in this codebase already
uses, which removes the risk entirely rather than patching the revert.
`update_vendor_field()`'s own session=None vs session=<caller's> behavior
(the actual thing this file tests) is unaffected either way.

Every test below may still fail on the SAME real-data assumptions the old
version hardcoded (e.g. "V00400 starts Category=Normal") if the real
sheet has drifted since — that's the same accepted "dataset drift"
category already tracked in memory (demo15-dataset-quirks), and now it
fails safely against a throwaway temp DB instead of corrupting real data.
"""
import os
import shutil
import sys
from datetime import date

import openpyxl
import pytest
from sqlalchemy.exc import NoResultFound
from sqlalchemy.orm import Session as SASession

from backend.ingestion.load_excel import EXCEL_PATH
from backend.shared.enums import VendorCategory


def _fresh_db_and_master(tmp_path):
    """Points PAYMENTS_DB_PATH at a fresh temp file and forces every
    stateful backend.* module to re-bind against it, THEN ingests a temp
    copy of the real Excel into it — so every test gets the real sheet's
    real vendor rows (V00400, V01542, ...) without ever opening the real
    file for writing or committing against the real DB.

    Must run before any `from backend.configuration.vendor_edits import
    ...` (or anything else that transitively imports backend.db.session)
    in every test function below — importing first would silently bind to
    a previous test's already-cached module (memory: demo15-dataset-quirks'
    "stale module identity" trap).
    """
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in ("backend.db.session", "backend.ingestion.load_excel", "backend.configuration.vendor_edits"):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(EXCEL_PATH, master_path)
    load(excel_path=master_path)
    return master_path


def _copy_excel(tmp_path, master_path, name="copy.xlsx"):
    dest = tmp_path / name
    shutil.copy(master_path, dest)
    return str(dest)


def _header_texts(excel_path):
    from backend.ingestion.column_mapping import build_sheet_map

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header_row = build_sheet_map(ws).header_row
    return {str(c.value).strip().lower() for c in ws[header_row] if c.value}


def _read_cell(excel_path, header_text, erp_code):
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, header_text, sheet_map.header_row)
    assert col is not None, f"column {header_text!r} not found"
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            return ws.cell(row=row, column=col).value
    raise AssertionError(f"{erp_code} not found in {excel_path}")


def _vendor_id(erp_code):
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    try:
        return session.query(Vendor).filter_by(erp_code=erp_code).one().id
    finally:
        session.close()


def test_edit_writes_to_correct_row_of_a_former_erp_code_collision_pair(tmp_path):
    """Entity-code-collision fix: the real sheet's erp_code "V00149" is on
    two rows — ARS "Instant Screening Services" and FP "G-Task Services" —
    a genuine two-entity code reuse, both live DB rows post-fix. Editing
    one of them must land in ITS OWN Excel row, never its same-code
    sibling's — _find_vendor_row() disambiguating by (entity, erp_code) is
    the whole point of this fix; before it, this scenario couldn't even
    arise (the sibling never existed as an editable row at all).

    Imports deliberately happen AFTER _fresh_db_and_master() (unlike most
    of this file's older tests) — importing SessionLocal/update_vendor_field
    beforehand binds them to whatever backend.db.session module happened to
    be cached first, not this test's own fresh tmp DB (see
    _fresh_db_and_master()'s own docstring on this exact trap)."""
    from backend.shared.enums import VendorCategory

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)

    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    ars_vendor = session.query(Vendor).filter_by(erp_code="V00149", entity="ARS").one()
    fp_vendor = session.query(Vendor).filter_by(erp_code="V00149", entity="FP").one()
    ars_id, fp_id = ars_vendor.id, fp_vendor.id
    session.close()

    old, new = update_vendor_field(fp_id, "category", VendorCategory.MUST_PAY, excel_path=excel_copy)
    assert new == VendorCategory.MUST_PAY

    # erp_code alone is ambiguous now (two entities share "V00149") — locate
    # each row explicitly by (erp_code, entity) rather than via the
    # single-match _read_cell() helper above.
    from backend.ingestion.column_mapping import build_sheet_map, find_column
    import openpyxl

    wb = openpyxl.load_workbook(excel_copy)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, "Category", sheet_map.header_row)
    ars_row = fp_row = None
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value != "V00149":
            continue
        entity = ws.cell(row=row, column=sheet_map.entity_col).value
        if entity == "ARS":
            ars_row = row
        elif entity == "FP":
            fp_row = row
    assert ars_row is not None and fp_row is not None

    # ...and ONLY the FP row's — the ARS sibling's own cell is untouched.
    assert ws.cell(row=fp_row, column=col).value == "Must Pay"
    assert ws.cell(row=ars_row, column=col).value != "Must Pay"

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=fp_id).one().category == VendorCategory.MUST_PAY
        assert verify.query(Vendor).filter_by(id=ars_id).one().category != VendorCategory.MUST_PAY
    finally:
        verify.close()


def test_category_updates_db_excel_and_audit_log(tmp_path):
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.shared.enums import ChangeSource, VendorCategory

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    old, new = update_vendor_field(vendor_id, "category", VendorCategory.MUST_PAY, excel_path=excel_copy)

    assert old == VendorCategory.NORMAL
    assert new == VendorCategory.MUST_PAY
    assert _read_cell(excel_copy, "Category", "V00400") == "Must Pay"  # published for real

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().category == VendorCategory.MUST_PAY
        entry = verify.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="category").one()
        assert entry.old_value == "Normal"  # clean Finance-facing text, not a Python repr
        assert entry.new_value == "Must Pay"
        assert entry.source == ChangeSource.UI_EDIT.value
    finally:
        verify.close()


def test_category_updates_to_inactive_writes_db_excel_and_audit_log(tmp_path):
    """docs/14 Inactive design (renamed 2026-07-24 from Exit): Inactive is
    a 4th valid category value, accepted through the exact same path as
    Must Pay/Commitment/Normal — same validation, same Excel write-back,
    same audit-log shape, no special-casing."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal
    from backend.shared.enums import ChangeSource, VendorCategory

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    old, new = update_vendor_field(vendor_id, "category", VendorCategory.INACTIVE, excel_path=excel_copy)

    assert old == VendorCategory.NORMAL
    assert new == VendorCategory.INACTIVE
    assert _read_cell(excel_copy, "Category", "V00400") == "Inactive"  # published for real

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().category == VendorCategory.INACTIVE
        entry = verify.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="category").one()
        assert entry.old_value == "Normal"
        assert entry.new_value == "Inactive"
        assert entry.source == ChangeSource.UI_EDIT.value
    finally:
        verify.close()


def test_commitment_months_updates_db_and_excel(tmp_path):
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V01542")

    old, new = update_vendor_field(vendor_id, "commitment_months", 6, excel_path=excel_copy)

    assert old == 1
    assert new == 6
    assert _read_cell(excel_copy, "Commitment Months", "V01542") == 6

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().commitment_months == 6
    finally:
        verify.close()


def test_assigned_week_updates_db_and_excel_and_can_be_cleared(tmp_path):
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    old, new = update_vendor_field(vendor_id, "assigned_week", 3, excel_path=excel_copy)
    assert old == 2
    assert new == 3
    assert _read_cell(excel_copy, "Assigned Week", "V00400") == 3

    # Clearing back to None is valid too.
    old2, new2 = update_vendor_field(vendor_id, "assigned_week", None, excel_path=excel_copy)
    assert old2 == 3
    assert new2 is None
    assert _read_cell(excel_copy, "Assigned Week", "V00400") is None

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().assigned_week is None
    finally:
        verify.close()


@pytest.mark.parametrize(
    "field,bad_value",
    [
        ("category", "not_a_real_category"),
        ("commitment_months", 0),
        ("commitment_months", -3),
        ("commitment_months", 2.5),
        ("assigned_week", 0),
        ("assigned_week", 6),
        ("assigned_week", "W2"),
        ("made_up_field", "anything"),
    ],
)
def test_invalid_values_rejected_before_touching_db_or_excel(tmp_path, field, bad_value):
    from backend.configuration.vendor_edits import FIELD_TO_EXCEL_HEADER, update_vendor_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)

    # Validation raises before session/Excel are ever touched, regardless of
    # session ownership — a caller-supplied session is fine here.
    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
        with pytest.raises(ValueError):
            update_vendor_field(vendor.id, field, bad_value, session=session, excel_path=excel_copy)

        assert session.query(AuditLog).filter_by(vendor_id=vendor.id).count() == 0
        header_texts = _header_texts(excel_copy)
        for header in FIELD_TO_EXCEL_HEADER.values():
            assert header.lower() not in header_texts  # no column was ever created
    finally:
        session.rollback()
        session.close()


def test_category_accepts_a_live_custom_category_added_via_configuration(tmp_path):
    """Category-configuration task: `_validate()`'s "category" branch used
    to be a fixed `VendorCategory(new_value)` enum lookup — a category
    Finance adds through Configuration (priority_bucket_edits.add_bucket)
    must be accepted here too, not just by the enum's 4 fixed members."""
    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    from backend.configuration.priority_bucket_edits import add_bucket
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    try:
        add_bucket("P6", "P6", 0.70, 0.15, 4, "Contractor", session=session)
        session.commit()
    finally:
        session.close()

    old, new = update_vendor_field(vendor_id, "category", "Contractor", excel_path=excel_copy)
    assert new == "Contractor"
    assert _read_cell(excel_copy, "Category", "V00400") == "Contractor"  # no fixed Excel-label translation, own name

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().category == "Contractor"
    finally:
        verify.close()


def test_assigned_week_validated_against_calendar_week_count(tmp_path):
    """CLAUDE.md rule 7's confirmed fix: assigned_week's valid range comes
    from the calendar (this cycle's actual week count), not a hardcoded 5.

    Real, unmodified state: the real ledger's latest month is May-2026 (31
    days -> 5 weeks), so W5 is accepted with no construction needed.
    Constructed: adding a later MonthlyLedger row for Feb-2099 (28 days ->
    4 weeks) makes that the new latest month across the whole DB — W5 must
    now be rejected while W4 is still accepted.
    """
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import MonthlyLedger, Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code="V00400").one()

        _, new = update_vendor_field(vendor.id, "assigned_week", 5, session=session, excel_path=excel_copy)
        assert new == 5  # 5-week month (real May-2026) accepts W5

        session.add(
            MonthlyLedger(
                vendor_id=vendor.id,
                month=date(2099, 2, 1),
                payable=0,
                payment=0,
                opening_balance=0,
                closing_balance=0,
            )
        )
        session.flush()

        with pytest.raises(ValueError):
            update_vendor_field(vendor.id, "assigned_week", 5, session=session, excel_path=excel_copy)

        _, new2 = update_vendor_field(vendor.id, "assigned_week", 4, session=session, excel_path=excel_copy)
        assert new2 == 4  # 4-week month (constructed Feb-2099) still accepts W4
    finally:
        session.rollback()
        session.close()


def test_owns_session_commits_then_publishes_in_order(tmp_path, monkeypatch):
    """Proof point (1): session=None commits, then publishes — in that order."""
    from backend.configuration.vendor_edits import update_vendor_field

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    call_order = []
    real_commit = SASession.commit

    def _tracked_commit(self):
        call_order.append("commit")
        return real_commit(self)

    real_replace = os.replace

    def _tracked_replace(src, dst):
        call_order.append("replace")
        return real_replace(src, dst)

    monkeypatch.setattr(SASession, "commit", _tracked_commit)
    monkeypatch.setattr(os, "replace", _tracked_replace)

    update_vendor_field(vendor_id, "commitment_months", 4, excel_path=excel_copy)
    assert call_order == ["commit", "replace"]


def test_owns_session_excel_stage_failure_rolls_back_and_never_publishes(tmp_path, monkeypatch):
    """Proof point (2), failure before commit is even attempted: staging the
    Excel edit fails -> rollback, no column ever created, no publish."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog, Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)

    def _raise_on_save(self, filename):
        raise OSError("simulated write-back failure")

    monkeypatch.setattr(openpyxl.Workbook, "save", _raise_on_save)

    vendor_id = _vendor_id("V00400")
    lookup = SessionLocal()
    original_category = lookup.query(Vendor).filter_by(id=vendor_id).one().category
    lookup.close()

    with pytest.raises(OSError):
        update_vendor_field(vendor_id, "category", VendorCategory.MUST_PAY, excel_path=excel_copy)

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().category == original_category  # never committed
        assert verify.query(AuditLog).filter_by(vendor_id=vendor_id).count() == 0
    finally:
        verify.close()

    header_texts = _header_texts(excel_copy)
    assert "category" not in header_texts  # never published


def test_owns_session_db_commit_failure_after_excel_staged_rolls_back_and_never_publishes(tmp_path, monkeypatch):
    """Proof point (2), failure after staging succeeds: the DB commit itself
    fails -> rollback, no publish. The mocked commit() never actually
    touches the temp DB, so this is safe either way."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    vendor_id = _vendor_id("V00400")

    def _raise_on_commit(self):
        raise RuntimeError("simulated DB commit failure")

    monkeypatch.setattr(SASession, "commit", _raise_on_commit)

    with pytest.raises(RuntimeError):
        update_vendor_field(vendor_id, "commitment_months", 9, excel_path=excel_copy)

    verify = SessionLocal()
    try:
        assert verify.query(Vendor).filter_by(id=vendor_id).one().commitment_months == 1  # temp DB untouched
    finally:
        verify.close()

    header_texts = _header_texts(excel_copy)
    assert "commitment months" not in header_texts  # staged edit was never published


def test_caller_supplied_session_failure_does_not_rollback_others_pending_changes(tmp_path):
    """Proof point (3): a failure inside update_vendor_field must not touch
    a caller-supplied session's OTHER pending, uncommitted work."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    session = SessionLocal()
    try:
        other_vendor = session.query(Vendor).filter_by(erp_code="V01542").one()
        other_vendor.commitment_months = 9  # unrelated pending change, not yet committed

        with pytest.raises(NoResultFound):
            update_vendor_field(-1, "category", VendorCategory.MUST_PAY, session=session, excel_path=excel_copy)

        assert other_vendor.commitment_months == 9  # survived — not rolled back
    finally:
        session.rollback()
        session.close()


def test_caller_supplied_session_never_publishes(tmp_path, monkeypatch):
    """Proof point (4): a caller-supplied session never calls os.replace(),
    under any circumstance — even on a successful edit."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    calls = []
    monkeypatch.setattr(os, "replace", lambda *args, **kwargs: calls.append((args, kwargs)))

    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
        old, new = update_vendor_field(vendor.id, "category", VendorCategory.MUST_PAY, session=session, excel_path=excel_copy)

        assert calls == []  # os.replace never called
        assert new == VendorCategory.MUST_PAY
        assert vendor.category == VendorCategory.MUST_PAY  # DB-side flush is visible in the shared session
        # Nothing was published — the real column must not exist in the Excel copy either.
        header_texts = _header_texts(excel_copy)
        assert "category" not in header_texts
    finally:
        session.rollback()
        session.close()


@pytest.mark.parametrize(
    "field,value1,value2",
    [
        ("category", VendorCategory.MUST_PAY, VendorCategory.COMMITMENT),
        ("commitment_months", 3, 6),
        ("assigned_week", 1, 4),
    ],
)
def test_column_reused_not_duplicated_across_vendors(tmp_path, field, value1, value2):
    """One shared code path (_find_or_create_column) covers all three
    fields — a second vendor's edit must reuse the column the first
    vendor's edit created, never duplicate it. Uses session=None for both
    edits since only an owning call ever touches Excel."""
    from backend.configuration.vendor_edits import FIELD_TO_EXCEL_HEADER, update_vendor_field

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    v1_id = _vendor_id("V00400")
    v2_id = _vendor_id("V01542")

    update_vendor_field(v1_id, field, value1, excel_path=excel_copy)
    update_vendor_field(v2_id, field, value2, excel_path=excel_copy)

    header_texts = list(_header_texts(excel_copy))
    assert header_texts.count(FIELD_TO_EXCEL_HEADER[field].lower()) == 1


def test_assigned_week_edit_reflected_in_build_weekly_view(tmp_path):
    """Confirms assigned_week's stored type (plain int 1-5 or None) matches
    exactly what planner.py already expects when grouping by week. Uses a
    caller-supplied session throughout — this only relies on the DB-side
    flush being visible within that same session, which update_vendor_field
    still does regardless of session ownership; nothing here checks Excel."""
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal
    from backend.models.new_model_2.allocator import generate_plan
    from backend.weekly_planning.planner import build_weekly_view

    master_path = _fresh_db_and_master(tmp_path)
    excel_copy = _copy_excel(tmp_path, master_path)
    session = SessionLocal()
    try:
        vendor = session.query(Vendor).filter_by(erp_code="V00400").one()
        assert vendor.assigned_week == 2  # real, confirmed baseline

        update_vendor_field(vendor.id, "assigned_week", 4, session=session, excel_path=excel_copy)
        assert vendor.assigned_week == 4

        plan = generate_plan(available_funds=100_000_000, session=session)
        view = build_weekly_view(plan["allocations"], session=session, persist=False)
        row = view["detail"][view["detail"]["vendor_id"] == vendor.id].iloc[0]
        assert row["assigned_week"] == 4  # picked up the edit correctly, grouped under the new week
    finally:
        session.rollback()
        session.close()
