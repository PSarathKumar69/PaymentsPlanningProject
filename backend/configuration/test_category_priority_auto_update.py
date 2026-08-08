"""Category<->Priority Tag auto-update on single-field UI edits (this task).

Reuses column_mapping.derive_category_and_priority_tag() (already built and
tested for ingestion-time mapping, backend/ingestion/test_category_priority_mapping.py)
— nothing here re-derives the mapping tables, only exercises
update_vendor_field()'s new wiring: DB write, Excel write-back (both
fields, one workbook cycle), audit_log (direct + derived, distinct
source), the Commitment/commitment_months edge case, and a live
generate_plan() before/after proof that the model actually plans
differently once a vendor's category/tag changes via a UI edit.

Same fresh-DB-per-test / synthetic-fixture convention as
test_priority_tag_mapping.py and test_category_priority_mapping.py — never
touches the real, externally-mutable master Excel (see memory:
demo15-dataset-quirks) or backend/db/app.db (unlike test_vendor_edits.py's
existing real-DB convention).

IMPORTANT ordering rule (memory: demo15-dataset-quirks' "stale module
identity" trap): _fresh_db_and_master(tmp_path) — which pops
backend.db.session/backend.ingestion.load_excel/backend.configuration.vendor_edits
from sys.modules before rebinding PAYMENTS_DB_PATH — must run BEFORE any
`from backend.configuration.vendor_edits import ...` (or anything that
transitively imports backend.db.session) in every test function. Importing
first would silently bind to the PREVIOUS parametrize case's (or a prior
test's) already-cached module, whose update_vendor_field()/SessionLocal
still point at that now-deleted temp DB — writes then "succeed" against a
stale engine while a fresh read-back sees nothing changed.
"""
import os
import shutil
import sys

import openpyxl
import pytest

_FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "..", "ingestion", "test_fixtures")
_FIXTURE = "single_header_row_sheet.xlsx"


def _fresh_db_and_master(tmp_path):
    """VT0001 = Category "Normal"/Priority Tag "P2", VT0002 = Category
    "Must Pay"/Priority Tag "P0" (fixture's own baseline — both agree, no
    ingestion-time conflict noise to account for here)."""
    os.environ["PAYMENTS_DB_PATH"] = str(tmp_path / "test.db")
    for mod in (
        "backend.db.session",
        "backend.ingestion.load_excel",
        "backend.configuration.vendor_edits",
    ):
        sys.modules.pop(mod, None)
    from backend.ingestion.load_excel import load

    master_path = str(tmp_path / "master.xlsx")
    shutil.copy(os.path.join(_FIXTURES_DIR, _FIXTURE), master_path)
    load(excel_path=master_path)
    return master_path


def _vendor_id(erp_code):
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    try:
        return session.query(Vendor).filter_by(erp_code=erp_code).one().id
    finally:
        session.close()


def _vendor(erp_code):
    from backend.db.models import Vendor
    from backend.db.session import SessionLocal

    session = SessionLocal()
    try:
        return session.query(Vendor).filter_by(erp_code=erp_code).one()
    finally:
        session.close()


def _read_cell(excel_path, header, erp_code):
    from backend.ingestion.column_mapping import build_sheet_map, find_column

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws)
    col = find_column(ws, header)
    assert col is not None, f"column {header!r} not found"
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value == erp_code:
            return ws.cell(row=row, column=col).value
    raise AssertionError(f"{erp_code} not found")


# ---- editing category derives priority_tag (all 4 cases) -------------------


@pytest.mark.parametrize(
    "category_value,expected_tag",
    [("must_pay", "P0"), ("commitment", "P1"), ("normal", "P2"), ("inactive", "P5")],
)
def test_edit_category_derives_priority_tag(tmp_path, category_value, expected_tag):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.shared.enums import VendorCategory

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0002")  # Must Pay/P0 baseline — every case here is a real transition

    result = update_vendor_field(vendor_id, "category", VendorCategory(category_value), excel_path=excel_copy)
    assert result.new_value == VendorCategory(category_value)

    vendor = _vendor("VT0002")
    assert vendor.category == VendorCategory(category_value)
    assert vendor.priority_tag == expected_tag

    if category_value == "must_pay":  # VT0002's existing baseline — a true no-op, nothing to derive
        assert result.sibling_update is None
    else:
        assert result.sibling_update == ("priority_tag", "P0", expected_tag)
    assert _read_cell(excel_copy, "Priority Tag", "VT0002") == expected_tag
    assert _read_cell(excel_copy, "Category", "VT0002") == {
        "must_pay": "Must Pay", "commitment": "Commitment", "normal": "Normal", "inactive": "Inactive",
    }[category_value]


# ---- editing priority_tag derives category (all recognized tags) ----------


@pytest.mark.parametrize(
    "tag_value,expected_category",
    [("P0", "must_pay"), ("P1", "commitment"), ("P2", "normal"), ("P3", "normal"), ("P4", "normal"), ("P5", "inactive")],
)
def test_edit_priority_tag_derives_category(tmp_path, tag_value, expected_category):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.shared.enums import VendorCategory

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")  # Normal/P2 baseline

    result = update_vendor_field(vendor_id, "priority_tag", tag_value, excel_path=excel_copy)
    assert result.new_value == tag_value

    vendor = _vendor("VT0001")
    assert vendor.priority_tag == tag_value
    assert vendor.category == VendorCategory(expected_category)

    if expected_category != "normal":  # "normal" is VT0001's existing category — no sibling change, see below
        assert result.sibling_update is not None
        assert result.sibling_update[0] == "category"
        assert _read_cell(excel_copy, "Category", "VT0001") == {
            "must_pay": "Must Pay", "commitment": "Commitment", "normal": "Normal", "inactive": "Inactive",
        }[expected_category]
    assert _read_cell(excel_copy, "Priority Tag", "VT0001") == tag_value


# ---- no sibling change when it already agrees -> no audit noise -----------


def test_no_sibling_audit_entry_when_already_agrees(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog
    from backend.db.session import SessionLocal

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")  # Normal/P2

    # P2 -> P3 both derive category=Normal — VT0001 is already Normal.
    result = update_vendor_field(vendor_id, "priority_tag", "P3", excel_path=excel_copy)
    assert result.sibling_update is None

    session = SessionLocal()
    try:
        derived_entries = session.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="category").all()
        assert derived_entries == []
    finally:
        session.close()


# ---- clearing priority_tag / a P6+ bucket tag derives nothing -------------


def test_clearing_priority_tag_derives_no_category_change(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")

    result = update_vendor_field(vendor_id, "priority_tag", None, excel_path=excel_copy)
    assert result.new_value is None
    assert result.sibling_update is None
    assert _vendor("VT0001").category == "normal"  # untouched


# ---- distinct audit_log source for the derived sibling change -------------


def test_sibling_audit_entry_has_distinct_derived_source(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.models import AuditLog
    from backend.db.session import SessionLocal
    from backend.shared.enums import ChangeSource

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")  # Normal/P2

    update_vendor_field(vendor_id, "priority_tag", "P1", excel_path=excel_copy)  # -> derives category=Commitment

    session = SessionLocal()
    try:
        direct = session.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="priority_tag").one()
        assert direct.source == ChangeSource.UI_EDIT.value

        derived = session.query(AuditLog).filter_by(vendor_id=vendor_id, field_name="category").one()
        assert derived.source == ChangeSource.UI_EDIT_DERIVED.value
        assert derived.source != direct.source
        assert derived.old_value == "Normal"
        assert derived.new_value == "Commitment"
    finally:
        session.close()


# ---- Commitment/commitment_months edge case: non-blocking warning ---------


def test_commitment_transition_with_unconfirmed_commitment_months_warns(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import COMMITMENT_MONTHS_UNCONFIRMED_DEFAULT, update_vendor_field
    from backend.db.session import SessionLocal
    from backend.db.models import Vendor

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")  # Normal/P2, commitment_months=1 (fixture) == the unconfirmed default

    session = SessionLocal()
    assert session.query(Vendor).filter_by(id=vendor_id).one().commitment_months == COMMITMENT_MONTHS_UNCONFIRMED_DEFAULT
    session.close()

    result = update_vendor_field(vendor_id, "priority_tag", "P1", excel_path=excel_copy)  # -> Commitment
    assert result.commitment_months_warning is not None
    assert "VT0001" in result.commitment_months_warning
    assert "commitment_months" in result.commitment_months_warning


def test_commitment_transition_no_warning_when_already_commitment(tmp_path):
    """No warning on a second edit that doesn't cross INTO Commitment — only
    the transition is flagged, not every edit to an already-Commitment vendor."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")

    update_vendor_field(vendor_id, "priority_tag", "P1", excel_path=excel_copy)  # Normal -> Commitment (warns)
    result = update_vendor_field(vendor_id, "commitment_months", 6, excel_path=excel_copy)  # already Commitment
    assert result.commitment_months_warning is None


def test_no_warning_when_commitment_months_already_confirmed(tmp_path):
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    vendor_id = _vendor_id("VT0001")

    update_vendor_field(vendor_id, "commitment_months", 6, excel_path=excel_copy)  # Finance sets a real figure first
    result = update_vendor_field(vendor_id, "priority_tag", "P1", excel_path=excel_copy)  # now transitions
    assert result.commitment_months_warning is None


# ---- live proof: generate_plan() actually plans differently ---------------


def test_generate_plan_before_after_category_edit_changes_treatment(tmp_path):
    """VT0001 starts Normal/P2 under a real funding shortfall (bucket-ceiling
    cut, NOT guaranteed). Edit priority_tag -> P1 (derives category
    Commitment). Regenerate: VT0001 must now be guaranteed at 100% of its
    OWN (Commitment-formula) required amount — a different number AND a
    different code path than before, not just a display change."""
    master_path = _fresh_db_and_master(tmp_path)
    from backend.configuration.vendor_edits import update_vendor_field
    from backend.db.session import SessionLocal
    from backend.models.new_model_2.allocator import generate_plan
    from backend.shared.enums import AllocationStatus

    session = SessionLocal()
    try:
        # Tight funds: covers only VT0002's guaranteed (Must Pay) requirement,
        # leaving VT0001 (Normal/P2) to be cut down by the bucket-ceiling math.
        before = generate_plan(available_funds=1.0, session=session)
    finally:
        session.close()
    before_row = before["allocations"].set_index("vendor_id").loc[_vendor_id("VT0001")]
    assert before_row["status"] != AllocationStatus.GUARANTEED.value
    required_before = before_row["required_amount"]
    assert before_row["allocated_amount"] < required_before  # bucket-ceiling cut it below 100%

    excel_copy = str(tmp_path / "edit.xlsx")
    shutil.copy(master_path, excel_copy)
    # commitment_months=2 first (a real, Finance-confirmed figure — not this
    # task's unconfirmed-default edge case, that's covered separately above)
    # so the two formulas actually diverge: VT0001's opening_balance is
    # 50000, and Normal's tranche-walk rule also happens to land on exactly
    # 50000 here (only one real tranche) — Commitment's opening_balance/
    # commitment_months amortization must use a DIFFERENT commitment_months
    # to prove a different formula engaged, not a coincidental same number.
    update_vendor_field(_vendor_id("VT0001"), "commitment_months", 2, excel_path=excel_copy)
    update_vendor_field(_vendor_id("VT0001"), "priority_tag", "P1", excel_path=excel_copy)  # -> Commitment

    session = SessionLocal()
    try:
        after = generate_plan(available_funds=1_000_000.0, session=session)
    finally:
        session.close()
    after_row = after["allocations"].set_index("vendor_id").loc[_vendor_id("VT0001")]
    assert after_row["status"] == AllocationStatus.GUARANTEED.value
    required_after = after_row["required_amount"]
    assert after_row["allocated_amount"] == pytest.approx(required_after)
    # Different formula engaged (Commitment's opening_balance/commitment_months
    # amortization vs. Normal's tranche-walk rule), not a coincidental match —
    # the whole point of this task's model-level verification.
    assert required_before != pytest.approx(required_after)
