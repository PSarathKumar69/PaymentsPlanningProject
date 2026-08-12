"""Vendor field edits — Configuration module, vendor-data half only
(docs/11-configuration-module.md). System-config editing (Model 1 weights,
Model 3 thresholds) and replacement-Excel-upload diffing are separate,
later work — not in scope here.

Every edit spans two systems: the DB and the real Excel file — but only
when this call *owns* its own session (no session= passed in) does it
touch Excel at all. Session ownership is the deciding line:

- Owns the session (the normal call, no session= argument): stage the
  Excel edit to a temp file, commit the DB, and only then publish the
  staged file (atomic os.replace) — narrowing the drift window to "DB
  committed, but the Excel replace hasn't happened or failed yet," rather
  than the older "Excel already saved, DB commit then fails" ordering. On
  any failure before a successful publish: roll back, never publish.
- Caller supplies the session (composing into a larger transaction): this
  function is DB-only. It flushes the vendor-field change and audit_log
  row into the shared session and returns — it never commits, never rolls
  back, and never touches the real Excel file. It has no way to know
  whether or when the caller's transaction will actually commit, and an
  Excel publish can't be undone the way a DB row can, so both are left
  entirely to the caller (any exception here propagates untouched, so the
  caller's own rollback covers this function's own pending changes too).

# RESIDUAL RISK — flagged, not fully solved: this is not a real two-phase
# commit across the DB and the filesystem. If the DB commit succeeds but
# the subsequent os.replace() fails or the process dies in that narrow
# window, the DB and Excel will disagree (DB updated, Excel still old)
# with nothing to reconcile them automatically. A full fix would need a
# write-ahead log / reconciliation pass across both systems, which is a
# bigger change than this pass calls for — flagging for Sarath rather than
# building it now. See test_db_commit_failure_after_excel_write_staged for
# what's proven today: a commit failure no longer leaves Excel ahead of
# the DB (the case that was previously untested).
"""
import os
import tempfile
from datetime import datetime

import openpyxl
from sqlalchemy import func

from backend.db.models import AuditLog, MonthlyLedger, Vendor, VendorExtraField
from backend.db.session import SessionLocal
from backend.ingestion import column_mapping_store
from backend.ingestion.column_mapping import (
    CATEGORY_EXCEL_LABEL,
    FIELD_TO_EXCEL_HEADER,
    build_sheet_map,
    derive_category_and_priority_tag,
    find_column,
    normalize_entity_text,
    resolve_header,
)
from backend.configuration.priority_bucket_edits import list_buckets
from backend.ingestion.load_excel import EXCEL_PATH, persist_excel_path_to_db, sync_excel_path_from_db
from backend.models.new_model_2.allocator import _seed_and_read_buckets
from backend.shared.enums import ChangeSource, VendorCategory, VendorPriorityTag
from backend.weekly_planning.calendar_utils import weeks_in_month

# DEFAULT — confirm with Sarath: category/commitment_months don't exist as
# columns in the real Excel yet, and assigned_week is currently packed with
# within-week-order into one legacy "W2-P3" column. Rather than re-packing
# edits into that legacy string format, this writes each field to its own
# clean, dedicated column (added to the sheet on first use if missing) and
# leaves the legacy combined column untouched — within_week_order is
# code-computed elsewhere and was never meant to be hand-edited anyway.
# (FIELD_TO_EXCEL_HEADER/CATEGORY_EXCEL_LABEL live in column_mapping.py, not
# here, so load_excel.py can read the same columns back on re-ingestion
# without a circular import — see that module's docstring.)


def _current_max_assigned_week(session):
    """How many week tags (W1-W<n>) are valid right now — derived from the
    calendar via the latest ingested ledger month (CLAUDE.md rule 7: never
    hardcode the week count to 5; some months only have 4). Reuses
    calendar_utils.weeks_in_month, the same function planner.py's own tests
    already exercise for this.
    """
    latest_month = session.query(func.max(MonthlyLedger.month)).scalar()
    if latest_month is None:
        return 5  # no ledger data at all yet (e.g. a fresh/empty DB) — fall back to the max possible
    return weeks_in_month(latest_month.year, latest_month.month)


def _validate(field, new_value, session=None):
    if field not in FIELD_TO_EXCEL_HEADER:
        raise ValueError(f"unknown field {field!r}; must be one of {sorted(FIELD_TO_EXCEL_HEADER)}")

    if field == "category":
        # Category-configuration task: used to be `VendorCategory(new_value)`
        # — hardcapping every edit to the fixed 4-value Python enum, so a
        # Configuration-added category (e.g. "Contractor") could never
        # actually be assigned to a vendor. Validated live against the two
        # fixed categories plus whatever's currently in the priority_buckets
        # table's category_name column instead — same DB-backed pattern the
        # "priority_tag" branch below already uses for its own bucket list.
        category = new_value.value if isinstance(new_value, VendorCategory) else str(new_value).strip()
        fixed = {VendorCategory.MUST_PAY.value, VendorCategory.COMMITMENT.value}
        live_categories = {b["category_name"] for b in list_buckets(session)}
        valid_categories = fixed | live_categories
        if category not in valid_categories:
            raise ValueError(f"invalid category {new_value!r}; must be one of {sorted(valid_categories)}")
        return category

    if field == "commitment_months":
        if isinstance(new_value, bool) or not isinstance(new_value, int) or new_value < 1:
            raise ValueError(f"commitment_months must be a positive integer, got {new_value!r}")
        return new_value

    if field == "priority_tag":
        if new_value is None:
            return None
        tag = new_value.value if isinstance(new_value, VendorPriorityTag) else str(new_value).strip().upper()
        # Bug fix (this task): used to validate against the fixed
        # VendorPriorityTag Python enum (P0-P5 only) — a Configuration-added
        # P6+ bucket (priority_bucket_edits.py::add_bucket) could never
        # actually be assigned to a vendor. Validated live against P0/P1
        # (fixed, auto-derived — never Finance-assignable in practice, but
        # kept permissive here exactly like the old enum was) plus whatever
        # buckets currently exist in the priority_buckets table — the same
        # DB-backed lookup allocator.py itself reads (_seed_and_read_buckets,
        # idempotent: seeds today's P2/P3/P4/P5 defaults on an empty table,
        # otherwise just reads back Configuration's current rows).
        auto_tags = {VendorPriorityTag.P0.value, VendorPriorityTag.P1.value}
        bucket_order, _, _, _, _ = _seed_and_read_buckets(session)
        valid_tags = auto_tags | set(bucket_order)
        if tag not in valid_tags:
            raise ValueError(f"invalid priority_tag {new_value!r}; must be one of {sorted(valid_tags)} or None")
        return tag

    # assigned_week
    if new_value is None:
        return None
    max_week = _current_max_assigned_week(session)
    if isinstance(new_value, bool) or not isinstance(new_value, int) or not (1 <= new_value <= max_week):
        raise ValueError(
            f"assigned_week must be an integer 1-{max_week} (this month's actual week count) or None, "
            f"got {new_value!r}"
        )
    return new_value


# Category<->Priority Tag auto-update (this task): editing one field
# through the UI must derive and write the sibling field too — same
# reconciliation column_mapping.derive_category_and_priority_tag() already
# does at ingestion time, reused here rather than re-deriving the mapping
# tables inline (one source of truth for this mapping, not two).
def _derive_sibling(field, new_value, session=None):
    """(sibling_field, sibling_value), or (None, None) if this edit derives
    nothing. `new_value` is the ALREADY-VALIDATED value for `field` (a
    plain category string for "category", a plain tag string/None for
    "priority_tag" — see _validate() above).

    - field == "category": always derives priority_tag (that category's own
      default bucket_key, Normal -> P2 — see column_mapping.py's
      derive_category_and_priority_tag()/_dynamic_category_maps()).
    - field == "priority_tag": derives category via the live priority_tag
      -> category_name mapping IF new_value is a recognized bucket_key
      (P0/P1 fixed, or any live priority_buckets row — Configuration-
      task fix: this used to be fixed at P0-P5 by design, so a
      Configuration-added custom bucket had no defined category to derive;
      now every live bucket carries its own category_name, so this works
      for any of them). Clearing the tag to None still derives nothing
      (nothing to derive category FROM).
    - any other field (commitment_months/assigned_week): no sibling.
    """
    if field == "category":
        category, priority_tag, _ = derive_category_and_priority_tag(new_value, None, session=session)
        return "priority_tag", priority_tag
    if field == "priority_tag":
        if new_value is None:
            return None, None
        category, priority_tag, _ = derive_category_and_priority_tag(None, new_value, session=session)
        if category is None:
            return None, None
        return "category", category
    return None, None


# The Commitment-category required-amount formula (min_funds_v2.py) uses
# `vendor.commitment_months or 1` — 1 is also the Vendor.commitment_months
# DB column's own default (models.py), so a vendor that just became
# Commitment (directly, or via priority_tag deriving it) almost always
# still carries this UNCONFIRMED value, not one Finance actually reviewed.
COMMITMENT_MONTHS_UNCONFIRMED_DEFAULT = 1


def _commitment_months_warning(vendor, old_category, new_category):
    """Non-blocking flag (this task's documented decision — see module
    docstring below `update_vendor_field()`) for the real edge case where an
    edit moves a vendor INTO Commitment while its commitment_months is still
    the unconfirmed default: required_amount_v2() will silently use that 1-
    month amortization until Finance sets a real value. None when the
    vendor was already Commitment before this edit (nothing new to flag),
    or when commitment_months already looks like a real, confirmed figure."""
    if new_category != VendorCategory.COMMITMENT or old_category == VendorCategory.COMMITMENT:
        return None
    if vendor.commitment_months in (None, COMMITMENT_MONTHS_UNCONFIRMED_DEFAULT):
        return (
            f"{vendor.erp_code} is now Commitment category, but commitment_months is "
            f"{vendor.commitment_months!r} (unconfirmed default) — the required-amount amortization formula "
            "will use this value until Finance sets it explicitly."
        )
    return None


class VendorFieldEditResult(tuple):
    """update_vendor_field()'s return value. Behaves as the plain
    (old_value, new_value) 2-tuple every existing caller already unpacks
    (`old, new = update_vendor_field(...)`) — changing that shape would
    touch a dozen unrelated call sites for a feature only the PATCH router
    needs. New callers (backend/api/routers/vendors.py's patch_vendor())
    read `.sibling_update` / `.commitment_months_warning` directly instead.

    sibling_update: None, or (sibling_field, old_value, new_value) when
    this edit actually changed the derived sibling (skipped when the
    sibling already agreed — e.g. editing priority_tag P2->P3 derives
    category=Normal, which the vendor already had).
    commitment_months_warning: see _commitment_months_warning() above.
    """

    def __new__(cls, old_value, new_value, sibling_update=None, commitment_months_warning=None):
        obj = super().__new__(cls, (old_value, new_value))
        obj.sibling_update = sibling_update
        obj.commitment_months_warning = commitment_months_warning
        return obj

    @property
    def old_value(self):
        return self[0]

    @property
    def new_value(self):
        return self[1]


def _find_or_create_column(ws, header_text, header_row):
    col = find_column(ws, header_text, header_row)
    if col is not None:
        return col
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col, value=header_text)
    return new_col


def _find_vendor_row(ws, sheet_map, erp_code, entity):
    """Entity-code-collision fix: erp_code alone can now match more than one
    Excel row (two different legal entities sharing a bare code, e.g. ARS/FP
    both "V00149") — matched on (entity, erp_code) together, normalized the
    same way load_excel.py's upsert does, so an edit/export always lands on
    the correct vendor's own row, never a same-code sibling's."""
    target_entity = normalize_entity_text(entity)
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        if ws.cell(row=row, column=sheet_map.erp_code_col).value != erp_code:
            continue
        if normalize_entity_text(ws.cell(row=row, column=sheet_map.entity_col).value) == target_entity:
            return row
    raise ValueError(f"vendor {erp_code!r} (entity {entity!r}) not found in the Excel sheet")


def build_vendor_row_index(ws, sheet_map):
    """Batched form of _find_vendor_row() above — one O(n) scan of the
    sheet building a full (normalized_entity, erp_code) -> row lookup,
    instead of calling _find_vendor_row() once per vendor inside a loop
    (each call is its own O(n) scan -> O(n^2) total). Perf fix (prod bug,
    confirmed 2026-08: finalized-plan-export taking 10+ seconds for ~450
    vendors — this quadratic openpyxl cell-scan was part of it, alongside
    the separate per-vendor DB round-trips fixed elsewhere in that same
    endpoint).

    Single-vendor callers (this module's own update_vendor_field()/
    update_extra_field(), one edit at a time) keep calling
    _find_vendor_row() directly — building this whole index for a single
    lookup would be strictly more work, not less. Use this one instead
    only where the same sheet is being looked up for many vendors in one
    request (e.g. new_model_2.py's finalized-plan-export).

    First occurrence wins on a duplicate (entity, erp_code) pair — same
    "first-row-wins" convention load_excel.py's own upsert and
    find_duplicate_erp_codes() already use, so a lookup here always
    resolves to the same row load() would have kept."""
    index = {}
    for row in range(sheet_map.data_start_row, ws.max_row + 1):
        erp_code = ws.cell(row=row, column=sheet_map.erp_code_col).value
        if erp_code is None:
            continue
        entity = ws.cell(row=row, column=sheet_map.entity_col).value
        index.setdefault((normalize_entity_text(entity), erp_code), row)
    return index


def _excel_cell_value(field, new_value):
    if field == "category":
        # .get(..., new_value): a custom category (Configuration-task) has
        # no separate Excel-label translation — Finance's typed name IS
        # already the display text, unlike Must Pay/Commitment/Normal/
        # Inactive's historical value-vs-label split.
        return CATEGORY_EXCEL_LABEL.get(new_value, new_value)
    # priority_tag is already a plain string (_validate above never returns
    # an enum instance) — falls through to the generic return below, same
    # as commitment_months/assigned_week.
    return new_value  # commitment_months: int; assigned_week: int or None (clears the cell); priority_tag: str or None


def _clean_value(field, value):
    """Human-readable audit_log text — never a raw Python repr like
    "VendorCategory.NORMAL". Reuses the same Finance-facing label the
    Excel cell gets, so the audit log and the Excel file always agree on
    what a value "looks like" (one shared label per field, not two)."""
    if value is None:
        return None
    if field == "category":
        return CATEGORY_EXCEL_LABEL.get(value, value)
    # priority_tag is already a plain string — falls through to str(value)
    # below (a no-op for it), same as commitment_months/assigned_week.
    return str(value)


def _save_workbook_to_temp(wb, excel_path):
    """Shared tail of every staging helper below: saves `wb` to a fresh temp
    file (same directory as excel_path, so the later os.replace is
    same-filesystem and atomic) and returns its path. Nothing is published
    to excel_path yet — that's always the caller's job, after its own DB
    commit succeeds."""
    directory = os.path.dirname(os.path.abspath(excel_path)) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=directory)
    os.close(fd)
    wb.save(tmp_path)
    return tmp_path


def _stage_excel_cell_write(header_text, erp_code, entity, cell_value, excel_path, header_overrides=None, sheet_start_month=None):
    """Generic version of the staging step: writes one (header, vendor row)
    cell to a fresh temp file. Shared by update_extra_field() (data-
    pipeline-upload task's generic passthrough columns, header is the
    literal Excel text) — _stage_field_writes() below is the known-field
    equivalent for update_vendor_field(), writing one or more cells in a
    single workbook load/save cycle.

    header_overrides/sheet_start_month (P1 demo-readiness fix): without
    these, build_sheet_map() below falls back to the fixed default headers/
    Apr-2025 start — if the real sheet needed an AI-resolved header for a
    REQUIRED field (erp_code/entity/vendor_name/opening_balance/
    total_payable/total_payment), that raised MissingRequiredColumnsError
    here even though ingestion itself worked fine."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws, header_overrides=header_overrides, sheet_start_month=sheet_start_month)

    col = _find_or_create_column(ws, header_text, sheet_map.header_row)
    row = _find_vendor_row(ws, sheet_map, erp_code, entity)
    # openpyxl's ws.cell(..., value=None) is a no-op (None means "getter
    # mode", not "clear") — direct attribute assignment is required to
    # actually blank a cell (needed when assigned_week is cleared to None).
    ws.cell(row=row, column=col).value = cell_value

    return _save_workbook_to_temp(wb, excel_path)


def _stage_field_writes(erp_code, entity, field_values, excel_path, header_overrides=None, sheet_start_month=None):
    """Known-field (category/commitment_months/assigned_week/priority_tag)
    staging — writes ONE OR MORE fields for the same vendor row in a single
    workbook load/save cycle. Needed since the Category<->Priority Tag
    auto-update task (this task): editing one field can derive a change to
    its sibling too, and each field has its own dedicated Excel column
    (CLAUDE.md rule 6) — two separate load-modify-save calls would each
    reopen the ORIGINAL file fresh, so the second call's write would
    silently discard the first's. field_values: [(field, new_value), ...],
    each new_value already validated/derived, not yet converted to its
    Excel cell representation (done here via _excel_cell_value()). Each
    field's OWN header is resolved through header_overrides too (these are
    AI-mappable fields, column_mapping.ALL_MAPPABLE_FIELDS) rather than
    always the fixed FIELD_TO_EXCEL_HEADER default, so an edit lands on the
    same column ingestion/the grid already resolved it to."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sheet_map = build_sheet_map(ws, header_overrides=header_overrides, sheet_start_month=sheet_start_month)
    row = _find_vendor_row(ws, sheet_map, erp_code, entity)

    for field, new_value in field_values:
        header_text = resolve_header(field, header_overrides)
        col = _find_or_create_column(ws, header_text, sheet_map.header_row)
        ws.cell(row=row, column=col).value = _excel_cell_value(field, new_value)

    return _save_workbook_to_temp(wb, excel_path)


def update_vendor_field(vendor_id, field, new_value, source=ChangeSource.UI_EDIT, session=None, excel_path=None):
    """Validates, updates the DB, records an audit_log row, and — only if
    this call owns its own session — writes back to the real Excel.

    Owns the session (session=None, the normal call): stage the Excel
    edit(s), commit the DB, then publish the staged file (os.replace) — in
    that order. A failure anywhere before a successful publish rolls back
    and never publishes (see module docstring for the residual-risk window
    this does and doesn't cover).

    Caller supplies `session`: DB-only. Flushes the change(s) into the
    shared session and returns; never commits, never rolls back, never
    touches Excel — the caller owns that transaction and must handle its
    own commit/rollback (and, for now, its own Excel write-back).

    Category<->Priority Tag auto-update (this task): editing "category" or
    "priority_tag" derives and writes the SIBLING field too (DB + its own
    Excel column + its own audit_log entry, source=UI_EDIT_DERIVED) — see
    _derive_sibling(). Skipped entirely when the sibling already agrees
    (no-op, no audit noise). The Commitment/commitment_months edge case
    (docs/14, min_funds_v2.py) is flagged non-blocking rather than required
    in the same request or silently ignored: PATCH /vendors/{id} is a
    single-field edit by design (VendorPatchRequest carries one field/value)
    — requiring commitment_months bundled into the same request would break
    that contract and every existing caller for an edge case that's still
    real money math either way. A warning that Finance can act on (or
    ignore, if 1 really is correct) matches this codebase's existing
    "flag, never silently block" convention (sheet_start_month_warning(),
    the category/priority ingestion conflict log) far better than a hard
    requirement would.

    Returns a VendorFieldEditResult — unpacks as (old_value, new_value)
    exactly like before this task; new callers read
    `.sibling_update`/`.commitment_months_warning` off it directly.
    """
    excel_path = excel_path or EXCEL_PATH

    owns_session = session is None
    session = session or SessionLocal()
    tmp_path = None
    try:
        # assigned_week's valid range depends on the calendar (this month's
        # actual week count), so validation needs a session now — moved
        # inside the try so a validation failure is cleaned up the same way
        # as any other failure below (rollback if this call owns the
        # session, always close if it does).
        new_value = _validate(field, new_value, session=session)
        vendor = session.query(Vendor).filter_by(id=vendor_id).one()
        old_category = vendor.category  # captured before any mutation below

        old_value = getattr(vendor, field)
        setattr(vendor, field, new_value)
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=vendor_id,
                field_name=field,
                old_value=_clean_value(field, old_value),
                new_value=_clean_value(field, new_value),
                source=source.value,
            )
        )

        sibling_field, sibling_value = _derive_sibling(field, new_value, session=session)
        sibling_update = None
        if sibling_field is not None and getattr(vendor, sibling_field) != sibling_value:
            old_sibling_value = getattr(vendor, sibling_field)
            setattr(vendor, sibling_field, sibling_value)
            session.add(
                AuditLog(
                    timestamp=datetime.now(),
                    vendor_id=vendor_id,
                    field_name=sibling_field,
                    old_value=_clean_value(sibling_field, old_sibling_value),
                    new_value=_clean_value(sibling_field, sibling_value),
                    # Never `source` (that's the DIRECT edit's own origin) —
                    # this is a system-derived side effect, always tagged so
                    # regardless of who/what triggered the direct edit.
                    source=ChangeSource.UI_EDIT_DERIVED.value,
                )
            )
            sibling_update = (sibling_field, old_sibling_value, sibling_value)

        new_category = new_value if field == "category" else (sibling_value if sibling_field == "category" else vendor.category)
        commitment_months_warning = _commitment_months_warning(vendor, old_category, new_category)

        session.flush()  # surface any DB-side error before touching Excel at all

        if not owns_session:
            # Caller owns the transaction: DB-only, nothing published.
            return VendorFieldEditResult(old_value, new_value, sibling_update, commitment_months_warning)

        header_overrides = column_mapping_store.load_overrides(session)
        sheet_start_month = column_mapping_store.get_sheet_start_month(session)
        field_writes = [(field, new_value)]
        if sibling_update is not None:
            field_writes.append((sibling_field, sibling_value))
        sync_excel_path_from_db()  # Vercel fix — this container may be warm from before the latest upload/edit
        tmp_path = _stage_field_writes(
            vendor.erp_code, vendor.entity, field_writes, excel_path, header_overrides, sheet_start_month
        )
        session.commit()  # publish only happens below, after this succeeds
        os.replace(tmp_path, excel_path)
        persist_excel_path_to_db()  # Vercel fix — durable copy, so other containers see this edit too
        tmp_path = None  # published — nothing left to clean up
        return VendorFieldEditResult(old_value, new_value, sibling_update, commitment_months_warning)
    except Exception:
        if owns_session:
            # Caller-supplied sessions are the caller's transaction to roll
            # back or not — rolling it back here would silently discard any
            # other pending, uncommitted work they'd already added to it.
            session.rollback()
        raise
    finally:
        if tmp_path is not None:
            os.remove(tmp_path)  # commit (or something before it) failed — discard the staged edit
        if owns_session:
            session.close()


def update_extra_field(vendor_id, column_name, new_value, source=ChangeSource.UI_EDIT, session=None, excel_path=None):
    """Generic-passthrough-column edit (data-pipeline-upload task,
    docs/11) — same DB-write/audit-log/atomic-Excel-write-back contract as
    update_vendor_field() above, extended to an arbitrary column position
    (VendorExtraField.column_name, the literal Excel header text) instead of
    one of the three hardcoded known fields. No validation against a known
    set — by definition this is a column the mapping layer doesn't
    recognize (Part B of that task); whatever Finance typed/picked is
    stored and written back as-is (never guessed at semantically).
    new_value is always coerced to text (None clears it).

    Same session-ownership contract as update_vendor_field(): owns its own
    session (session=None) -> stage, commit, publish, in that order; a
    caller-supplied session is DB-only, no Excel touch, no commit — the
    caller owns that transaction.

    Returns (old_value, new_value).
    """
    excel_path = excel_path or EXCEL_PATH
    new_value = None if new_value is None else str(new_value)

    owns_session = session is None
    session = session or SessionLocal()
    tmp_path = None
    try:
        vendor = session.query(Vendor).filter_by(id=vendor_id).one()
        extra_row = session.query(VendorExtraField).filter_by(vendor_id=vendor_id, column_name=column_name).first()
        old_value = extra_row.value if extra_row is not None else None
        if extra_row is None:
            session.add(VendorExtraField(vendor_id=vendor_id, column_name=column_name, value=new_value))
        else:
            extra_row.value = new_value
        session.add(
            AuditLog(
                timestamp=datetime.now(),
                vendor_id=vendor_id,
                field_name=column_name,
                old_value=old_value,
                new_value=new_value,
                source=source.value,
            )
        )
        session.flush()  # surface any DB-side error before touching Excel at all

        if not owns_session:
            return old_value, new_value

        header_overrides = column_mapping_store.load_overrides(session)
        sheet_start_month = column_mapping_store.get_sheet_start_month(session)
        sync_excel_path_from_db()  # Vercel fix — this container may be warm from before the latest upload/edit
        tmp_path = _stage_excel_cell_write(
            column_name, vendor.erp_code, vendor.entity, new_value, excel_path, header_overrides, sheet_start_month
        )
        session.commit()  # publish only happens below, after this succeeds
        os.replace(tmp_path, excel_path)
        persist_excel_path_to_db()  # Vercel fix — durable copy, so other containers see this edit too
        tmp_path = None  # published — nothing left to clean up
        return old_value, new_value
    except Exception:
        if owns_session:
            session.rollback()
        raise
    finally:
        if tmp_path is not None:
            os.remove(tmp_path)
        if owns_session:
            session.close()
