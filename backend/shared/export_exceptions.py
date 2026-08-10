"""Shared "Exceptions" tab builder for the two Excel exports New Model 2
offers (backend/api/routers/new_model_2.py's min-funds-verification-export
and finalized-plan-export) — investigated, not guessed: see the
investigation report this task followed for why each category below is
in scope and why a second WORKSHEET (not footer rows on the main sheet)
is the safe way to attach it.

Genuine anomalies only (real, unresolved money/data problems) — never
routine, expected exclusions. A zero-outstanding-balance vendor is the
one routine case, and it gets a single neutral count line, not a row per
vendor. Every category here composes from functions that already exist
elsewhere in this codebase — no new arithmetic:

  - Duplicate ERP-code groups: column_mapping.find_duplicate_erp_codes()
    against a read-only load of the live master sheet — the one vendor(s)
    silently dropped at ingestion (first-row-wins) never becomes a Vendor
    row, so there is nothing to find in the DB itself.
  - Required-amount capped: min_funds_v2.min_funds_breakdown_v2()'s
    returned rule/oldest_amount/second_amount/current_amount vs total —
    _capped() only ever mutates `total`, so comparing it against the
    pre-cap component amounts tells us whether the cap actually fired,
    with no separate calculation.
  - v2_no_history_before_as_of: same breakdown's own `rule` field.
  - Inactive vendors with real money still owed
    (Vendor.is_active is False and live_outstanding_balance() > 0).
"""
from openpyxl.styles import Font, PatternFill

from backend.db.models import Vendor
from backend.ingestion import column_mapping_store
from backend.ingestion.column_mapping import build_sheet_map, find_duplicate_erp_codes
from backend.shared.constants import MONEY_EPSILON
from backend.shared.min_funds import _load_vendors_and_ledger, live_outstanding_balance
from backend.shared.min_funds_v2 import min_funds_breakdown_v2

_COLUMNS = ["Vendor Name", "ERP Code", "Entity", "Category", "Reason"]

# Rules whose pre-cap raw total is the sum/lookup of these specific detail
# fields — mirrors min_funds_v2._compute_v2()'s own branches exactly (the
# constituent amounts it returns are never re-capped, only `total` is).
_RAW_TOTAL_FIELDS = {
    "v2_oldest_and_current": ("oldest_amount", "second_amount"),
    "v2_oldest_and_second": ("oldest_amount", "second_amount"),
    "v2_oldest_only": ("oldest_amount",),
    "v2_current_zero_oldest_only": ("oldest_amount",),
    "v2_only_current": ("current_amount",),
}


def build_exceptions(session, ws_for_duplicates, as_of):
    """Returns {"rows": [...], "routine_excluded_count": int}. Each row is
    {vendor_name, erp_code, entity, category, reason} — plain strings,
    ready to write straight into a sheet.

    `ws_for_duplicates` is an already-open openpyxl worksheet of the live
    master sheet (read-only use only — find_duplicate_erp_codes() never
    mutates it). Callers that already have one open (finalized-plan-export)
    should pass it directly; a caller with none open yet
    (min-funds-verification-export) opens EXCEL_PATH read-only just for
    this.
    """
    rows = []

    header_overrides = column_mapping_store.load_overrides(session)
    sheet_start_month = column_mapping_store.get_sheet_start_month(session)
    sheet_map = build_sheet_map(
        ws_for_duplicates, header_overrides=header_overrides, sheet_start_month=sheet_start_month
    )
    for group in find_duplicate_erp_codes(ws_for_duplicates, sheet_map):
        winner = group["rows"][0]
        for skipped in group["rows"][1:]:
            rows.append(
                {
                    "vendor_name": skipped["vendor_name"],
                    "erp_code": group["erp_code"],
                    "entity": group["entity"],
                    "category": "Duplicate ERP code — dropped at ingestion",
                    "reason": (
                        f"Same entity+ERP-code ({group['entity']}/{group['erp_code']}) as "
                        f"{winner['vendor_name']!r} (row {winner['row']}, kept) — this row "
                        f"(row {skipped['row']}) was skipped entirely at upload and never became "
                        "a vendor record. Assign a distinct ERP code to resolve."
                    ),
                }
            )

    vendors, ledger_by_vendor = _load_vendors_and_ledger(session)
    for vendor in vendors:
        detail = min_funds_breakdown_v2(vendor, ledger_by_vendor[vendor.id], as_of=as_of)
        rule = detail["rule"]

        if rule == "v2_no_history_before_as_of":
            rows.append(
                {
                    "vendor_name": vendor.vendor_name,
                    "erp_code": vendor.erp_code,
                    "entity": vendor.entity,
                    "category": "No ledger history before planning month",
                    "reason": (
                        f"Planning month's aging reference ({as_of}) predates this vendor's earliest "
                        f"recorded ledger row — showing the nearest real bill ({detail['current_month']}, "
                        f"Rs {detail['current_amount']:.2f}) as a best-effort default."
                    ),
                }
            )
            continue

        raw_fields = _RAW_TOTAL_FIELDS.get(rule)
        if raw_fields is None:
            continue  # commitment_outstanding_over_1_months, v2_no_outstanding — never capped
        raw_total = sum(detail[f] or 0.0 for f in raw_fields)
        if raw_total > detail["total"] + MONEY_EPSILON:
            rows.append(
                {
                    "vendor_name": vendor.vendor_name,
                    "erp_code": vendor.erp_code,
                    "entity": vendor.entity,
                    "category": "Required amount capped",
                    "reason": (
                        f"Min Funds Required capped at live outstanding balance Rs {detail['total']:.2f} "
                        f"(uncapped tranche total would have been Rs {raw_total:.2f}) — a negative-payment "
                        "correction or rounding drift in the ledger."
                    ),
                }
            )

    inactive_with_balance = [
        v for v in session.query(Vendor).filter(Vendor.is_active == False).order_by(Vendor.id).all()  # noqa: E712
        if live_outstanding_balance(v) > MONEY_EPSILON
    ]
    for vendor in inactive_with_balance:
        rows.append(
            {
                "vendor_name": vendor.vendor_name,
                "erp_code": vendor.erp_code,
                "entity": vendor.entity,
                "category": "Inactive vendor with unresolved balance",
                "reason": (
                    f"No longer active in the latest sheet, but still shows "
                    f"Rs {live_outstanding_balance(vendor):.2f} outstanding — real unresolved money on a "
                    "vendor not in the active plan."
                ),
            }
        )

    all_active = session.query(Vendor).filter(Vendor.is_active.isnot(False)).all()
    routine_excluded_count = sum(1 for v in all_active if live_outstanding_balance(v) <= MONEY_EPSILON)

    return {"rows": rows, "routine_excluded_count": routine_excluded_count}


_HEADER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ROW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ROW_FONT = Font(color="9C0006")
_NEUTRAL_FONT = Font(italic=True, color="595959")
_COLUMN_WIDTHS = (32, 12, 10, 34, 95)


def write_exceptions_sheet(wb, exceptions):
    """Adds a new "Exceptions" WORKSHEET (never extra rows on an existing
    sheet — see this module's own docstring for why) to `wb`, red-styled,
    and returns it. Genuine anomaly rows first, then one neutral (not red)
    summary line for the routine zero-balance exclusion count.
    """
    ws = wb.create_sheet("Exceptions")

    for col, header in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    for row in exceptions["rows"]:
        for col, key in enumerate(("vendor_name", "erp_code", "entity", "category", "reason"), start=1):
            cell = ws.cell(row=row_idx, column=col, value=row[key])
            cell.fill = _ROW_FILL
            cell.font = _ROW_FONT
        row_idx += 1

    if not exceptions["rows"]:
        ws.cell(row=row_idx, column=1, value="No exceptions found.").font = _NEUTRAL_FONT
        row_idx += 1

    row_idx += 1  # blank separator before the neutral summary line
    ws.cell(
        row=row_idx,
        column=1,
        value=(
            f"{exceptions['routine_excluded_count']} vendor(s) excluded from the main sheet — "
            "zero outstanding balance (routine)."
        ),
    ).font = _NEUTRAL_FONT

    for col, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    return ws
