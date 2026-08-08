"""Tests against the real ingested database + the real master Excel file —
no synthetic data, always session.rollback()'d, never committed. Same
convention as backend/configuration/test_priority_bucket_edits.py.

The investigation behind this module found exactly two live anomalies in
the current real data (V00393's defensive-cap hit, and the ARS "INC"
Phyllo Inc./PrimeRole, Inc. duplicate pair) — these tests pin that down as
a real-data regression check, not a hypothetical. If the real sheet
changes enough to resolve either one, these tests are the signal that
happened; update them rather than deleting them blind.
"""
import openpyxl

from backend.db.session import SessionLocal
from backend.ingestion.load_excel import EXCEL_PATH
from backend.shared.export_exceptions import build_exceptions, write_exceptions_sheet
from backend.shared.min_funds import live_outstanding_balance
from backend.db.models import Vendor


def _real_exceptions(session, as_of):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    return build_exceptions(session, wb.active, as_of)


def test_flags_the_real_ars_inc_duplicate_pair():
    session = SessionLocal()
    try:
        exceptions = _real_exceptions(session, as_of=None)
        row = next((r for r in exceptions["rows"] if r["erp_code"] == "INC" and r["entity"] == "ARS"), None)
        assert row is not None, "the confirmed real ARS/INC duplicate pair should always appear"
        assert row["vendor_name"] == "PrimeRole, Inc."  # the skipped (dropped) row — row 140's Phyllo Inc. is the winner
        assert row["category"] == "Duplicate ERP code — dropped at ingestion"
        assert "Phyllo Inc." in row["reason"]
        assert "skipped entirely" in row["reason"]
    finally:
        session.rollback()
        session.close()


def test_flags_the_real_v00393_capped_case():
    session = SessionLocal()
    try:
        exceptions = _real_exceptions(session, as_of=None)
        row = next((r for r in exceptions["rows"] if r["erp_code"] == "V00393"), None)
        assert row is not None, "the confirmed real V00393 defensive-cap hit should always appear"
        assert row["category"] == "Required amount capped"
        assert "capped at live outstanding balance" in row["reason"]
        assert "uncapped tranche total would have been" in row["reason"]
    finally:
        session.rollback()
        session.close()


def test_routine_excluded_count_matches_zero_balance_active_vendors():
    session = SessionLocal()
    try:
        exceptions = _real_exceptions(session, as_of=None)
        active = session.query(Vendor).filter(Vendor.is_active.isnot(False)).all()
        expected = sum(1 for v in active if live_outstanding_balance(v) <= 0.01)
        assert exceptions["routine_excluded_count"] == expected
    finally:
        session.rollback()
        session.close()


def test_exact_exception_row_count_against_todays_real_data():
    """Pins the exact current real-data shape down, not just "at least
    these two" — fails loudly (not silently) if a THIRD anomaly ever
    appears and nobody looked at it."""
    session = SessionLocal()
    try:
        exceptions = _real_exceptions(session, as_of=None)
        categories = sorted(r["category"] for r in exceptions["rows"])
        assert categories == ["Duplicate ERP code — dropped at ingestion", "Required amount capped"]
    finally:
        session.rollback()
        session.close()


def test_write_exceptions_sheet_is_a_second_worksheet_not_extra_rows():
    """The whole point of this task's design choice: appending must never
    touch an existing sheet's own used range."""
    session = SessionLocal()
    try:
        exceptions = _real_exceptions(session, as_of=None)
    finally:
        session.rollback()
        session.close()

    wb = openpyxl.Workbook()
    main_ws = wb.active
    main_ws["A1"] = "existing data"
    main_ws["B1"] = "more existing data"
    original_max_row = main_ws.max_row
    original_title = main_ws.title

    exceptions_ws = write_exceptions_sheet(wb, exceptions)

    assert wb.sheetnames == [original_title, "Exceptions"]
    assert main_ws.max_row == original_max_row
    assert main_ws["A1"].value == "existing data"
    assert wb.active.title == original_title  # adding the tab never changes which sheet is active
    assert exceptions_ws.title == "Exceptions"


def test_write_exceptions_sheet_styles_rows_red_and_summary_neutral():
    fake_exceptions = {
        "rows": [
            {
                "vendor_name": "Test Vendor",
                "erp_code": "T001",
                "entity": "TE",
                "category": "Required amount capped",
                "reason": "test reason",
            }
        ],
        "routine_excluded_count": 5,
    }
    wb = openpyxl.Workbook()
    ws = write_exceptions_sheet(wb, fake_exceptions)

    header_cell = ws.cell(row=1, column=1)
    row_cell = ws.cell(row=2, column=1)
    assert header_cell.value == "Vendor Name"
    assert header_cell.fill.fgColor.rgb == "00C00000"
    assert row_cell.value == "Test Vendor"
    assert row_cell.fill.fgColor.rgb == "00FFC7CE"  # red-tinted, not the neutral summary style

    summary_row = ws.max_row
    summary_cell = ws.cell(row=summary_row, column=1)
    assert "5 vendor(s) excluded" in summary_cell.value
    assert "zero outstanding balance (routine)" in summary_cell.value
    # Neutral, not red — no fill set, italic grey font instead.
    assert summary_cell.fill.fgColor.rgb in ("00000000", None)
    assert summary_cell.font.color.rgb == "00595959"


def test_write_exceptions_sheet_with_no_rows_says_so_and_still_shows_summary():
    empty_exceptions = {"rows": [], "routine_excluded_count": 0}
    wb = openpyxl.Workbook()
    ws = write_exceptions_sheet(wb, empty_exceptions)
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "No exceptions found." in values
    assert any("0 vendor(s) excluded" in (v or "") for v in values)


if __name__ == "__main__":
    # ponytail: smallest runnable self-check, no framework/fixtures.
    test_flags_the_real_ars_inc_duplicate_pair()
    test_flags_the_real_v00393_capped_case()
    test_routine_excluded_count_matches_zero_balance_active_vendors()
    test_exact_exception_row_count_against_todays_real_data()
    test_write_exceptions_sheet_is_a_second_worksheet_not_extra_rows()
    test_write_exceptions_sheet_styles_rows_red_and_summary_neutral()
    test_write_exceptions_sheet_with_no_rows_says_so_and_still_shows_summary()
    print("all export_exceptions self-checks passed")
