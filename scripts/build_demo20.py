"""One-off: build a 20-real-vendor demo Excel from the real master sheet.

Reads data/Vendor's Details.xlsx (read-only, data_only=True so formula cells
give their last-cached literal value instead of the formula string — the
source sheet has 300+ formulas including cross-workbook VLOOKUPs that would
otherwise re-save with no cached value and read back as None, per
load_excel.py's own note on this exact openpyxl behavior).

Copies header rows 1-3 (including the two merged header-block cells in row 2)
and the first 20 real vendor data rows (rows 4-23) verbatim as literal
values + number formats + column widths into a new workbook.

Not part of the app — run once, then discard.
"""
import copy

import openpyxl

SRC = "data/Vendor's Details.xlsx"
DST = "data/Vendor's Details (Demo 20).xlsx"
HEADER_ROWS = 3
DEMO_VENDOR_ROWS = 20

src_wb = openpyxl.load_workbook(SRC, data_only=True)
src_ws = src_wb.active

dst_wb = openpyxl.Workbook()
dst_ws = dst_wb.active
dst_ws.title = src_ws.title

last_row = HEADER_ROWS + DEMO_VENDOR_ROWS
for row in range(1, last_row + 1):
    for col in range(1, src_ws.max_column + 1):
        src_cell = src_ws.cell(row=row, column=col)
        dst_cell = dst_ws.cell(row=row, column=col, value=src_cell.value)
        dst_cell.number_format = src_cell.number_format
        if src_cell.has_style:
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)

for merged_range in src_ws.merged_cells.ranges:
    if merged_range.max_row <= last_row:
        dst_ws.merge_cells(str(merged_range))

for col_letter, dim in src_ws.column_dimensions.items():
    dst_ws.column_dimensions[col_letter].width = dim.width

dst_wb.save(DST)
print(f"Saved {DST}: {last_row} rows written ({HEADER_ROWS} header + {DEMO_VENDOR_ROWS} vendor rows)")
