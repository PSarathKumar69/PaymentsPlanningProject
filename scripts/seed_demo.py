"""Dev-only utility: build a small demo Excel workbook — the SAME full
column layout as the real master sheet (all 48 columns: ERP Code, Entity,
Unique, Vendor Name, VN, Op. Balance, every monthly Payable/Payment column,
Total Payable/Payment, Closing Balance, W1-W5, Total, Priority, Commitment
Months, Assigned Week), just with only 10 real vendors' rows — then runs the
REAL ingestion pipeline (backend/ingestion/load_excel.py) against it into a
separate demo DB. This exercises the actual Excel -> DB path, not a
hand-copied DB shortcut, and never touches the real master sheet or the
real dev DB.

Usage:
    python -m scripts.seed_demo

Produces:
    data/demo/Vendor's Details (Demo10).xlsx
    backend/db/demo.db  (delete + re-run any time to reset to a clean demo state)
"""
import os

# Must happen before any backend.db.session import (this repo's own or a
# transitive one via backend.ingestion.load_excel) — the engine binds to
# PAYMENTS_DB_PATH at import time.
os.environ.setdefault("PAYMENTS_DB_PATH", "backend/db/demo.db")

import openpyxl

from backend.configuration.vendor_edits import update_vendor_field
from backend.db.models import Vendor
from backend.db.session import SessionLocal, init_db
from backend.ingestion.column_mapping import DATA_START_ROW, HEADER_ROW
from backend.ingestion.load_excel import load, print_report
from backend.shared.enums import ChangeSource

SOURCE_EXCEL = "data/Vendor's Details.xlsx"
DEMO_EXCEL = "data/demo/Vendor's Details (Demo10).xlsx"

# Picked for demo breadth, not arbitrary: a payment-status mix (paid in
# full / partial / not paid), all 5 assigned weeks represented, an aging
# bucket spread (31-60 through 120+ — the real sheet currently has no
# vendor whose OWN OLDEST bucket is 0-30, so that end isn't representable
# without fabricating data), and a balance spread (~6.5L to ~7.2Cr) so
# Model 2's shortfall behavior is visible.
DEMO_ERP_CODES = [
    "V00400", "V01542", "V01524", "V01227", "V00370",
    "V00445", "V01222", "V01054", "V01083", "V01531",
]

# V00400 has no Assigned Week in the real sheet right now (checked both the
# legacy "Priority" column and the dedicated "Assigned Week" column — both
# blank), so a fresh ingest leaves it out of every weekly view/override.
# Assigned here the same way Finance would in the real Configuration module
# — a real edit, written back + audited — purely so this demo vendor
# exercises the full feature set. Confirm the week number with Sarath if it
# matters for the walkthrough; any of 1-5 works equally well here.
ASSIGNED_WEEK_FIXUPS = {"V00400": 3}


def build_demo_excel(source_path=SOURCE_EXCEL, output_path=DEMO_EXCEL, erp_codes=DEMO_ERP_CODES):
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    src_ws = src_wb.active

    erp_col = next(
        c for c in range(1, src_ws.max_column + 1)
        if str(src_ws.cell(row=HEADER_ROW, column=c).value or "").strip() == "ERP Code"
    )

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    # Header block (rows 1-3), every column, verbatim — including the
    # merged "Payable"/"Payment" block-label cells.
    for r in range(1, DATA_START_ROW):
        for c in range(1, src_ws.max_column + 1):
            out_ws.cell(row=r, column=c, value=src_ws.cell(row=r, column=c).value)
    for merged_range in src_ws.merged_cells.ranges:
        out_ws.merge_cells(str(merged_range))

    out_row = DATA_START_ROW
    found = set()
    for r in range(DATA_START_ROW, src_ws.max_row + 1):
        erp_code = src_ws.cell(row=r, column=erp_col).value
        if erp_code in erp_codes:
            for c in range(1, src_ws.max_column + 1):
                out_ws.cell(row=out_row, column=c, value=src_ws.cell(row=r, column=c).value)
            out_row += 1
            found.add(erp_code)

    missing = set(erp_codes) - found
    if missing:
        raise ValueError(f"ERP codes not found in {source_path}: {missing}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    out_wb.save(output_path)
    print(f"Wrote {out_row - DATA_START_ROW} vendor rows ({src_ws.max_column} columns) to {output_path}")
    return output_path


def seed_demo_db(excel_path):
    # Drop + recreate tables through the engine rather than deleting the
    # .db file directly — an OS-level delete fails with a sharing violation
    # on Windows if any process (even just a server left running against
    # this demo DB) still has the file open.
    from backend.db.models import Base
    from backend.db.session import engine

    Base.metadata.drop_all(engine)
    init_db()
    session = SessionLocal()
    try:
        report = load(excel_path=excel_path, session=session)
        session.commit()
    finally:
        session.close()
    print_report(report)

    # Real Configuration-module edit, written back into the demo Excel and
    # audited — same code path Finance's own edits use (update_vendor_field
    # owns its own session here, so this commits and publishes independently
    # of the ingestion transaction above).
    session = SessionLocal()
    try:
        for erp_code, week in ASSIGNED_WEEK_FIXUPS.items():
            vendor = session.query(Vendor).filter_by(erp_code=erp_code).one()
            update_vendor_field(
                vendor.id, "assigned_week", week, source=ChangeSource.UI_EDIT, session=None, excel_path=excel_path
            )
            print(f"Assigned week {week} to {erp_code} (real sheet had no week tag for it)")
    finally:
        session.close()


if __name__ == "__main__":
    seed_demo_db(build_demo_excel())
